#!/usr/bin/env python3
"""Seed `pricings` + `pricing_lines` (bảng giá) from customer tariff
documents.

Each tariff layout is wildly different across customers (see
`docs/BANG_GIA_DATA_FLOW.md`). Rather than build a brittle generic
parser, we ship one named extractor per known layout and dispatch by
`--format` flag (or by filename hint if `--format auto`).

Built-in formats:

  pan      — PAN tariff. Sheet `Trucking (HD)`, header row 6, AK/AL/AM
             columns under the "ĐƠN GIÁ T04.26 ĐÃ CỘNG GIÁ GỬI" super-
             header carry (Vỏ, Hàng 20', Hàng 40&45) prices per route.

  hap      — HAP tariff. Sheet `CUOC`, header rows 1-3, cols C/D
             (Hàng 20'/40') and E/F (Vỏ 20'/40') per route.

  newway   — NEWWAY settlement-style data, no clean tariff sheet.
             Best-effort: extracts distinct (route × work_type) →
             modal unit_price. Often produces 0 rows; flagged clearly.

Idempotent on (client, work_type, pickup_id, dropoff_id).

Driver_salary + allowance default to 0 — kế toán fills via
PricingDetail UI per the team's pricing-data-flow decision.

Examples:
    # PAN tariff
    ./scripts/seeds/seed_pricing_from_files.py \\
        --format pan --client-code PAN \\
        --files "docs/PAN- BK SL T04.26 (HD).xlsx"

    # HAP tariff
    ./scripts/seeds/seed_pricing_from_files.py \\
        --format hap --client-code HAP \\
        --files "docs/Phúc Lộc - Shipside T4.26 HAP.xlsx"

    # batch — multiple customers, --format auto picks per file by name
    ./scripts/seeds/seed_pricing_from_files.py --format auto \\
        --client-code PAN --files "docs/PAN- BK SL T04.26 (HD).xlsx"

    # dry-run + droplet
    DATABASE_URL=postgresql://prod-host/db \\
        ./scripts/seeds/seed_pricing_from_files.py --prod --dry-run \\
        --format hap --client-code HAP \\
        --files /tmp/HAP_tariff.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import sys
import time
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from _common import (
    add_common_args,
    add_files_arg,
    assert_safe_target,
    configure_logging,
    filter_existing_files,
    open_session,
)


@dataclass
class TariffRow:
    client_code: str
    pickup_raw: str
    dropoff_raw: str
    work_type: str          # F20 | F40 | E20 | E40
    unit_price: int
    note: str = ""


# ---------------------------------------------------------------------------
# Format detectors / extractors
# ---------------------------------------------------------------------------

FORMATS = ("pan", "hap", "newway", "auto")


def detect_format(filename: str) -> str:
    """Heuristic — `--format auto` falls through here."""
    n = filename.lower()
    if "pan" in n and ("bk" in n or "sl" in n):
        return "pan"
    if "hap" in n or "shipside" in n:
        return "hap"
    if "newway" in n or "hechun" in n:
        return "newway"
    return ""


def parse_pan(filepath: Path, client_code: str) -> list[TariffRow]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "Trucking (HD)" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Trucking (HD)"]
    rows: list[TariffRow] = []
    for r in range(7, ws.max_row + 1):
        route = ws.cell(r, 2).value
        if not route or not isinstance(route, str):
            continue
        route = route.strip()
        if any(t in route.upper() for t in ("TỔNG", "TOTAL", "GHI CHU", "CỘNG")):
            continue
        empty_p = ws.cell(r, 36).value
        full20 = ws.cell(r, 37).value
        full40 = ws.cell(r, 38).value
        pickup, dropoff = _split_route(route)
        for price, wt in (
            (empty_p, "E40"), (empty_p, "E20"),
            (full20, "F20"), (full40, "F40"),
        ):
            if isinstance(price, (int, float)) and price > 0:
                rows.append(TariffRow(
                    client_code=client_code,
                    pickup_raw=pickup, dropoff_raw=dropoff,
                    work_type=wt, unit_price=int(round(price)),
                    note=f"Trucking (HD) row {r}",
                ))
    wb.close()
    return rows


def parse_hap(filepath: Path, client_code: str) -> list[TariffRow]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "CUOC" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["CUOC"]
    rows: list[TariffRow] = []
    for r in range(4, ws.max_row + 1):
        route = ws.cell(r, 2).value
        if not route or not isinstance(route, str):
            continue
        route = route.strip()
        if any(t in route.upper() for t in ("TỔNG", "TOTAL", "CỘNG")):
            continue
        full20 = ws.cell(r, 3).value
        full40 = ws.cell(r, 4).value
        empty20 = ws.cell(r, 5).value
        empty40 = ws.cell(r, 6).value
        pickup, dropoff = _split_route(route)
        for price, wt in (
            (full20, "F20"), (full40, "F40"),
            (empty20, "E20"), (empty40, "E40"),
        ):
            if isinstance(price, (int, float)) and price > 0:
                rows.append(TariffRow(
                    client_code=client_code,
                    pickup_raw=pickup, dropoff_raw=dropoff,
                    work_type=wt, unit_price=int(round(price)),
                    note=f"CUOC row {r}",
                ))
    wb.close()
    return rows


def parse_newway(filepath: Path, client_code: str) -> list[TariffRow]:
    """Settlement-data shape — modal unit_price per (route, work_type).
    Output is best-effort. Routes without a parseable pickup/dropoff dash
    fall through to the upserter which will skip them."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = "Tháng 4" if "Tháng 4" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_seen: dict[tuple[str, str], dict[int, int]] = {}
    for r in range(8, ws.max_row + 1):
        route = ws.cell(r, 11).value
        if not route or not isinstance(route, str):
            continue
        route = route.strip()
        price = ws.cell(r, 12).value
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        empty20 = ws.cell(r, 5).value
        empty40 = ws.cell(r, 6).value
        full20 = ws.cell(r, 7).value
        full40 = ws.cell(r, 8).value
        wt = ("F40" if full40 else "F20" if full20
              else "E40" if empty40 else "E20" if empty20 else None)
        if wt is None:
            continue
        rows_seen.setdefault((route, wt), {})[int(round(price))] = (
            rows_seen.get((route, wt), {}).get(int(round(price)), 0) + 1
        )
    wb.close()
    out: list[TariffRow] = []
    for (route, wt), counts in rows_seen.items():
        best_price = max(counts.items(), key=lambda kv: kv[1])[0]
        pickup, dropoff = _split_route(route)
        out.append(TariffRow(
            client_code=client_code,
            pickup_raw=pickup, dropoff_raw=dropoff,
            work_type=wt, unit_price=best_price,
            note=f"settlement modal across {sum(counts.values())} trips",
        ))
    return out


PARSERS = {"pan": parse_pan, "hap": parse_hap, "newway": parse_newway}


def _split_route(route: str) -> tuple[str, str]:
    for sep in (" – ", " — ", " - ", "–", "—", "-"):
        if sep in route:
            parts = [p.strip() for p in route.split(sep, 1)]
            if len(parts) == 2 and all(parts):
                return parts[0], parts[1]
    return route.strip(), ""


# ---------------------------------------------------------------------------
# Upsert
# ---------------------------------------------------------------------------

async def _upsert(db, rows: list[TariffRow], dry_run: bool, log) -> dict[str, int]:
    from sqlalchemy import select  # type: ignore
    from app.models.domain import Client, Pricing, PricingLine  # type: ignore
    from app.services.location_resolver import (  # type: ignore
        LocationResolverService, ResolverSource,
    )

    resolver = LocationResolverService(db)

    # Clients lookup
    client_map: dict[str, Client] = {}
    for c in (await db.execute(select(Client))).scalars().all():
        if c.code:
            client_map[c.code.upper()] = c

    pricings_added = pricings_existing = 0
    lines_added = lines_existing = 0
    skipped_no_client = skipped_no_locations = 0

    for row in rows:
        client = client_map.get(row.client_code.upper())
        if client is None:
            skipped_no_client += 1
            continue

        pickup_loc = dropoff_loc = None
        if row.pickup_raw:
            r = await resolver.resolve_or_create(
                row.pickup_raw, source=ResolverSource.MANUAL, user_id=None,
            )
            pickup_loc = r.location
        if row.dropoff_raw:
            r = await resolver.resolve_or_create(
                row.dropoff_raw, source=ResolverSource.MANUAL, user_id=None,
            )
            dropoff_loc = r.location
        if pickup_loc is None or dropoff_loc is None:
            skipped_no_locations += 1
            continue

        existing = (await db.execute(
            select(Pricing).where(
                Pricing.client_id == client.id,
                Pricing.work_type == row.work_type,
                Pricing.pickup_location_id == pickup_loc.id,
                Pricing.dropoff_location_id == dropoff_loc.id,
            )
        )).scalar_one_or_none()
        if existing is None:
            if dry_run:
                pricings_added += 1
                lines_added += 1
                log.info("  [dry-run] +Pricing %s %s %s→%s @%d",
                         row.client_code, row.work_type,
                         pickup_loc.name, dropoff_loc.name, row.unit_price)
                continue
            pricing = Pricing(
                client_id=client.id,
                work_type=row.work_type,
                pickup_location_id=pickup_loc.id,
                dropoff_location_id=dropoff_loc.id,
                is_active=True,
            )
            db.add(pricing); await db.flush()
            pricings_added += 1
        else:
            pricing = existing
            pricings_existing += 1

        line_existing = (await db.execute(
            select(PricingLine).where(
                PricingLine.pricing_id == pricing.id, PricingLine.quantity == 1,
            )
        )).scalar_one_or_none()
        if line_existing is None:
            if dry_run:
                lines_added += 1
                continue
            db.add(PricingLine(
                pricing_id=pricing.id, quantity=1,
                unit_price=row.unit_price,
                driver_salary=0, allowance=0,
            ))
            lines_added += 1
        else:
            lines_existing += 1

    if not dry_run:
        await db.commit()

    return {
        "pricings_added": pricings_added,
        "pricings_existing": pricings_existing,
        "lines_added": lines_added,
        "lines_existing": lines_existing,
        "skipped_no_client": skipped_no_client,
        "skipped_no_locations": skipped_no_locations,
    }


# ---------------------------------------------------------------------------
# Argparse
# ---------------------------------------------------------------------------

def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="seed_pricing_from_files",
        description="Seed bảng giá (Pricing + PricingLine) from customer tariff "
                    "documents.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "--format",
        choices=FORMATS,
        required=True,
        help="Tariff layout. Use 'auto' to detect from filename (pan/hap/newway).",
    )
    p.add_argument(
        "--client-code",
        required=True,
        help="Client.code that owns these tariffs (e.g. PAN, HAP, NEWWAY).",
    )
    add_files_arg(p, required=True)
    add_common_args(p)
    return p.parse_args(argv)


async def run(args: argparse.Namespace) -> int:
    log = configure_logging(args.log_level)
    db_url = assert_safe_target(args, log)
    files = filter_existing_files(args.files, log,
                                   allow_missing=args.allow_missing_files)
    if not files:
        log.error("No files to process.")
        return 2

    started = time.monotonic()
    grand_added = grand_existing = 0

    async with open_session(db_url) as db:
        for fp in files:
            fmt = args.format
            if fmt == "auto":
                fmt = detect_format(fp.name)
                if fmt not in PARSERS:
                    log.warning("[skip] %s — couldn't auto-detect format", fp.name)
                    continue
            parser = PARSERS[fmt]
            log.info("=== %s — format=%s, client=%s ===", fp.name, fmt, args.client_code)
            rows = parser(fp, args.client_code)
            log.info("  parsed %d tariff rows", len(rows))
            if not rows:
                log.warning("  ⚠ NO ROWS extracted — confirm file structure")
                continue
            stats = await _upsert(db, rows, args.dry_run, log)
            log.info("  pricings: +%d, existing %d", stats["pricings_added"], stats["pricings_existing"])
            log.info("  lines:    +%d, existing %d", stats["lines_added"], stats["lines_existing"])
            if stats["skipped_no_client"]:
                log.warning("  ⚠ %d rows skipped — client %s not found",
                            stats["skipped_no_client"], args.client_code)
            if stats["skipped_no_locations"]:
                log.warning("  ⚠ %d rows skipped — pickup/dropoff couldn't be resolved",
                            stats["skipped_no_locations"])
            grand_added += stats["lines_added"]
            grand_existing += stats["lines_existing"]

    log.info("Done in %.2fs — total lines: +%d, existing %d",
             time.monotonic() - started, grand_added, grand_existing)
    log.info("Note: driver_salary + allowance default to 0 — fill via PricingDetail UI.")
    return 0


def main() -> int:
    return asyncio.run(run(parse_args(sys.argv[1:])))


if __name__ == "__main__":
    sys.exit(main())
