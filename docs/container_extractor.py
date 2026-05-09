#!/usr/bin/env python3
"""
Container Extraction Tool
========================
Extracts container data from shipping Excel files and outputs CSV with:
- Container Number
- Cont Type (E20, E40, F20, F40)
- Pick Up Point
- Drop Off Point

Supports 4 file formats:
1. GLORY SHANGHAI style (Bay Plan + System Export)
2. CONSCIENCE style (Bay Plan + System Export)
3. HAIAN BETA style (Loading List)
4. Phúc Lộc Shipside style (Invoice)

Usage:
    python container_extractor.py <input_file_or_directory> [--output output.csv]
"""

import argparse
import csv
import os
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd


def detect_file_type(filepath):
    """Detect which file format based on filename and sheet names."""
    fname = os.path.basename(filepath).upper()
    
    try:
        if filepath.endswith('.xls'):
            xl = pd.ExcelFile(filepath)
            sheets = [s.upper() for s in xl.sheet_names]
        else:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sheets = [s.upper() for s in wb.sheetnames]
            wb.close()
    except Exception:
        return 'UNKNOWN'
    
    # Check for HAIAN BETA style
    if 'HAIAN' in fname or 'TOTAL' in sheets:
        return 'HAIAN_BETA'
    
    # Check for Phúc Lộc style  
    if 'CUOC' in sheets or 'BẢNG KÊ SS' in sheets or 'BANG KE SS' in sheets:
        return 'PHUC_LOC'
    
    # Check for GLORY/CONSCIENCE style (Bay Plan + System Export)
    # These have Bay Plan sheets with vessel name as sheet name
    if filepath.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            for sheet_name in wb.sheetnames:
                # Bay Plan sheet typically named after vessel
                if any(kw in sheet_name.upper() for kw in ['CONSCIENCE', 'GLORY', 'N', 'S']):
                    if len(wb.sheetnames) >= 2:
                        wb.close()
                        return 'BAY_PLAN_SYSTEM_EXPORT'
            wb.close()
        except Exception:
            pass
    
    return 'UNKNOWN'


def extract_bay_plan_system_export(filepath):
    """
    Extract from GLORY/CONSCIENCE style files.
    Uses Bay Plan sheet (has port sections) for location data.
    Falls back to System Export for F/E status if Bay Plan lacks it.
    """
    results = []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # --- Identify Bay Plan vs System Export sheets ---
    bay_plan_sheet = None
    system_export_sheet = None
    
    for sheet_name in wb.sheetnames:
        sn = sheet_name.upper().strip()
        # System Export sheets are named "Sheet", "Sheet1", etc.
        if sn in ('SHEET', 'SHEET1') or (sn.startswith('SHEET') and sn[5:].isdigit()):
            system_export_sheet = wb[sheet_name]
        else:
            bay_plan_sheet = wb[sheet_name]
    
    # Get vessel name from first row of Bay Plan, or from sheet name
    vessel_name = ''
    if bay_plan_sheet:
        vessel_name = str(bay_plan_sheet.cell(row=1, column=1).value or '').strip()
    
    # Also try to extract vessel name from the Bay Plan sheet name itself
    bay_plan_sheet_name = ''
    for sheet_name in wb.sheetnames:
        sn = sheet_name.upper().strip()
        if sn not in ('SHEET', 'SHEET1') and not (sn.startswith('SHEET') and sn[5:].isdigit()):
            bay_plan_sheet_name = sheet_name
            break
    
    # If vessel name not found in cell A1, use the sheet name
    if not vessel_name and bay_plan_sheet_name:
        vessel_name = bay_plan_sheet_name
    
    # --- Build F/E lookup from System Export ---
    fe_lookup = {}
    se_vessel_name = ''
    se_voyage = ''
    if system_export_sheet:
        # Find header row with container number column
        header_row_idx = None
        cont_col_idx = None
        fe_col_idx = None
        vessel_col_idx = None
        voyage_col_idx = None
        
        for row_idx, row in enumerate(system_export_sheet.iter_rows(min_row=1, max_row=1, values_only=False), 1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if 'Container' in val or 'หมายเลขตู้' in val:
                        cont_col_idx = cell.column
                        header_row_idx = row_idx
                    elif val in ('空/重(E/F)', 'Lỗ/Hàng', 'Rỗng / Hàng') or 'E/F' in val or '空/重' in val:
                        fe_col_idx = cell.column
                    elif 'Tên tàu' in val or 'tên tàu' in val:
                        vessel_col_idx = cell.column
                    elif 'Chuyến' in val or 'chuyến tàu' in val:
                        voyage_col_idx = cell.column
        
        # Also try to find F/E by scanning known column positions
        # GLORY: Container=col C(3), F/E=col AC(29)
        # CONSCIENCE: Container=col B(2), F/E=col BH(60)
        if cont_col_idx is None:
            # Try column C and B
            for test_col in [3, 2]:
                val = system_export_sheet.cell(row=2, column=test_col).value
                if val and isinstance(val, str) and len(val.strip()) >= 7 and re.match(r'^[A-Z]{4}\d{7}$', val.strip()):
                    cont_col_idx = test_col
                    header_row_idx = 1
                    break
        
        if fe_col_idx is None:
            # Try known F/E column positions
            for test_col in [29, 60, 28, 59]:  # AC, BH, AB, BG
                val = system_export_sheet.cell(row=2, column=test_col).value
                if val and str(val).strip().upper() in ('E', 'F', 'H', 'R'):
                    fe_col_idx = test_col
                    break
        
        if cont_col_idx and fe_col_idx and header_row_idx:
            for row in system_export_sheet.iter_rows(min_row=header_row_idx+1, values_only=False):
                cont_val = None
                fe_val = None
                for cell in row:
                    if cell.column == cont_col_idx:
                        if cell.value and isinstance(cell.value, str):
                            v = cell.value.strip()
                            if re.match(r'^[A-Z]{4}\d{7}$', v):
                                cont_val = v
                    elif cell.column == fe_col_idx:
                        if cell.value and str(cell.value).strip().upper() in ('E', 'F', 'H', 'R'):
                            fe_val = str(cell.value).strip().upper()
                    elif vessel_col_idx and cell.column == vessel_col_idx:
                        if cell.value and not se_vessel_name:
                            se_vessel_name = str(cell.value).strip()
                    elif voyage_col_idx and cell.column == voyage_col_idx:
                        if cell.value and not se_voyage:
                            se_voyage = str(cell.value).strip()
                
                if cont_val and fe_val:
                    fe_lookup[cont_val] = fe_val
    
    # Use vessel name from System Export only if Bay Plan doesn't have one
    # Don't append SE voyage if Bay Plan already has a complete vessel+voyage name
    if not vessel_name and se_vessel_name:
        vessel_name = se_vessel_name
        if se_voyage and se_voyage not in vessel_name:
            vessel_name = f"{vessel_name} {se_voyage}".strip()
    # If Bay Plan has vessel name, don't override with SE data
    
    # --- Extract from Bay Plan sheet ---
    if bay_plan_sheet:
        # Step 1: Find port section headers in row 2
        port_col_starts = {}  # {column_number: port_code}
        
        for cell in bay_plan_sheet[2]:
            if cell.value and isinstance(cell.value, str):
                val = cell.value.strip()
                # Port codes are short alphabetic strings like HKG, SHA, TAG, NSA
                if len(val) <= 5 and val.isalpha():
                    port_col_starts[cell.column] = val
        
        # Step 2: Find header row (contains "Container" or "Số Container")
        header_row = None
        for r in range(3, 8):
            for cell in bay_plan_sheet[r]:
                if cell.value and isinstance(cell.value, str) and 'Container' in cell.value:
                    header_row = r
                    break
            if header_row:
                break
        
        if header_row is None:
            print(f"  Warning: Could not find header row in Bay Plan sheet")
            wb.close()
            return results
        
        # Step 3: Map each "Số Container" column to its port section + size/type columns
        sections = []
        for cell in bay_plan_sheet[header_row]:
            if cell.value and isinstance(cell.value, str) and 'Container' in cell.value:
                cont_col = cell.column
                
                # Find the port for this section (nearest port header to the left)
                assigned_port = None
                for pc in sorted(port_col_starts.keys()):
                    if pc <= cont_col:
                        assigned_port = port_col_starts[pc]
                
                # Find Size and Type columns in the same section (within +/- 5 cols)
                size_col = None
                type_col = None
                for cell2 in bay_plan_sheet[header_row]:
                    if abs(cell2.column - cont_col) <= 5 and cell2.column >= cont_col - 2:
                        if cell2.value and isinstance(cell2.value, str):
                            if 'Kích thước' in cell2.value or 'Size' in cell2.value:
                                size_col = cell2.column
                            elif 'Loại' in cell2.value and 'Container' not in cell2.value:
                                type_col = cell2.column
                
                sections.append({
                    'cont_col': cont_col,
                    'size_col': size_col,
                    'type_col': type_col,
                    'port': assigned_port
                })
        
        # Step 4: Extract container data from each section
        for section in sections:
            cont_col = section['cont_col']
            size_col = section['size_col']
            type_col = section['type_col']
            port = section['port']
            
            for row in bay_plan_sheet.iter_rows(min_row=header_row+1, values_only=False):
                cont_val = None
                size_val = None
                type_val = None
                
                for cell in row:
                    if cell.column == cont_col:
                        cont_val = cell.value
                    elif cell.column == size_col:
                        size_val = cell.value
                    elif cell.column == type_col:
                        type_val = cell.value
                
                # Validate container number: must be like ABCU1234567 (4 letters + 7 digits)
                if not cont_val or not isinstance(cont_val, str):
                    continue
                
                cont_num = cont_val.strip()
                if not re.match(r'^[A-Z]{4}\d{7}$', cont_num):
                    continue
                
                # Determine F/E status
                fe = fe_lookup.get(cont_num, 'E')  # Default E for Bay Plan exports
                
                # Build cont type
                cont_type = build_cont_type(fe, size_val, type_val)
                
                # Pickup = HAIPHONG (POL for all these vessels)
                pickup = 'HAIPHONG'
                dropoff = port or ''
                
                results.append({
                    'container_number': cont_num,
                    'cont_type': cont_type,
                    'pickup': pickup,
                    'dropoff': dropoff,
                    'vessel': vessel_name,
                    'source': os.path.basename(filepath)
                })
    
    wb.close()
    return results


def extract_haian_beta(filepath):
    """
    Extract from HAIAN BETA style Loading List.
    Has: POD, ContainerNo., F/E, SIZE columns.
    PORT OF LOADING from header.
    """
    results = []
    
    df = pd.read_excel(filepath, sheet_name='TOTAL', header=None)
    
    # Find header row (row with 'CONTAINERNo.')
    header_row = None
    for i in range(len(df)):
        for j in range(len(df.columns)):
            val = df.iloc[i, j]
            if isinstance(val, str) and 'CONTAINERNo' in val:
                header_row = i
                break
        if header_row is not None:
            break
    
    if header_row is None:
        print(f"Warning: Could not find header row in {filepath}")
        return results
    
    # Set column names from header row
    data = df.iloc[header_row+1:].copy()
    data.columns = df.iloc[header_row]
    data = data[data['CONTAINERNo.'].notna()].reset_index(drop=True)
    
    # Get POL from header
    pol = 'HAIPHONG'
    for i in range(min(10, len(df))):
        for j in range(len(df.columns)):
            val = df.iloc[i, j]
            if isinstance(val, str) and 'PORT OF LOADING' in val.upper():
                # Extract port name
                match = re.search(r'PORT OF LOADING:\s*(.+)', val, re.IGNORECASE)
                if match:
                    pol = match.group(1).strip()
                break
    
    # Get vessel name - scan for VESSEL: keyword then find nearest non-empty cell to the right
    vessel_name = ''
    voyage = ''
    for i in range(min(10, len(df))):
        for j in range(len(df.columns)):
            val = df.iloc[i, j]
            if isinstance(val, str):
                val_upper = val.upper().strip()
                if val_upper.startswith('VESSEL') and ':' in val_upper and 'OPR' not in val_upper:
                    # Find the nearest non-empty cell to the right (may skip NaN cells)
                    for k in range(j+1, min(j+5, len(df.columns))):
                        cell_val = df.iloc[i, k]
                        if pd.notna(cell_val) and str(cell_val).strip():
                            vessel_name = str(cell_val).strip()
                            break
                    # Also check if vessel name is embedded in the same cell
                    if not vessel_name:
                        match = re.search(r'VESSEL:\s*(.+)', val, re.IGNORECASE)
                        if match:
                            vessel_name = match.group(1).strip()
                elif val_upper.startswith('VOY') and ':' in val_upper:
                    for k in range(j+1, min(j+5, len(df.columns))):
                        cell_val = df.iloc[i, k]
                        if pd.notna(cell_val) and str(cell_val).strip():
                            voyage = str(cell_val).strip()
                            break
    
    if voyage:
        vessel_name = f"{vessel_name} {voyage}".strip()
    
    # Extract containers
    for _, row in data.iterrows():
        cont_num = str(row.get('CONTAINERNo.', '')).strip()
        if not cont_num or cont_num == 'nan' or len(cont_num) < 7:
            continue
        
        # Validate container number format (4 letters + 7 digits)
        if not re.match(r'^[A-Z]{4}\d{7}$', cont_num):
            continue
        
        pod = str(row.get('POD', '')).strip()
        if pod == 'nan':
            pod = ''
        
        fe = str(row.get('F/E', '')).strip().upper()
        if fe == 'NAN':
            fe = 'E'
        
        size_type = str(row.get('SIZE', '')).strip()
        if size_type == 'nan':
            size_type = ''
        
        # Parse combined size+type like "40HC", "20DC"
        size, ctype = parse_size_type(size_type)
        
        cont_type = build_cont_type(fe, size, ctype)
        
        results.append({
            'container_number': cont_num,
            'cont_type': cont_type,
            'pickup': pol,
            'dropoff': pod,
            'vessel': vessel_name,
            'source': os.path.basename(filepath)
        })
    
    return results


def extract_phuc_loc(filepath):
    """
    Extract from Phúc Lộc Shipside Invoice.
    Has: SỐCONT, LOẠI, H/R, NƠI LẤY CONTAINER, NƠI TRẢ CONTAINER
    """
    results = []
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Find the main data sheet
    ws = None
    for sheet_name in wb.sheetnames:
        if 'BẢNG KÊ' in sheet_name or 'BANG KE' in sheet_name:
            ws = wb[sheet_name]
            break
    
    if ws is None:
        # Try first sheet
        ws = wb[wb.sheetnames[0]]
    
    # Find header row (contains SỐCONT)
    header_row = None
    for r in range(1, min(30, ws.max_row+1)):
        for cell in ws[r]:
            if cell.value and isinstance(cell.value, str) and 'SỐCONT' in cell.value:
                header_row = r
                break
        if header_row:
            break
    
    if header_row is None:
        print(f"Warning: Could not find header row in {filepath}")
        wb.close()
        return results
    
    # Map column positions from header
    col_map = {}
    for cell in ws[header_row]:
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip()
            if 'SỐCONT' in val or 'SOCONT' in val:
                col_map['container'] = cell.column
            elif val == 'LOẠI' or val == 'LOAI':
                col_map['size'] = cell.column
            elif 'H/R' in val:
                col_map['hr'] = cell.column
            elif 'TÀU' in val or 'TAU' in val:
                col_map['vessel'] = cell.column
            elif 'NƠI LẤY' in val or 'NOI LAY' in val:
                col_map['pickup'] = cell.column
            elif 'NƠI TRẢ' in val or 'NOI TRA' in val:
                col_map['dropoff'] = cell.column
            elif 'CHUYẾN' in val or 'CHUYEN' in val:
                col_map['voyage'] = cell.column
    
    # Extract data
    for r in range(header_row + 1, ws.max_row + 1):
        cont_num = ws.cell(row=r, column=col_map.get('container', 2)).value
        if not cont_num:
            break
        
        cont_num = str(cont_num).strip()
        if len(cont_num) < 7:
            continue
        
        # Validate container number format (4 letters + 7 digits)
        if not re.match(r'^[A-Z]{4}\d{7}$', cont_num):
            continue
        
        size = ws.cell(row=r, column=col_map.get('size', 3)).value
        hr = ws.cell(row=r, column=col_map.get('hr', 4)).value
        vessel = ws.cell(row=r, column=col_map.get('vessel', 5)).value
        pickup = ws.cell(row=r, column=col_map.get('pickup', 7)).value
        dropoff = ws.cell(row=r, column=col_map.get('dropoff', 8)).value
        
        # Map H/R to F/E
        if hr:
            hr = str(hr).strip().upper()
            fe = 'F' if hr == 'H' else 'E' if hr == 'R' else hr
        else:
            fe = 'E'
        
        cont_type = build_cont_type(fe, size)
        
        results.append({
            'container_number': cont_num,
            'cont_type': cont_type,
            'pickup': str(pickup).strip() if pickup else '',
            'dropoff': str(dropoff).strip() if dropoff else '',
            'vessel': str(vessel).strip() if vessel else '',
            'source': os.path.basename(filepath)
        })
    
    wb.close()
    return results


def parse_size_type(size_type_str):
    """Parse combined size+type string like '40HC', '20DC', '20RF'."""
    if not size_type_str or str(size_type_str) == 'nan':
        return None, None
    
    s = str(size_type_str).strip()
    match = re.match(r'^(\d{2})\s*(.*)$', s)
    if match:
        return match.group(1), match.group(2)
    
    # Try just digits
    if s.isdigit():
        return s, None
    
    return None, s


def build_cont_type(fe, size, ctype=None):
    """
    Build container type code: E20, E40, F20, F40
    
    Args:
        fe: 'F' (full/hàng) or 'E' (empty/rỗng/vỏ)
        size: Container size as string ('20', '40', '45')
        ctype: Container type code ('HC', 'DC', 'GP', 'RF', 'TK')
    
    Returns:
        Simplified type: E20, E40, F20, F40
        For 45' containers: E45, F45
    """
    # Normalize F/E
    if fe and str(fe).strip().upper() in ('H', 'RỖNG', 'RO'):
        fe = 'E'
    elif fe and str(fe).strip().upper() in ('HÀNG', 'HANG'):
        fe = 'F'
    
    fe = str(fe).strip().upper() if fe else 'E'
    if fe not in ('E', 'F'):
        fe = 'E'  # default to empty
    
    # Normalize size
    size = str(size).strip() if size else '20'
    # Extract leading digits
    size_match = re.match(r'^(\d+)', size)
    if size_match:
        size = size_match.group(1)
    else:
        size = '20'  # default
    
    return f"{fe}{size}"


def extract_file(filepath):
    """Auto-detect file type and extract containers."""
    file_type = detect_file_type(filepath)
    print(f"  Detected type: {file_type}")
    
    if file_type == 'HAIAN_BETA':
        return extract_haian_beta(filepath)
    elif file_type == 'PHUC_LOC':
        return extract_phuc_loc(filepath)
    elif file_type == 'BAY_PLAN_SYSTEM_EXPORT':
        return extract_bay_plan_system_export(filepath)
    else:
        print(f"  Warning: Unknown file type for {filepath}, attempting all extractors...")
        # Try each extractor
        for extractor in [extract_haian_beta, extract_phuc_loc, extract_bay_plan_system_export]:
            try:
                results = extractor(filepath)
                if results:
                    return results
            except Exception as e:
                continue
        return []


def main():
    parser = argparse.ArgumentParser(description='Extract container data from shipping Excel files')
    parser.add_argument('input', help='Input Excel file or directory containing Excel files')
    parser.add_argument('--output', '-o', help='Output CSV file path', default=None)
    
    args = parser.parse_args()
    
    input_path = args.input
    
    # Collect all Excel files
    if os.path.isdir(input_path):
        files = []
        for f in os.listdir(input_path):
            if f.endswith(('.xlsx', '.xls')) and not f.startswith('~'):
                files.append(os.path.join(input_path, f))
    else:
        files = [input_path]
    
    if not files:
        print("No Excel files found!")
        sys.exit(1)
    
    print(f"Found {len(files)} file(s) to process")
    
    all_results = []
    for filepath in files:
        print(f"\nProcessing: {os.path.basename(filepath)}")
        try:
            results = extract_file(filepath)
            print(f"  Extracted {len(results)} containers")
            all_results.extend(results)
        except Exception as e:
            print(f"  Error: {e}")
            import traceback
            traceback.print_exc()
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        output_path = os.path.join(os.path.dirname(files[0]), 'containers_extracted.csv')
    
    # Write CSV
    if all_results:
        fieldnames = ['container_number', 'cont_type', 'pickup', 'dropoff', 'vessel', 'source']
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_results)
        
        print(f"\n✅ Extracted {len(all_results)} total containers → {output_path}")
        
        # Print summary
        from collections import Counter
        type_counts = Counter(r['cont_type'] for r in all_results)
        print("\nSummary by container type:")
        for ct, count in sorted(type_counts.items()):
            print(f"  {ct}: {count}")
        
        source_counts = Counter(r['source'] for r in all_results)
        print("\nSummary by source file:")
        for src, count in sorted(source_counts.items()):
            print(f"  {src}: {count}")
    else:
        print("\n❌ No containers extracted!")


if __name__ == '__main__':
    main()
