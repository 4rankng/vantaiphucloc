"""Tests for the customer settlement export pipeline.

Covers:
- Vietnamese number-to-words helper (a few representative bills)
- Per-container price split (single, equal-N, mixed-type)
- Period helper (PAN-style 26→25)
- End-to-end: SettlementData assembly and Excel generation
"""

import io
from datetime import date

import openpyxl
import pytest

from app.contexts.billing.domain.value_objects import (
    SettlementPeriod,
    settlement_period_for as _settlement_period,
)
from app.contexts.billing.infrastructure.excel_writer import (
    generate_pan_bk_sl_workbook,
    settlement_filename,
)
from app.contexts.billing.infrastructure.settlement_loader import (
    SqlSettlementDataLoader,
    _split_unit_price_per_container,
)
from app.models.domain import (
    Partner,
    TripOrder,
    TripOrderContainer,
)
from app.utils.number_to_words_vi import number_to_vietnamese_words


def settlement_period_for(year: int, month: int):
    """Test shim — keeps the (start, end) tuple shape the legacy tests use."""
    p = _settlement_period(year, month)
    return p.start, p.end


async def load_settlement_data(db_session, partner_id, period_start, period_end):
    loader = SqlSettlementDataLoader(db_session)
    return await loader.load(
        partner_id=partner_id,
        period=SettlementPeriod(start=period_start, end=period_end),
    )


# ---------------------------------------------------------------------------
# number_to_vietnamese_words
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "amount,expected_endswith",
    [
        (0, "Không đồng"),
        (1, "Một đồng"),
        (15, "Mười lăm đồng"),
        (21, "Hai mươi mốt đồng"),
        (105, "Một trăm linh năm đồng"),
        (1_000, "Một nghìn đồng"),
        (1_005_000, "Một triệu không trăm linh năm nghìn đồng"),
    ],
)
def test_number_to_words_simple(amount: int, expected_endswith: str):
    assert number_to_vietnamese_words(amount) == expected_endswith


def test_number_to_words_billion_range():
    text = number_to_vietnamese_words(1_234_567_890)
    assert text.startswith("Một tỷ ")
    assert text.endswith("đồng")
    assert "triệu" in text and "nghìn" in text


# ---------------------------------------------------------------------------
# Per-container price split
# ---------------------------------------------------------------------------

def _fake_container(idx: int, work_type: str = "F20") -> TripOrderContainer:
    c = TripOrderContainer(
        trip_order_id=1,
        container_number=f"TEST{idx:07d}",
        work_type=work_type,
    )
    c.id = idx
    return c


def test_split_single_container_takes_full_price():
    conts = [_fake_container(1)]
    out = _split_unit_price_per_container(1_500_000, conts)
    assert out == {1: 1_500_000}


def test_split_two_containers_equal_with_remainder_on_last():
    conts = [_fake_container(1), _fake_container(2)]
    out = _split_unit_price_per_container(999_999, conts)
    assert out == {1: 499_999, 2: 500_000}
    assert sum(out.values()) == 999_999


def test_split_three_mixed_type_keeps_total_intact():
    conts = [
        _fake_container(1, "F20"),
        _fake_container(2, "F40"),
        _fake_container(3, "E20"),
    ]
    out = _split_unit_price_per_container(2_000_000, conts)
    assert sum(out.values()) == 2_000_000


# ---------------------------------------------------------------------------
# Period helper
# ---------------------------------------------------------------------------

def test_settlement_period_april_2026():
    start, end = settlement_period_for(2026, 4)
    assert start == date(2026, 3, 26)
    assert end == date(2026, 4, 25)


def test_settlement_period_january_crosses_year():
    start, end = settlement_period_for(2026, 1)
    assert start == date(2025, 12, 26)
    assert end == date(2026, 1, 25)


# ---------------------------------------------------------------------------
# End-to-end (DB → Excel) with the sqlite test fixture
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_load_settlement_data_aggregates_routes(db_session):
    partner = Client(
        code="PAN",
        name="CÔNG TY TNHH PAN HẢI AN",
        phone="0225",
        tax_code="0201815115",
        address="Lô KB5, KCN Nam Đình Vũ, Hải An, Hải Phòng",
        is_active=True,
    )
    db_session.add(partner)
    await db_session.flush()

    from app.models.domain import Location
    pickup = Location(name="PAN HA", is_active=True)
    dropoff = Location(name="HẢI AN", is_active=True)
    db_session.add_all([pickup, dropoff])
    await db_session.flush()

    t1 = TripOrder(
        trip_date=date(2026, 4, 1),
        partner_id=partner.id,
        pickup_location_id=pickup.id,
        dropoff_location_id=dropoff.id,
        unit_price=1_309_742,
        driver_salary=400_000,
        allowance=0,
        status="PENDING",
    )
    db_session.add(t1)
    await db_session.flush()
    db_session.add_all([
        TripOrderContainer(trip_order_id=t1.id, container_number="TEST0000001", work_type="F20"),
        TripOrderContainer(trip_order_id=t1.id, container_number="TEST0000002", work_type="F20"),
    ])

    t2 = TripOrder(
        trip_date=date(2026, 4, 2),
        partner_id=partner.id,
        pickup_location_id=pickup.id,
        dropoff_location_id=dropoff.id,
        unit_price=681_050,
        driver_salary=420_000,
        allowance=0,
        status="PENDING",
    )
    db_session.add(t2)
    await db_session.flush()
    db_session.add(
        TripOrderContainer(trip_order_id=t2.id, container_number="TEST0000003", work_type="F40")
    )
    await db_session.flush()

    data = await load_settlement_data(
        db_session, partner.id, date(2026, 3, 26), date(2026, 4, 25)
    )

    assert len(data.trip_lines) == 3
    assert len(data.route_summary) == 1
    summary = data.route_summary[0]
    assert summary.f20_count == 2
    assert summary.f40_count == 1
    assert summary.empty_count == 0
    assert summary.total_amount == 1_309_742 + 681_050
    assert data.total_pre_vat == 1_309_742 + 681_050


@pytest.mark.asyncio
async def test_generate_workbook_produces_two_sheets_with_data(db_session):
    partner = Client(
        code="PAN",
        name="CÔNG TY TNHH PAN HẢI AN",
        phone="0225",
        tax_code="0201815115",
        address="Hải Phòng",
        is_active=True,
    )
    db_session.add(partner)
    await db_session.flush()

    from app.models.domain import Location
    pickup = Location(name="PAN HA", is_active=True)
    dropoff = Location(name="HẢI AN", is_active=True)
    db_session.add_all([pickup, dropoff])
    await db_session.flush()

    trip = TripOrder(
        trip_date=date(2026, 4, 10),
        partner_id=partner.id,
        pickup_location_id=pickup.id,
        dropoff_location_id=dropoff.id,
        unit_price=654_871,
        driver_salary=400_000,
        allowance=0,
        status="PENDING",
    )
    db_session.add(trip)
    await db_session.flush()
    db_session.add(
        TripOrderContainer(trip_order_id=trip.id, container_number="HACU1234567", work_type="F20")
    )
    await db_session.flush()

    data = await load_settlement_data(
        db_session, partner.id, date(2026, 3, 26), date(2026, 4, 25)
    )
    blob = generate_pan_bk_sl_workbook(data)

    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert wb.sheetnames == ["BKTT T4.26", "SL T4.26"]

    sl = wb["SL T4.26"]
    assert sl.cell(row=11, column=4).value == "HACU1234567"
    assert sl.cell(row=11, column=5).value == 1
    assert sl.cell(row=11, column=14).value == 654_871

    bktt = wb["BKTT T4.26"]
    assert bktt.cell(row=11, column=2).value == "PAN HA"
    assert bktt.cell(row=11, column=3).value == "HẢI AN"
    assert bktt.cell(row=11, column=11).value == 654_871

    fname = settlement_filename(data)
    assert fname == "PAN_BK_SL_T04.26_HD.xlsx"


@pytest.mark.asyncio
async def test_workbook_with_no_trips_is_still_valid(db_session):
    partner = Client(
        code="PAN",
        name="CÔNG TY TNHH PAN HẢI AN",
        phone="0225",
        is_active=True,
    )
    db_session.add(partner)
    await db_session.flush()

    data = await load_settlement_data(
        db_session, partner.id, date(2026, 3, 26), date(2026, 4, 25)
    )
    blob = generate_pan_bk_sl_workbook(data)

    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "BKTT T4.26" in wb.sheetnames
    assert "SL T4.26" in wb.sheetnames
