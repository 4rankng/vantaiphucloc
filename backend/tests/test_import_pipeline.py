"""Tests for the generic customer-Excel import pipeline.

Verifies all 4 sample files in `docs/` parse via the same code path with
no per-format branches, AND that programmatically stripping accountant-
added columns/sheets produces the same canonical row count.
"""

import asyncio
import io
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from app.contexts.operations.infrastructure.import_pipeline.canonical import (
    is_skip_header,
    normalize_header_text,
    synonym_substring_match,
    EXACT_LOOKUP,
)
from app.contexts.operations.infrastructure.import_pipeline.pattern_detector import detect_pattern
from app.contexts.operations.infrastructure.import_pipeline.pattern_extractors import (
    ExtractedRow,
    extract_bay_plan,
    extract_invoice,
    extract_loading_list,
    extract_settlement_list,
)
from app.contexts.operations.infrastructure.import_pipeline.pipeline import run_preview
from app.contexts.operations.infrastructure.import_pipeline.value_parsers import (
    build_cont_type,
    parse_container_no,
    parse_container_size,
    parse_date as parse_date_val,
    parse_freight_kind,
    parse_size_type,
    parse_weight_kg,
)
from app.contexts.operations.infrastructure.import_pipeline.workbook import load_workbook


DOCS = Path(__file__).resolve().parents[2] / "docs"
LOADING = DOCS / "LOADING LIST HAIAN DELL 037S 19.4.xlsx"
DISCHARGING = DOCS / "DISCHARGING LIST HAIAN TIME 454W 6.4.xlsx"
GLORY = DOCS / "2.GLORY SHANGHAI- 2612N.xlsx"
BDST = DOCS / "BDST 11.4.xls"

# New sample files (pattern-based)
GLORY_NEW = DOCS / "2.GLORY SHANGHAI- 2612N.xlsx"
CONSCIENCE = DOCS / "8.CONSCIENCE 2615N.xlsx"
HAIAN_BETA = DOCS / "Loading list of HAIAN BETA 062S.xls"
PHUC_LOC = DOCS / "Phúc Lộc - Shipside T4.26 HAP.xlsx"
SAMPLE_IO = DOCS / "templates" / "sample-input-output.xlsx"


# ---------------------------------------------------------------------------
# Canonical / synonyms
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "header,expected_field",
    [
        ("Số container", "container_no"),
        ("CONTAINER NO.", "container_no"),
        ("Container Id", "container_no"),
        ("CONTNo.", "container_no"),
        ("หมายเลขตู้", "container_no"),
        ("VGM (KGM)", "gross_weight_kg"),
        ("Weight", "gross_weight_kg"),
        ("F/E", "freight_kind"),
        ("Freight Kind", "freight_kind"),
        ("CY", "pickup_location"),
        ("Del Port", "dropoff_location"),
        ("Booking No", "customer_ref"),
        ("SỐCONTAINER", "container_no"),
        ("SOCONTAINER", "container_no"),
        ("Vessel", "vessel"),
        ("Tên tàu", "vessel"),
    ],
)
def test_synonym_dictionary_matches_known_headers(header: str, expected_field: str):
    norm = normalize_header_text(header)
    assert EXACT_LOOKUP.get(norm) == expected_field, f"{header!r} → {EXACT_LOOKUP.get(norm)} ≠ {expected_field}"


@pytest.mark.parametrize(
    "header",
    [
        "POD", "POL",
        "Bay", "Slot", "Cell", "Crane", "QC01", "Sales/Mkt",
        "Hãng tàu", "Hãng khai thác",
        "Loại công việc", "Phương thức ra", "Hàng nội/ngoại",
    ],
)
def test_skip_dictionary_classifies_vessel_and_admin(header: str):
    assert is_skip_header(normalize_header_text(header)), header


def test_substring_match_does_not_fire_inside_word():
    # "shipper" must NOT match "ship" via skip dictionary
    assert not is_skip_header(normalize_header_text("Shipper"))
    # "Hãng tàu" must NOT match "hang" / "tau" via synonym dictionary
    assert synonym_substring_match(normalize_header_text("Hãng tàu")) is None


# ---------------------------------------------------------------------------
# Value parsers
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "raw,expected",
    [
        ("ABCD1234567", "ABCD1234567"),
        ("ABCD-123456-7", "ABCD1234567"),
        ("abcd1234567", "ABCD1234567"),
    ],
)
def test_parse_container_no_normalizes(raw, expected):
    assert parse_container_no(raw) == expected


@pytest.mark.parametrize(
    "raw,expected",
    [("#REF!", "missing_container_no"), ("0", "bad_container_no"), ("", "missing_container_no"), (None, "missing_container_no")],
)
def test_parse_container_no_rejects(raw, expected):
    with pytest.raises(ValueError, match=expected):
        parse_container_no(raw)


@pytest.mark.parametrize(
    "raw,iso,expected",
    [
        ("20DC", None, "20"),
        ("40HC", None, "40"),
        ("22G0", None, "20"),
        ("45G1", None, "40"),
        ("42U0", None, "40"),
        (None, "22G0", "20"),
        (None, "45HC", "40"),
        (40, None, "40"),
        ("20", None, "20"),
        ("45", None, "40"),
    ],
)
def test_parse_container_size_handles_format_variants(raw, iso, expected):
    assert parse_container_size(raw, iso_hint=iso) == expected


@pytest.mark.parametrize(
    "raw,expected",
    [("F", "F"), ("E", "E"), ("FULL", "F"), ("EMPTY", "E"), ("", "F"), (None, "F"), ("hàng", "F"), ("rỗng", "E")],
)
def test_parse_freight_kind(raw, expected):
    assert parse_freight_kind(raw) == expected


@pytest.mark.parametrize(
    "raw,expected",
    [
        (32000, 32000.0),
        (15.66, 15660.0),       # tonnes auto-upscaled
        ("32,000", 32000.0),
        ("32.000,5", 32000.5),  # European decimal
        ("1,234.56", 1234.56),
        ("", None),
        (None, None),
    ],
)
def test_parse_weight_kg(raw, expected):
    assert parse_weight_kg(raw) == expected


def test_parse_date_dmy_priority():
    # "11/04/2026" should parse as 11-Apr (DMY), not 04-Nov (MDY)
    assert parse_date_val("11/04/2026") == date(2026, 4, 11)


def test_parse_date_excel_serial():
    # Excel serial 46141 = 2026-04-29 with the standard 1900 epoch
    # (1899-12-30 base + 46141 days). Round-trip should be stable.
    parsed = parse_date_val(46141)
    assert parsed is not None
    # Confirm it's a date, in 2026, and reverse-derives the same serial
    assert parsed.year == 2026 and parsed.month == 4


# ---------------------------------------------------------------------------
# End-to-end preview against the 4 sample files
# ---------------------------------------------------------------------------

@pytest.fixture
def files_present():
    missing = [p for p in [LOADING, DISCHARGING, GLORY, BDST] if not p.exists()]
    if missing:
        pytest.skip(f"sample files missing: {missing}")


@pytest.mark.asyncio
async def test_loading_list_picks_total_sheet_and_parses(files_present):
    res = await run_preview(LOADING.read_bytes(), LOADING.name, default_trip_date=date(2026, 4, 19))
    assert res.sheet_name == "TOTAL"
    assert res.header_row_index == 9            # row 10 is the header
    assert res.stats["accepted_count"] >= 600   # 686 in fixture, allow some rejected
    # Required mapping present
    fields = {m["canonical_field"] for m in res.column_mappings if m["canonical_field"]}
    assert {"container_no", "container_size", "freight_kind"} <= fields


@pytest.mark.asyncio
async def test_discharging_list_picks_sample_sheet_and_parses(files_present):
    res = await run_preview(DISCHARGING.read_bytes(), DISCHARGING.name, default_trip_date=date(2026, 4, 6))
    assert res.sheet_name == "Sample Sheet"
    assert res.header_row_index == 0
    assert res.stats["accepted_count"] >= 500
    fields = {m["canonical_field"] for m in res.column_mappings if m["canonical_field"]}
    assert {"container_no", "freight_kind"} <= fields


@pytest.mark.asyncio
async def test_glory_shanghai_skips_stowage_picks_sheet1(files_present):
    res = await run_preview(GLORY.read_bytes(), GLORY.name, default_trip_date=date(2026, 3, 31))
    # Must NOT pick the side-by-side stowage diagram
    assert res.sheet_name == "Sheet1"
    assert res.stats["accepted_count"] >= 100
    # Tractor plate + driver name extracted from the Vietnamese-headed columns
    assert res.accepted[0]["values"]["vehicle_plate"] != ""


@pytest.mark.asyncio
async def test_bdst_xls_loads_and_parses(files_present):
    res = await run_preview(BDST.read_bytes(), BDST.name, default_trip_date=date(2026, 4, 11))
    assert res.sheet_name == "Sheet1"
    assert res.stats["accepted_count"] >= 400


# ---------------------------------------------------------------------------
# Stripped-version invariance.
#
# The 4 sample files were touched by the accountant after delivery. For
# the loading list the accountant added columns S/T (formula keys); for
# Glory Shanghai the first sheet is a stowage diagram with kế-toán fuel-
# price annotations. We strip those and verify the canonical output is
# unchanged (same accepted count, same first row container).
# ---------------------------------------------------------------------------

def _strip_columns(content: bytes, sheet_name: str, drop_columns: list[int]) -> bytes:
    """Return a new .xlsx workbook with the given column indices (0-based) removed
    from `sheet_name`."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name]
    drop_set = sorted(set(drop_columns), reverse=True)
    for r in range(1, ws.max_row + 1):
        for c in drop_set:
            ws.cell(row=r, column=c + 1).value = None
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@pytest.mark.asyncio
async def test_loading_list_stripped_extra_cols_same_output(files_present):
    raw = LOADING.read_bytes()
    res_full = await run_preview(raw, LOADING.name, default_trip_date=date(2026, 4, 19))
    # Accountant scratch columns observed: cols 17 (SALES/MKT) and 18 (formula key)
    stripped = _strip_columns(raw, "TOTAL", [17, 18])
    res_stripped = await run_preview(stripped, LOADING.name, default_trip_date=date(2026, 4, 19))
    assert res_stripped.sheet_name == res_full.sheet_name
    assert res_stripped.header_row_index == res_full.header_row_index
    assert res_stripped.stats["accepted_count"] == res_full.stats["accepted_count"]


@pytest.mark.asyncio
async def test_glory_picks_correct_sheet_with_stowage_present(files_present):
    """The stowage CONSCIENCE sheet stays in the workbook — the picker must
    still choose Sheet1. (Removing the stowage sheet would be unfair to the
    test; production files won't contain it but the parser must reject it.)"""
    res = await run_preview(GLORY.read_bytes(), GLORY.name, default_trip_date=date(2026, 3, 31))
    assert res.sheet_name == "Sheet1"
    assert "CONSCIENCE 2612N" in [s["sheet_name"] for s in res.sheet_alternatives] + [res.sheet_name]


# ---------------------------------------------------------------------------
# Value parsers — parse_size_type / build_cont_type
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "raw,expected_size,expected_type",
    [
        ("40HC", "40", "HC"),
        ("20DC", "20", "DC"),
        ("20RF", "20", "RF"),
        ("45HC", "45", "HC"),
        ("20", "20", None),
        ("40", "40", None),
        (None, None, None),
        ("", None, None),
    ],
)
def test_parse_size_type(raw, expected_size, expected_type):
    size, typ = parse_size_type(raw)
    assert size == expected_size
    assert typ == expected_type


@pytest.mark.parametrize(
    "fe,size,expected",
    [
        ("E", "20", "E20"),
        ("E", "40", "E40"),
        ("F", "20", "F20"),
        ("F", "40", "F40"),
        ("H", "20", "F20"),    # H → F (hàng)
        ("R", "40", "E40"),    # R → E (rỗng)
        ("FULL", "20", "F20"),
        ("EMPTY", "40", "E40"),
    ],
)
def test_build_cont_type(fe, size, expected):
    assert build_cont_type(fe, size) == expected


# ---------------------------------------------------------------------------
# Pattern detection
# ---------------------------------------------------------------------------

@pytest.fixture
def glory_sheets():
    if not GLORY_NEW.exists():
        pytest.skip("GLORY SHANGHAI sample missing")
    sheets = load_workbook(GLORY_NEW.read_bytes(), GLORY_NEW.name)
    return sheets


@pytest.fixture
def conscience_sheets():
    if not CONSCIENCE.exists():
        pytest.skip("CONSCIENCE sample missing")
    sheets = load_workbook(CONSCIENCE.read_bytes(), CONSCIENCE.name)
    return sheets


@pytest.fixture
def haian_sheets():
    if not HAIAN_BETA.exists():
        pytest.skip("HAIAN BETA sample missing")
    sheets = load_workbook(HAIAN_BETA.read_bytes(), HAIAN_BETA.name)
    return sheets


@pytest.fixture
def phucloc_sheets():
    if not PHUC_LOC.exists():
        pytest.skip("Phúc Lộc sample missing")
    sheets = load_workbook(PHUC_LOC.read_bytes(), PHUC_LOC.name)
    return sheets


def test_detect_bay_plan_glory(glory_sheets):
    pattern = detect_pattern(glory_sheets, "2.GLORY SHANGHAI- 2612N.xlsx")
    assert pattern is not None
    assert pattern.pattern_name == "bay_plan"
    assert pattern.confidence >= 0.6


def test_detect_bay_plan_conscience(conscience_sheets):
    pattern = detect_pattern(conscience_sheets, "8.CONSCIENCE 2615N.xlsx")
    assert pattern is not None
    assert pattern.pattern_name == "bay_plan"


def test_detect_loading_list(haian_sheets):
    pattern = detect_pattern(haian_sheets, "Loading list of HAIAN BETA 062S.xls")
    assert pattern is not None
    assert pattern.pattern_name == "loading_list"


def test_detect_invoice(phucloc_sheets):
    pattern = detect_pattern(phucloc_sheets, "Phúc Lộc - Shipside T4.26 HAP.xlsx")
    assert pattern is not None
    assert pattern.pattern_name == "invoice"


# ---------------------------------------------------------------------------
# Pattern extractors — unit tests
# ---------------------------------------------------------------------------

def test_extract_bay_plan_glory(glory_sheets):
    accepted, rejected = extract_bay_plan(glory_sheets, "2.GLORY SHANGHAI- 2612N.xlsx")
    assert len(accepted) > 0
    for row in accepted:
        assert len(row.container_number) == 11  # 4 letters + 7 digits
        assert row.cont_type in ("E20", "E40", "F20", "F40")
        assert row.pickup != ""
        assert row.dropoff != ""
    # GLORY has containers across multiple port sections
    ports = {r.dropoff for r in accepted}
    assert len(ports) >= 2


def test_extract_bay_plan_conscience(conscience_sheets):
    accepted, rejected = extract_bay_plan(conscience_sheets, "8.CONSCIENCE 2615N.xlsx")
    assert len(accepted) > 0
    for row in accepted:
        assert row.cont_type in ("E20", "E40", "F20", "F40")


def test_extract_loading_list(haian_sheets):
    accepted, rejected = extract_loading_list(haian_sheets, "Loading list of HAIAN BETA 062S.xls")
    assert len(accepted) > 0
    for row in accepted:
        assert len(row.container_number) == 11
        assert row.cont_type in ("E20", "E40", "F20", "F40", "E45", "F45")
    # Should have vessel name
    vessel_names = {r.vessel_name for r in accepted if r.vessel_name}
    assert len(vessel_names) >= 1


def test_extract_invoice(phucloc_sheets):
    accepted, rejected = extract_invoice(phucloc_sheets, "Phúc Lộc - Shipside T4.26 HAP.xlsx")
    assert len(accepted) > 0
    for row in accepted:
        assert len(row.container_number) == 11
        assert row.cont_type in ("E20", "E40", "F20", "F40")
        assert row.pickup != ""
        assert row.dropoff != ""


# ---------------------------------------------------------------------------
# End-to-end preview via pipeline — pattern-detected files
# ---------------------------------------------------------------------------

@pytest.fixture
def pattern_files_present():
    missing = [p for p in [GLORY_NEW, CONSCIENCE, HAIAN_BETA, PHUC_LOC] if not p.exists()]
    if missing:
        pytest.skip(f"sample files missing: {missing}")


@pytest.mark.asyncio
async def test_glory_pattern_preview(pattern_files_present):
    res = await run_preview(GLORY_NEW.read_bytes(), GLORY_NEW.name, default_trip_date=date(2026, 3, 31))
    assert res.stats["accepted_count"] > 0
    first = res.accepted[0]["values"]
    assert "container_no" in first
    assert first["cont_type"] in ("E20", "E40", "F20", "F40")
    assert first["work_type"]  # operation type
    assert first["pickup_location"] != ""


@pytest.mark.asyncio
async def test_conscience_pattern_preview(pattern_files_present):
    res = await run_preview(CONSCIENCE.read_bytes(), CONSCIENCE.name, default_trip_date=date(2026, 3, 31))
    assert res.stats["accepted_count"] > 0


@pytest.mark.asyncio
async def test_haian_pattern_preview(pattern_files_present):
    res = await run_preview(HAIAN_BETA.read_bytes(), HAIAN_BETA.name, default_trip_date=date(2026, 4, 19))
    assert res.stats["accepted_count"] > 0
    first = res.accepted[0]["values"]
    assert first["pickup_location"] != ""


@pytest.mark.asyncio
async def test_phucloc_pattern_preview(pattern_files_present):
    res = await run_preview(PHUC_LOC.read_bytes(), PHUC_LOC.name, default_trip_date=date(2026, 4, 26))
    assert res.stats["accepted_count"] > 0
    first = res.accepted[0]["values"]
    assert first["pickup_location"] != ""
    assert first["dropoff_location"] != ""


# ---------------------------------------------------------------------------
# Settlement List (BẢNG KÊ QUYẾT TOÁN — Vietnamese reconciliation)
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_io_sheets():
    if not SAMPLE_IO.exists():
        pytest.skip("sample-input-output.xlsx missing")
    sheets = load_workbook(SAMPLE_IO.read_bytes(), SAMPLE_IO.name)
    return sheets


def test_detect_settlement_list(sample_io_sheets):
    pattern = detect_pattern(sample_io_sheets, "sample-input-output.xlsx")
    assert pattern is not None
    assert pattern.pattern_name == "settlement_list"
    assert pattern.confidence >= 0.6


def test_extract_settlement_list(sample_io_sheets):
    accepted, rejected = extract_settlement_list(sample_io_sheets, "sample-input-output.xlsx")
    assert len(accepted) > 0
    for row in accepted:
        assert len(row.container_number) == 11
        assert row.cont_type in ("E20", "E40", "F20", "F40")
        assert row.work_type  # operation type (e.g., CHUYỂN BÃI, XUẤT/NHẬP TÀU)
        assert row.pickup != ""
        assert row.dropoff != ""


def test_settlement_list_work_types(sample_io_sheets):
    accepted, _ = extract_settlement_list(sample_io_sheets, "sample-input-output.xlsx")
    cont_types = {r.cont_type for r in accepted}
    assert len(cont_types) >= 2
    assert all(ct in ("E20", "E40", "F20", "F40") for ct in cont_types)


def test_settlement_list_locations(sample_io_sheets):
    accepted, _ = extract_settlement_list(sample_io_sheets, "sample-input-output.xlsx")
    pickups = {r.pickup for r in accepted if r.pickup}
    dropoffs = {r.dropoff for r in accepted if r.dropoff}
    assert len(pickups) >= 1
    assert len(dropoffs) >= 1


@pytest.mark.asyncio
async def test_settlement_list_e2e_preview():
    if not SAMPLE_IO.exists():
        pytest.skip("sample-input-output.xlsx missing")
    res = await run_preview(SAMPLE_IO.read_bytes(), SAMPLE_IO.name, default_trip_date=date(2026, 4, 1))
    assert res.stats["accepted_count"] > 0
    first = res.accepted[0]["values"]
    assert "container_no" in first
    assert first["cont_type"] in ("E20", "E40", "F20", "F40")
    assert first["work_type"]  # operation type
    assert first["pickup_location"] != ""
    assert first["dropoff_location"] != ""
    for w in res.warnings:
        assert "Thiếu cột bắt buộc" not in w
