"""Tests for route pricing Excel import parser."""
from __future__ import annotations

from io import BytesIO

import openpyxl

from app.contexts.route_pricing.infrastructure.route_pricing_import import (
    _find_header_row,
    _normalize_operation_type,
    _parse_int_price,
    parse_route_pricing_bytes,
)


def _make_xlsx(sheet_name: str, rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── _find_header_row ─────────────────────────────────────────────


def test_find_header_row_standard():
    rows = [["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "F20", "F40", "E20", "E40", "TÁC NGHIỆP"]]
    assert _find_header_row(rows) == 0


def test_find_header_row_case_insensitive():
    rows = [["stt", "chủ hàng", "điểm đi", "điểm đến", "f20", "f40", "e20", "e40", "tác nghiệp"]]
    assert _find_header_row(rows) == 0


def test_find_header_row_whitespace_tolerant():
    rows = [[" STT ", " CHỦ HÀNG ", " ĐIỂM ĐI ", " ĐIỂM ĐẾN ", " F20 ", " F40 ", " E20 ", " E40 ", " TÁC NGHIỆP "]]
    assert _find_header_row(rows) == 0


def test_find_header_row_no_match():
    rows = [["foo", "bar", "baz"]]
    assert _find_header_row(rows) is None


def test_find_header_row_partial_headers():
    rows = [["STT", "CHỦ HÀNG", "ĐIỂM ĐI"]]
    assert _find_header_row(rows) is None


def test_find_header_row_requires_price_col():
    rows = [["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "TÁC NGHIỆP"]]
    assert _find_header_row(rows) is None


# ── _normalize_operation_type ────────────────────────────────────


def test_normalize_op_exact():
    assert _normalize_operation_type("XUẤT/NHẬP TÀU") == "XUẤT/NHẬP TÀU"


def test_normalize_op_case_insensitive():
    assert _normalize_operation_type("xuất/nhập tàu") == "XUẤT/NHẬP TÀU"


def test_normalize_op_whitespace():
    assert _normalize_operation_type("  chuyển   bãi  ") == "CHUYỂN BÃI"


def test_normalize_op_invalid():
    assert _normalize_operation_type("INVALID") is None


def test_normalize_op_empty():
    assert _normalize_operation_type("") is None


def test_normalize_op_all_valid():
    from app.contexts.route_pricing.domain.value_objects import VALID_OPERATION_TYPES
    for op in VALID_OPERATION_TYPES:
        assert _normalize_operation_type(op) == op


# ── _parse_int_price ─────────────────────────────────────────────


def test_price_none():
    assert _parse_int_price(None) is None


def test_price_empty_str():
    assert _parse_int_price("") is None


def test_price_int():
    assert _parse_int_price(100) == 100


def test_price_float_rounds():
    assert _parse_int_price(218352.5) == 218352  # banker's rounding


def test_price_str():
    assert _parse_int_price("500") == 500


def test_price_invalid():
    assert _parse_int_price("abc") is None


# ── parse_route_pricing_bytes ────────────────────────────────────


def test_parse_basic():
    xlsx = _make_xlsx("Sheet1", [
        ["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "F20", "F40", "E20", "E40", "TÁC NGHIỆP"],
        [1, "HPH", "HECHUN", "HẢI AN", 400000, 400000, None, None, "XUẤT/ NHẬP TÀU"],
        [2, "VNB", "VIMC", "VIMC", 67500, 135000, 67500, 135000, "XUẤT/ NHẬP TÀU"],
    ])
    result = parse_route_pricing_bytes(xlsx)
    assert result["sheet_name"] == "Sheet1"
    assert len(result["rows"]) == 2
    assert result["rows"][0]["client_raw"] == "HPH"
    assert result["rows"][0]["f20_price"] == 400000
    assert result["rows"][0]["e20_price"] is None
    assert result["rows"][0]["operation_type"] == "XUẤT/NHẬP TÀU"
    assert result["stats"]["total"] == 2


def test_parse_skips_empty_rows():
    xlsx = _make_xlsx("Sheet1", [
        ["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "F20", "F40", "E20", "E40", "TÁC NGHIỆP"],
        [1, "HPH", "HECHUN", "HẢI AN", 400000, None, None, None, "XUẤT/ NHẬP TÀU"],
        [None, None, None, None, None, None, None, None, None],
        [3, "VNB", "VIMC", "VIMC", 67500, 135000, None, None, "XUẤT/ NHẬP TÀU"],
    ])
    result = parse_route_pricing_bytes(xlsx)
    assert len(result["rows"]) == 2


def test_parse_skips_aggregate_rows():
    xlsx = _make_xlsx("Sheet1", [
        ["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "F20", "F40", "E20", "E40", "TÁC NGHIỆP"],
        [1, "HPH", "HECHUN", "HẢI AN", 400000, None, None, None, "XUẤT/ NHẬP TÀU"],
        [None, "TỔNG CỘNG", None, None, 999999, None, None, None, None],
    ])
    result = parse_route_pricing_bytes(xlsx)
    assert len(result["rows"]) == 1


def test_parse_no_matching_sheet():
    xlsx = _make_xlsx("RandomSheet", [["A", "B", "C"]])
    result = parse_route_pricing_bytes(xlsx)
    assert len(result["rows"]) == 0
    assert len(result["warnings"]) > 0


def test_parse_header_not_first_row():
    xlsx = _make_xlsx("Data", [
        ["BẢNG CƯỚC TUYẾN"],
        ["Ngày cập nhật"],
        ["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "F20", "F40", "E20", "E40", "TÁC NGHIỆP"],
        [1, "PAN", "PAN", "NAM ĐÌNH VŨ", 454158, 489065, 218352, 436705, "XUẤT/ NHẬP TÀU"],
    ])
    result = parse_route_pricing_bytes(xlsx)
    assert len(result["rows"]) == 1
    assert result["rows"][0]["client_raw"] == "PAN"


def test_parse_multiple_sheets():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Other"
    ws1.append(["A", "B"])
    ws2 = wb.create_sheet("CƯỚC TUYẾN")
    ws2.append(["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "F20", "F40", "E20", "E40", "TÁC NGHIỆP"])
    ws2.append([1, "HPH", "HECHUN", "HẢI AN", 400000, None, None, None, "XUẤT/ NHẬP TÀU"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = parse_route_pricing_bytes(buf.read())
    assert result["sheet_name"] == "CƯỚC TUYẾN"
    assert len(result["rows"]) == 1


def test_parse_price_rounding():
    xlsx = _make_xlsx("Sheet1", [
        ["STT", "CHỦ HÀNG", "ĐIỂM ĐI", "ĐIỂM ĐẾN", "F20", "F40", "E20", "E40", "TÁC NGHIỆP"],
        [1, "PAN", "PAN", "NAM ĐÌNH VŨ", 218352.5, None, None, None, "XUẤT/ NHẬP TÀU"],
    ])
    result = parse_route_pricing_bytes(xlsx)
    assert result["rows"][0]["f20_price"] == 218352  # banker's rounding
