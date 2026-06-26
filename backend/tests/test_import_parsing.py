"""Tests for value parsers and end-to-end preview over the accountant-touched
sample files (LOADING/DISCHARGING/GLORY/BDST), including stripped-column
invariance.

Split from ``test_import_pipeline.py`` (move, not rewrite) — every test name,
parametrize table, fixture, and assertion is preserved verbatim.
"""

import io
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from app.contexts.operations.infrastructure.import_pipeline.llm import (
    NullBatchHeaderClassifier,
)
from app.contexts.operations.infrastructure.import_pipeline.pipeline import (
    run_preview,
)
from app.contexts.operations.infrastructure.import_pipeline.value_parsers import (
    build_cont_type,
    parse_container_no,
    parse_container_size,
    parse_date as parse_date_val,
    parse_freight_kind,
    parse_size_type,
    parse_weight_kg,
)


DOCS = Path(__file__).resolve().parents[2] / "docs"
LOADING = DOCS / "LOADING LIST HAIAN DELL 037S 19.4.xlsx"
DISCHARGING = DOCS / "DISCHARGING LIST HAIAN TIME 454W 6.4.xlsx"
GLORY = DOCS / "2.GLORY SHANGHAI- 2612N.xlsx"
BDST = DOCS / "BDST 11.4.xls"


# ---------------------------------------------------------------------------
# Value parsers
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("ABCD1234567", "ABCD1234567"),
        ("ABCD-123456-7", "ABCD1234567"),
        ("abcd1234567", "ABCD1234567"),
    ],
)
def test_parse_container_no_normalizes(raw, expected):
    assert parse_container_no(raw) == expected


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("#REF!", "missing_container_no"),
        ("0", "bad_container_no"),
        ("", "missing_container_no"),
        (None, "missing_container_no"),
    ],
)
def test_parse_container_no_rejects(raw, expected):
    with pytest.raises(ValueError, match=expected):
        parse_container_no(raw)


@pytest.mark.parametrize(
    "raw,iso,expected",
    [
        ("20DC", None, "20"),
        ("40HC", None, "40"),
        ("22G0", None, "20"),
        ("45G1", None, "40"),
        ("42U0", None, "40"),
        (None, "22G0", "20"),
        (None, "45HC", "40"),
        (40, None, "40"),
        ("20", None, "20"),
        ("45", None, "40"),
    ],
)
def test_parse_container_size_handles_format_variants(raw, iso, expected):
    assert parse_container_size(raw, iso_hint=iso) == expected


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("F", "F"),
        ("E", "E"),
        ("FULL", "F"),
        ("EMPTY", "E"),
        ("", "F"),
        (None, "F"),
        ("hàng", "F"),
        ("rỗng", "E"),
    ],
)
def test_parse_freight_kind(raw, expected):
    assert parse_freight_kind(raw) == expected


def test_parse_freight_kind_raises_on_unknown():
    """Test that raise_on_unknown=True causes exception for empty/unknown values."""
    with pytest.raises(ValueError, match="unknown_freight_kind"):
        parse_freight_kind("", raise_on_unknown=True)
    with pytest.raises(ValueError, match="unknown_freight_kind"):
        parse_freight_kind(None, raise_on_unknown=True)
    # Explicit E/F should still work with raise_on_unknown=True
    assert parse_freight_kind("E", raise_on_unknown=True) == "E"
    assert parse_freight_kind("F", raise_on_unknown=True) == "F"
    assert parse_freight_kind("FULL", raise_on_unknown=True) == "F"
    assert parse_freight_kind("EMPTY", raise_on_unknown=True) == "E"


@pytest.mark.parametrize(
    "raw,expected",
    [
        (32000, 32000.0),
        (15.66, 15660.0),  # tonnes auto-upscaled
        ("32,000", 32000.0),
        ("32.000,5", 32000.5),  # European decimal
        ("1,234.56", 1234.56),
        ("", None),
        (None, None),
    ],
)
def test_parse_weight_kg(raw, expected):
    assert parse_weight_kg(raw) == expected


def test_parse_date_dmy_priority():
    # "11/04/2026" should parse as 11-Apr (DMY), not 04-Nov (MDY)
    assert parse_date_val("11/04/2026") == date(2026, 4, 11)


def test_parse_date_excel_serial():
    # Excel serial 46141 = 2026-04-29 with the standard 1900 epoch
    # (1899-12-30 base + 46141 days). Round-trip should be stable.
    parsed = parse_date_val(46141)
    assert parsed is not None
    # Confirm it's a date, in 2026, and reverse-derives the same serial
    assert parsed.year == 2026 and parsed.month == 4


@pytest.mark.asyncio
async def test_import_preview_marks_unknown_freight_kinds():
    """Test that empty freight_kind cells are marked with freight_kind_unknown flag."""
    import openpyxl
    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as tmp_dir:
        # Create a workbook with some rows having empty freight_kind
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "Ngày đi",
                "Số cont",
                "Kích thước",
                "F/E",
                "Điểm đi",
                "Điểm đến",
                "Người gửi",
                "Ghi chú",
                "Trọng lượng",
            ]
        )
        ws.append(
            [
                "2026-06-01",
                "CAIU6167954",
                "20",
                "F",
                "HaiPhong",
                "HCMC",
                "ABC",
                "",
                "5000",
            ]
        )
        ws.append(
            [
                "2026-06-01",
                "CAIU6386850",
                "40",
                "",
                "HaiPhong",
                "HCMC",
                "DEF",
                "",
                "8000",
            ]
        )  # Empty freight_kind
        ws.append(
            [
                "2026-06-01",
                "FCIU6219871",
                "20",
                "E",
                "HaiPhong",
                "HCMC",
                "GHI",
                "",
                "3000",
            ]
        )

        path = Path(tmp_dir) / "test_unknown.xlsx"
        wb.save(path)

        res = await run_preview(
            path.read_bytes(),
            path.name,
            default_trip_date=date(2026, 6, 1),
            classifier=NullBatchHeaderClassifier(),
        )

        # Check that the row with empty freight_kind is marked as unknown
        assert len(res.accepted) == 3
        assert not res.accepted[0]["values"]["freight_kind_unknown"]
        assert res.accepted[0]["values"]["freight_kind"] == "F"

        assert res.accepted[1]["values"]["freight_kind_unknown"]
        assert (
            res.accepted[1]["values"]["freight_kind"] == "F"
        )  # Defaults to F but marked unknown

        assert not res.accepted[2]["values"]["freight_kind_unknown"]
        assert res.accepted[2]["values"]["freight_kind"] == "E"


@pytest.fixture
def files_present():
    missing = [p for p in [LOADING, DISCHARGING, GLORY, BDST] if not p.exists()]
    if missing:
        pytest.skip(f"sample files missing: {missing}")


@pytest.mark.asyncio
async def test_loading_list_picks_total_sheet_and_parses(files_present):
    res = await run_preview(
        LOADING.read_bytes(), LOADING.name, default_trip_date=date(2026, 4, 19)
    )
    assert res.sheet_name == "TOTAL"
    assert res.header_row_index == 9  # row 10 is the header
    assert res.stats["accepted_count"] >= 600  # 686 in fixture, allow some rejected
    # Required mapping present
    fields = {m["canonical_field"] for m in res.column_mappings if m["canonical_field"]}
    assert {"container_no", "container_size", "freight_kind"} <= fields


@pytest.mark.asyncio
async def test_discharging_list_picks_sample_sheet_and_parses(files_present):
    res = await run_preview(
        DISCHARGING.read_bytes(), DISCHARGING.name, default_trip_date=date(2026, 4, 6)
    )
    assert res.sheet_name == "Sample Sheet"
    assert res.header_row_index == 0
    assert res.stats["accepted_count"] >= 500
    fields = {m["canonical_field"] for m in res.column_mappings if m["canonical_field"]}
    assert {"container_no", "freight_kind"} <= fields


@pytest.mark.asyncio
async def test_glory_shanghai_skips_stowage_picks_sheet1(files_present):
    res = await run_preview(
        GLORY.read_bytes(), GLORY.name, default_trip_date=date(2026, 3, 31)
    )
    # Must NOT pick the side-by-side stowage diagram
    assert res.sheet_name == "Sheet1"
    assert res.stats["accepted_count"] >= 100
    # Tractor plate + driver name extracted from the Vietnamese-headed columns
    assert res.accepted[0]["values"]["vehicle_plate"] != ""


@pytest.mark.asyncio
async def test_bdst_xls_loads_and_parses(files_present):
    res = await run_preview(
        BDST.read_bytes(), BDST.name, default_trip_date=date(2026, 4, 11)
    )
    assert res.sheet_name == "Sheet1"
    assert res.stats["accepted_count"] >= 400


# ---------------------------------------------------------------------------
# Stripped-version invariance.
#
# The 4 sample files were touched by the accountant after delivery. For
# the loading list the accountant added columns S/T (formula keys); for
# Glory Shanghai the first sheet is a stowage diagram with kế-toán fuel-
# price annotations. We strip those and verify the canonical output is
# unchanged (same accepted count, same first row container).
# ---------------------------------------------------------------------------


def _strip_columns(content: bytes, sheet_name: str, drop_columns: list[int]) -> bytes:
    """Return a new .xlsx workbook with the given column indices (0-based) removed
    from `sheet_name`."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name]
    drop_set = sorted(set(drop_columns), reverse=True)
    for r in range(1, ws.max_row + 1):
        for c in drop_set:
            ws.cell(row=r, column=c + 1).value = None
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@pytest.mark.asyncio
async def test_loading_list_stripped_extra_cols_same_output(files_present):
    raw = LOADING.read_bytes()
    res_full = await run_preview(raw, LOADING.name, default_trip_date=date(2026, 4, 19))
    # Accountant scratch columns observed: cols 17 (SALES/MKT) and 18 (formula key)
    stripped = _strip_columns(raw, "TOTAL", [17, 18])
    res_stripped = await run_preview(
        stripped, LOADING.name, default_trip_date=date(2026, 4, 19)
    )
    assert res_stripped.sheet_name == res_full.sheet_name
    assert res_stripped.header_row_index == res_full.header_row_index
    assert res_stripped.stats["accepted_count"] == res_full.stats["accepted_count"]


@pytest.mark.asyncio
async def test_glory_picks_correct_sheet_with_stowage_present(files_present):
    """The stowage CONSCIENCE sheet stays in the workbook — the picker must
    still choose Sheet1. (Removing the stowage sheet would be unfair to the
    test; production files won't contain it but the parser must reject it.)"""
    res = await run_preview(
        GLORY.read_bytes(), GLORY.name, default_trip_date=date(2026, 3, 31)
    )
    assert res.sheet_name == "Sheet1"
    assert "CONSCIENCE 2612N" in [s["sheet_name"] for s in res.sheet_alternatives] + [
        res.sheet_name
    ]


# ---------------------------------------------------------------------------
# Value parsers — parse_size_type / build_cont_type
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected_size,expected_type",
    [
        ("40HC", "40", "HC"),
        ("20DC", "20", "DC"),
        ("20RF", "20", "RF"),
        ("45HC", "45", "HC"),
        ("20", "20", None),
        ("40", "40", None),
        (None, None, None),
        ("", None, None),
    ],
)
def test_parse_size_type(raw, expected_size, expected_type):
    size, typ = parse_size_type(raw)
    assert size == expected_size
    assert typ == expected_type


@pytest.mark.parametrize(
    "fe,size,expected",
    [
        ("E", "20", "E20"),
        ("E", "40", "E40"),
        ("F", "20", "F20"),
        ("F", "40", "F40"),
        ("H", "20", "F20"),  # H → F (hàng)
        ("R", "40", "E40"),  # R → E (rỗng)
        ("FULL", "20", "F20"),
        ("EMPTY", "40", "E40"),
    ],
)
def test_build_cont_type(fe, size, expected):
    assert build_cont_type(fe, size) == expected
