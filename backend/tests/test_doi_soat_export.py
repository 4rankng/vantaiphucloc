from datetime import date
from io import BytesIO

import openpyxl
import pytest

from app.contexts.operations.infrastructure.excel.booked_trip_export import (
    generate_doi_soat_excel,
)
from app.models.domain import BookedTrip, Client, DeliveredTrip, Location


@pytest.mark.asyncio
async def test_doi_soat_export_includes_revenue_and_salary_columns(db_session):
    client = Client(id=1, code="KH01", name="Khach Hang 01", is_active=True)
    pickup = Location(id=1, name="Cang Tan Vu", is_active=True)
    dropoff = Location(id=2, name="Kho Hai Phong", is_active=True)
    db_session.add_all([client, pickup, dropoff])

    booked = BookedTrip(
        id=100,
        trip_date=date(2026, 6, 5),
        client_id=1,
        pickup_location_id=1,
        dropoff_location_id=2,
        work_type="NANG HA",
        cont_number="MSCU1234567",
        cont_type="F20",
    )
    delivered = DeliveredTrip(
        id=200,
        trip_date=date(2026, 6, 5),
        client_id=1,
        pickup_location_id=1,
        dropoff_location_id=2,
        work_type="NANG HA",
        cont_number="MSCU1234567",
        cont_type="F20",
        vehicle_plate="15C-12345",
        booked_trip_id=100,
        revenue=1_200_000,
        driver_salary=300_000,
        note="Tài xế báo kẹt cổng",
    )
    db_session.add_all([booked, delivered])
    await db_session.flush()

    content, _client_name = await generate_doi_soat_excel(
        db_session,
        client_id=1,
        date_from="2026-06-01",
        date_to="2026-06-30",
    )

    wb = openpyxl.load_workbook(BytesIO(content), data_only=False)
    ws = wb.active

    headers = [cell.value for cell in ws[10]]
    assert headers[13:17] == ["CƯỚC", "LƯƠNG", "TRẠNG THÁI", "GHI CHÚ"]

    row = [cell.value for cell in ws[12]]
    assert row[13] == 1_200_000
    assert row[14] == 300_000
    assert row[15] == "Đã ghép"
    assert row[16] == "Tài xế báo kẹt cổng"


@pytest.mark.asyncio
async def test_doi_soat_export_without_client_includes_all_clients(db_session):
    client_a = Client(id=11, code="KHA", name="Khach Hang A", is_active=True)
    client_b = Client(id=12, code="KHB", name="Khach Hang B", is_active=True)
    pickup = Location(id=11, name="Cang A", is_active=True)
    dropoff = Location(id=12, name="Kho B", is_active=True)
    db_session.add_all([client_a, client_b, pickup, dropoff])

    trip_a = DeliveredTrip(
        id=211,
        trip_date=date(2026, 6, 5),
        client_id=11,
        pickup_location_id=11,
        dropoff_location_id=12,
        work_type="NANG HA",
        cont_number="AAAA1234567",
        cont_type="F20",
        vehicle_plate="15C-11111",
        revenue=1_000_000,
        driver_salary=250_000,
    )
    trip_b = DeliveredTrip(
        id=212,
        trip_date=date(2026, 6, 6),
        client_id=12,
        pickup_location_id=11,
        dropoff_location_id=12,
        work_type="NANG HA",
        cont_number="BBBB1234567",
        cont_type="F40",
        vehicle_plate="15C-22222",
        revenue=2_000_000,
        driver_salary=500_000,
    )
    db_session.add_all([trip_a, trip_b])
    await db_session.flush()

    content, client_name = await generate_doi_soat_excel(
        db_session,
        client_id=None,
        date_from="2026-06-01",
        date_to="2026-06-30",
    )

    wb = openpyxl.load_workbook(BytesIO(content), data_only=False)
    ws = wb.active

    assert client_name == "Tất cả chủ hàng"
    assert ws["C7"].value == "Tất cả chủ hàng"
    rows = [
        [cell.value for cell in row] for row in ws.iter_rows(min_row=12, max_row=13)
    ]
    assert [row[2] for row in rows] == ["KHA", "KHB"]
    assert [row[3] for row in rows] == ["AAAA1234567", "BBBB1234567"]


@pytest.mark.asyncio
async def test_doi_soat_export_all_clients_does_not_cross_match_same_container(
    db_session,
):
    client_a = Client(id=21, code="KHA", name="Khach Hang A", is_active=True)
    client_b = Client(id=22, code="KHB", name="Khach Hang B", is_active=True)
    pickup = Location(id=21, name="Cang A", is_active=True)
    dropoff = Location(id=22, name="Kho B", is_active=True)
    db_session.add_all([client_a, client_b, pickup, dropoff])

    booked = BookedTrip(
        id=301,
        trip_date=date(2026, 6, 5),
        client_id=21,
        pickup_location_id=21,
        dropoff_location_id=22,
        work_type="NANG HA",
        cont_number="SAME1234567",
        cont_type="F20",
    )
    delivered = DeliveredTrip(
        id=302,
        trip_date=date(2026, 6, 5),
        client_id=22,
        pickup_location_id=21,
        dropoff_location_id=22,
        work_type="NANG HA",
        cont_number="SAME1234567",
        cont_type="F20",
        vehicle_plate="15C-22222",
    )
    db_session.add_all([booked, delivered])
    await db_session.flush()

    content, _client_name = await generate_doi_soat_excel(
        db_session,
        client_id=None,
        date_from="2026-06-01",
        date_to="2026-06-30",
    )

    wb = openpyxl.load_workbook(BytesIO(content), data_only=False)
    ws = wb.active
    rows = [
        [cell.value for cell in row] for row in ws.iter_rows(min_row=12, max_row=13)
    ]

    assert [row[2] for row in rows] == ["KHA", "KHB"]
    assert [row[15] for row in rows] == ["Chưa ghép", "Chưa ghép"]
