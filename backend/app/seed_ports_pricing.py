"""Seed locations (ports) and pricing lanes from the DS_PORT_CƯỚC_TUYẾN Excel.

Usage:
    cd backend && python -m app.seed_ports_pricing

Reads the Excel file at the hardcoded path below and populates:
  - 62 Location records from sheet "DS PORT"
  - 14 Partner records (client type) from CƯỚC TUYẾN "CHỦ HÀNG"
  - Pricing + PricingLine records from CƯỚC TUYẾN

Idempotent — safe to run multiple times. Skips existing records.

Run after ``alembic upgrade head``.
"""

import asyncio
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.database import async_session
from app.models.domain import Location, Client, Pricing, PricingLine

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

_EXCEL_PATH = (
    Path(__file__).resolve().parent.parent.parent
    / "docs" / "seed" / "DS_PORT_CU_OC_TUYEN.xlsx"
)

# Fallback: check inbound media directory
_MEDIA_PATH = (
    Path.home() / ".openclaw" / "media" / "inbound"
)



def _find_excel() -> Path:
    """Locate the Excel file."""
    if _EXCEL_PATH.exists():
        return _EXCEL_PATH
    # Search in media inbound
    for f in _MEDIA_PATH.glob("DS_PORT*.xlsx"):
        return f
    raise FileNotFoundError(
        f"Cannot find Excel file. Searched:\n"
        f"  {_EXCEL_PATH}\n"
        f"  {_MEDIA_PATH}/DS_PORT*.xlsx"
    )


def _normalize_operation_type(value: str | None) -> str:
    if not value:
        return "CHUYỂN BÃI"
    norm = value.strip().upper()
    if "CHUYỂN BÃI" in norm or "CHUYEN BAI" in norm:
        return "CHUYỂN BÃI"
    if "XUẤT" in norm or "NHẬP" in norm or "TAU" in norm or "TÀU" in norm:
        return "XUẤT/NHẬP TÀU"
    if "LẤY VỎ" in norm:
        return "LẤY VỎ HẠ HÀNG"
    if "ĐÓNG KHO" in norm:
        return "ĐÓNG KHO"
    if "SÀ LAN" in norm:
        return "CHẠY SÀ LAN"
    return "CHUYỂN BÃI"


def _read_excel(path: Path) -> tuple[list[str], list[dict]]:
    """Parse both sheets from the Excel file.

    Returns:
        ports: list of 62 port names
        pricing_rows: list of dicts with keys:
            client_name, pickup, dropoff, price, work_type
    """
    # data_only=True reads cached formula results
    wb = load_workbook(path, data_only=True)

    # Sheet 1: DS PORT (column B, rows 1-62)
    ws_ports = wb["DS PORT"]
    ports: list[str] = []
    for row in ws_ports.iter_rows(min_row=1, max_row=ws_ports.max_row, min_col=2, max_col=2, values_only=True):
        if row[0] and str(row[0]).strip():
            ports.append(str(row[0]).strip())

    # Sheet 2: CƯỚC TUYẾN (row 2 = header, data from row 3)
    ws_pricing = wb["CƯỚC TUYẾN"]
    pricing_rows: list[dict] = []
    for row in ws_pricing.iter_rows(min_row=3, max_row=ws_pricing.max_row, values_only=True):
        client_name = row[1]  # B: CHỦ HÀNG
        if not client_name or not str(client_name).strip():
            continue

        def _val(cell):
            """Convert cell value to int or None."""
            if cell is None:
                return None
            try:
                v = float(cell)
                iv = int(v)
                return iv if iv == v else int(round(v))
            except (ValueError, TypeError):
                return None

        prices = {
            "F20": _val(row[4]) if len(row) > 4 else None,
            "F40": _val(row[5]) if len(row) > 5 else None,
            "E20": _val(row[6]) if len(row) > 6 else None,
            "E40": _val(row[7]) if len(row) > 7 else None,
        }
        best_price = max((v for v in prices.values() if v), default=None)
        operation_type = str(row[8]).strip() if row[8] else None

        pricing_rows.append({
            "client_name": str(client_name).strip(),
            "pickup": str(row[2]).strip() if row[2] else None,
            "dropoff": str(row[3]).strip() if row[3] else None,
            "price": best_price,
            "work_type": _normalize_operation_type(operation_type),
        })

    return ports, pricing_rows


async def seed_ports_pricing() -> None:
    path = _find_excel()
    print(f"Reading Excel: {path.name}")
    ports, pricing_rows = _read_excel(path)
    print(f"  DS PORT: {len(ports)} locations")
    print(f"  CƯỚC TUYẾN: {len(pricing_rows)} rows")

    # Collect all location names: ports + locations referenced in pricing
    pricing_loc_names: set[str] = set()
    for r in pricing_rows:
        if r["pickup"]:
            pricing_loc_names.add(r["pickup"])
        if r["dropoff"]:
            pricing_loc_names.add(r["dropoff"])
    all_loc_names = list(dict.fromkeys(ports + sorted(pricing_loc_names - set(ports))))

    async with async_session() as db:
        # ── 1. Locations (62 ports + extras from pricing) ───────────────
        print(f"\n=== Seeding Locations ({len(ports)} ports + {len(pricing_loc_names - set(ports))} from pricing) ===")
        loc_map: dict[str, Location] = {}
        created_locs = 0

        # Load existing locations
        result = await db.execute(select(Location))
        for loc in result.scalars().all():
            loc_map[loc.name] = loc

        for name in all_loc_names:
            if name in loc_map:
                continue
            loc = Location(
                name=name,
                is_active=True,
                pending_geocode=True,
                created_via="seed_ports",
                location_review_needed=False,
            )
            db.add(loc)
            await db.flush()
            loc_map[name] = loc
            created_locs += 1
            print(f"  + {name} (id={loc.id})")

        await db.commit()
        print(f"  Locations: {created_locs} created, {len(all_loc_names) - created_locs} existing")

        # ── 2. Partners (14 chủ hàng as clients) ────────────────────────
        print("\n=== Seeding Clients (Khách hàng) ===")
        client_map: dict[str, Client] = {}

        result = await db.execute(select(Client))
        for p in result.scalars().all():
            if p.code:
                client_map[p.code] = p

        chu_hang_names = sorted({r["client_name"] for r in pricing_rows})
        created_partners = 0

        for name in chu_hang_names:
            code = name.upper().replace(" ", "_")
            if code in client_map:
                print(f"  = {name} (already exists as {code})")
                continue
            client = Client(
                code=code,
                name=name,
                is_active=True,
            )
            db.add(client)
            await db.flush()
            client_map[code] = client
            created_partners += 1
            print(f"  + {name} (code={code}, id={client.id})")

        await db.commit()
        print(f"  Clients: {created_partners} created, {len(chu_hang_names) - created_partners} existing")

        # ── 3. Pricings + PricingLines ──────────────────────────────────
        print("\n=== Seeding Pricings ===")

        # client_id on Pricing = the client (khách hàng/chủ hàng) directly.
        # No PHUCLOC intermediary — Phúc Lộc is the company itself, not a client_name.

        # Load existing pricings for dedup
        result = await db.execute(select(Pricing))
        existing_pricings: dict[tuple, Pricing] = {}
        for p in result.scalars().all():
            key = (
                p.client_id,
                p.work_type,
                p.pickup_location_id,
                p.dropoff_location_id,
            )
            existing_pricings[key] = p

        created_pricings = 0
        skipped_no_loc = 0
        skipped_no_price = 0
        skipped_existing = 0

        for row in pricing_rows:
            client_code = row["client_name"].upper().replace(" ", "_")
            client = client_map.get(client_code)
            if not client:
                print(f"  ⚠ Skip: client_name '{row['client_name']}' not found")
                continue

            pickup_name = row["pickup"]
            dropoff_name = row["dropoff"]
            pickup_loc = loc_map.get(pickup_name)
            dropoff_loc = loc_map.get(dropoff_name)

            if not pickup_loc or not dropoff_loc:
                skipped_no_loc += 1
                print(f"  ⚠ Skip: location not found — {pickup_name} → {dropoff_name}")
                continue

            work_type = row.get("work_type", "CHUYỂN BÃI")
            price = row.get("price")
            if price is None or price == 0:
                skipped_no_price += 1
                continue

            dedup_key = (
                client.id,
                work_type,
                pickup_loc.id,
                dropoff_loc.id,
            )

            if dedup_key in existing_pricings:
                skipped_existing += 1
                continue

            pricing = Pricing(
                client_id=client.id,
                work_type=work_type,
                pickup_location_id=pickup_loc.id,
                dropoff_location_id=dropoff_loc.id,
                is_active=True,
            )
            db.add(pricing)
            await db.flush()

            db.add(PricingLine(
                pricing_id=pricing.id,
                quantity=1,
                unit_price=price,
                driver_salary=0,
                allowance=0,
            ))
            await db.flush()

            existing_pricings[dedup_key] = pricing
            created_pricings += 1
            print(f"  + {client.code}: {pickup_name} → {dropoff_name} {work_type} = {price:,}")

        await db.commit()

        # ── Summary ─────────────────────────────────────────────────────
        print("\n" + "=" * 60)
        print("SEED COMPLETE — Ports & Pricing from Excel")
        print("=" * 60)
        print(f"  Locations:  {created_locs} created")
        print(f"  Partners:   {created_partners} created")
        print(f"  Pricings:   {created_pricings} created")
        print(f"  Skipped:    {skipped_existing} existing, {skipped_no_loc} missing location, {skipped_no_price} no price")


if __name__ == "__main__":
    asyncio.run(seed_ports_pricing())
