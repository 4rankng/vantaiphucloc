"""Shared Excel parsing and generation utilities for reconciliation flows."""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Constants ──

TEMPLATE_VERSION = "1.0"

# Standard header styles
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Date formats used across all Excel files
DATE_FORMATS = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"]

# Container number pattern
CONTAINER_RE = re.compile(r"\b[A-Z]{4}\d{7}\b")


def parse_date(raw: Any) -> date | None:
    """Parse a date value from Excel cell (handles various formats)."""
    if raw is None:
        return None
    if isinstance(raw, (date, datetime)):
        return raw.date() if isinstance(raw, datetime) else raw
    s = str(raw).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_amount(raw: Any) -> int | None:
    """Parse an amount value (strips non-digit chars, returns integer VND)."""
    if raw is None:
        return None
    s = re.sub(r"[^\d]", "", str(raw))
    return int(s) if s else None


def looks_like_container(val: Any) -> bool:
    """Check if a value looks like a container number."""
    if val is None:
        return False
    s = str(val).strip().upper()
    return bool(CONTAINER_RE.match(s))


def detect_column_mapping(headers: Sequence[str], expected_fields: dict[str, list[str]]) -> dict[str, int]:
    """Auto-detect column indices by matching headers against expected field names.

    Args:
        headers: List of header strings from the Excel file
        expected_fields: {field_name: [possible_header_names]}

    Returns:
        {field_name: column_index}
    """
    mapping = {}
    normalized_headers = [str(h).strip().lower() for h in headers]

    for field_name, aliases in expected_fields.items():
        for i, header in enumerate(normalized_headers):
            if any(alias.lower() in header for alias in aliases):
                mapping[field_name] = i
                break
    return mapping


def styled_workbook(title: str, headers: list[str]) -> tuple:
    """Create a styled workbook with header row.

    Returns: (workbook, worksheet, current_row)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    # Column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    return wb, ws, 2


def add_template_version(ws, col: int):
    """Add template version metadata to a hidden metadata cell beyond data columns."""
    version_cell = ws.cell(row=1, column=col + 1, value=f"TEMPLATE_V{TEMPLATE_VERSION}")
    version_cell.font = Font(color="D0D0D0", size=8)  # Light gray, tiny


def get_template_version(ws) -> str | None:
    """Extract template version from metadata cell."""
    for col in range(20, 30):  # Check columns T-AC
        val = ws.cell(row=1, column=col).value
        if val and str(val).startswith("TEMPLATE_V"):
            return str(val).replace("TEMPLATE_V", "")
    return None


def apply_header_style(
    ws,
    row: int,
    num_cols: int,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
) -> None:
    """Apply standard header styling to a row of cells.

    Defaults to the blue header style (white bold text on #4472C4 fill, centered).
    """
    _font = font or _HEADER_FONT
    _fill = fill or _HEADER_FILL
    _align = alignment or Alignment(horizontal="center")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _font
        cell.fill = _fill
        cell.alignment = _align


def auto_fit_columns(ws, max_width: int = 40) -> None:
    """Auto-size column widths based on cell content."""
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def workbook_to_bytes(wb) -> bytes:
    """Save workbook to BytesIO and return bytes."""
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def flatten_complex_sheet(raw_rows: list[tuple]) -> list[tuple]:
    """Detect and flatten side-by-side tables into a single flat table.
    
    If the sheet has multiple occurrences of a key header (like "container" or "số cont")
    on the same row, it slices the row into multiple sub-rows vertically and
    concatenates them.
    """
    if not raw_rows:
        return []
        
    header_idx = -1
    container_cols = []

    for i, row in enumerate(raw_rows[:20]):
        cols_found = []
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip().lower()

            # Exact matches — "container" alone (not "contact" etc.)
            if cell_str == "container":
                cols_found.append(col_idx)
                continue

            # Substring matches for more specific Vietnamese terms
            if any(pat in cell_str for pat in ["số cont", "số container", "container no", "số côn", "cont no"]):
                cols_found.append(col_idx)
        
        if len(cols_found) > 1:
            header_idx = i
            container_cols = cols_found
            break
            
    if header_idx == -1 or len(container_cols) <= 1:
        return raw_rows
        
    header_row = raw_rows[header_idx]
    headers = [str(c).strip().lower() if c is not None else "" for c in header_row]
    
    table_bounds = []
    for j, c_col in enumerate(container_cols):
        start = c_col
        prev_end = table_bounds[-1][1] if table_bounds else -1
        
        while start - 1 > prev_end and headers[start - 1] != "":
            start -= 1
            
        end = c_col
        limit = container_cols[j+1] if j < len(container_cols) - 1 else len(headers)
        while end + 1 < limit and headers[end + 1] != "":
            end += 1
            
        table_bounds.append((start, end))
        
    max_width = max(end - start + 1 for start, end in table_bounds)
    
    flattened = raw_rows[:header_idx]
    
    first_table_start, first_table_end = table_bounds[0]
    base_header_slice = list(raw_rows[header_idx][first_table_start:first_table_end + 1])
    while len(base_header_slice) < max_width:
        base_header_slice.append(None)
    flattened.append(tuple(base_header_slice))
    
    for row in raw_rows[header_idx + 1:]:
        for start, end in table_bounds:
            if start < len(row):
                slice_len = min(end - start + 1, len(row) - start)
                data_slice = list(row[start:start + slice_len])
                
                if any(c is not None and str(c).strip() != "" for c in data_slice):
                    while len(data_slice) < max_width:
                        data_slice.append(None)
                    flattened.append(tuple(data_slice))
                    
    return flattened
