"""AI file parsing pipeline orchestrator.

Coordinates the 5 stages: Sniff → Cache → Apply → Cleanup → Preview
Also provides template-based parsing for customer Excel files.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app.ai.parser import (
    ColumnMapping,
    ParseResult,
    ParsedCell,
    ParsedRow,
    _compute_file_hash,
    sniff_columns,
    CANONICAL_FIELDS,
)
from app.ai.mapping_cache import get_cached_mapping, save_mapping
from app.ai.gemini_client import call_gemini_row_cleanup, estimate_cost

_logger = logging.getLogger(__name__)

SAMPLE_ROWS = 20  # Rows to send to LLM for sniffing
PREVIEW_ROWS = 20  # Rows to show in preview

# ---------------------------------------------------------------------------
# Template-based Excel parsing (no AI)
# ---------------------------------------------------------------------------

# Mapping from raw header text to display name.
# None = skip column. Keys are stripped + uppercased for matching.
_TEMPLATE_HEADER_MAP: dict[str, str | None] = {
    "STT": None,
    "NGÀY ĐI": "Ngày đi",
    "CHỦ HÀNG": "Chủ hàng",
    "SỐCONTAINER": "Số Cont",
    "SỐ CONTAINER": "Số Cont",
    "SỐ CONT": "Số Cont",
    "F20'": "F20",
    "F40'": "F40",
    "E20'": "E20",
    "E40'": "E40",
    "SỐ XE CHẠY": "Số xe chạy",
    "SỐXE CHẠY": "Số xe chạy",
    "SỐ XE\nCHẠY": "Số xe chạy",
    "SỐ XE VC": "Số xe chạy",
    "BIỂN SỐ XE": "Số xe chạy",
    "BIỂN SỐ": "Số xe chạy",
    "XE CHẠY": "Số xe chạy",
    "SỐ XE": "Số xe chạy",
    "ĐIỂM ĐI": "Điểm đi",
    "ĐIỂM ĐẾN": "Điểm đến",
    "SỐ CHUYẾN": None,  # Skip — always 1 in system, never stored as 0.5
    "CƯỚC CHUYẾN": "Cước chuyến",
    "CƯỚC\nCHUYẾN": "Cước chuyến",
    "TỔNG TT": "Tổng TT",
    "TÁC NGHIỆP": "Tác Nghiệp",
    "GHI CHÚ": None,
    "GHI\nCHÚ": None,
}


def _clean_value(val: str) -> str:
    """Strip leading/trailing whitespace, dots, commas, semicolons."""
    return re.sub(r'^[\s,.;:]+|[\s,.;:]+$', '', val)


def _detect_duplicates(rows: list[dict[str, Any]], cont_key: str = "Số Cont") -> tuple[list[DuplicateGroup], list[str]]:
    """Detect exact duplicate containers in parsed rows.

    "Exact" means ALL column values match across rows, not just the container number.
    Rows with the same container but different dates/locations/amounts are NOT duplicates.

    Returns (duplicate_groups, warnings).
    """
    groups: list[DuplicateGroup] = []
    warnings: list[str] = []

    indexed: list[tuple[int, str]] = []
    for i, row in enumerate(rows):
        cn = str(row.get(cont_key) or "").strip().upper().replace(" ", "").replace("-", "")
        if cn:
            indexed.append((i, cn))

    if len(indexed) < 2:
        return groups, warnings

    def _row_signature(row: dict[str, Any]) -> str:
        """Normalized string of all cell values for full-row comparison."""
        parts = []
        for k, v in row.items():
            s = str(v).strip().upper().replace(" ", "").replace("-", "") if v is not None else ""
            parts.append(f"{k}={s}")
        return "|".join(parts)

    sig_groups: dict[tuple[str, str], list[int]] = {}
    for i, cn in indexed:
        sig = _row_signature(rows[i])
        sig_groups.setdefault((cn, sig), []).append(i)

    for (cn, sig), indices in sig_groups.items():
        if len(indices) > 1:
            groups.append(DuplicateGroup(
                type="exact",
                row_indices=indices,
                containers=[cn],
                message=f"Cont {cn} xuất hiện {len(indices)} lần (dòng {', '.join(str(j + 1) for j in indices)})",
            ))

    exact_count = len(groups)
    if exact_count:
        warnings.append(f"Tìm thấy {exact_count} nhóm cont trùng nhau")

    return groups, warnings


def _normalise_header(raw: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", raw.strip())).upper()


@dataclass
class DuplicateGroup:
    """A cluster of rows with similar container numbers."""
    type: str  # "exact"
    row_indices: list[int]  # 0-based index into rows list
    containers: list[str]
    message: str


@dataclass
class TemplateParseResult:
    filename: str
    sheet_name: str
    total_rows: int
    columns: list[str]
    rows: list[dict[str, Any]]
    duplicate_groups: list[DuplicateGroup] | None = None
    warnings: list[str] | None = None


def _find_template_sheet(wb) -> tuple[Any, int, dict[int, str]] | None:
    """Scan all sheets for the header row matching the template structure.

    Returns (worksheet, header_idx, col_map) or None if no match.
    """
    # Key headers that MUST be present to identify the template
    REQUIRED_HEADERS = {"NGÀY ĐI", "SỐCONTAINER", "SỐ CONTAINER", "SỐ CONT",
                        "ĐIỂM ĐI", "ĐIỂM ĐẾN", "CƯỚC CHUYẾN"}

    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(max_row=30, values_only=True):
            rows.append(list(row))

        for header_idx in range(min(len(rows), 20)):
            raw_headers = [_normalise_header(str(c)) if c is not None else "" for c in rows[header_idx]]
            # Check if enough required headers are present
            matched_required = sum(1 for h in REQUIRED_HEADERS if h in raw_headers)
            if matched_required < 3:
                continue

            # Build column map
            col_map: dict[int, str] = {}
            for idx, raw_h in enumerate(raw_headers):
                display = _TEMPLATE_HEADER_MAP.get(raw_h)
                if display is None:
                    continue
                col_map[idx] = display

            if len(col_map) >= 5:
                return ws, header_idx, col_map

    return None


def parse_template_excel(file: BinaryIO, filename: str) -> TemplateParseResult:
    """Parse a customer-template Excel file by detecting the header structure."""
    wb = load_workbook(file, read_only=True, data_only=True)

    result = _find_template_sheet(wb)
    if result is None:
        raise ValueError(
            "Không tìm thấy sheet hợp lệ trong file. "
            "File phải có các cột: Ngày đi, Số Container, Điểm đi, Điểm đến, Cước chuyến..."
        )

    ws, header_idx, col_map = result
    columns = list(col_map.values())

    # Read all rows from the matched sheet
    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    data_start = header_idx + 2  # skip header row + summary row

    # Parse data rows
    rows: list[dict[str, Any]] = []
    for row in all_rows[data_start:]:
        # Stop at summary/end
        first_val = row[0] if row else None
        if first_val is None or str(first_val).strip() == "" or str(first_val).strip().upper().startswith("CỘNG"):
            break

        record: dict[str, Any] = {}
        all_empty = True
        for col_idx, display_name in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, datetime):
                val = val.strftime("%d/%m/%Y")
            elif val is not None:
                val = _clean_value(str(val)) or None
            if val is not None:
                all_empty = False
            record[display_name] = val

        if all_empty:
            continue
        rows.append(record)

    # Derive "Loại Cont" from E20/F20/E40/F40 columns
    _WORK_TYPE_KEYS = {"E20", "F20", "E40", "F40"}
    for record in rows:
        cont_type = None
        for key in _WORK_TYPE_KEYS:
            if record.get(key) is not None:
                cont_type = key
                break
        record["Loại Cont"] = cont_type
        for key in _WORK_TYPE_KEYS:
            record.pop(key, None)

    columns = [c for c in columns if c not in _WORK_TYPE_KEYS]
    cont_idx = columns.index("Số Cont") + 1 if "Số Cont" in columns else len(columns)
    columns.insert(cont_idx, "Loại Cont")

    # Detect duplicate / near-duplicate containers
    dup_groups, dup_warnings = _detect_duplicates(rows)

    return TemplateParseResult(
        filename=filename,
        sheet_name=ws.title,
        total_rows=len(rows),
        columns=columns,
        rows=rows,
        duplicate_groups=dup_groups,
        warnings=dup_warnings,
    )


async def parse_file(
    file: BinaryIO,
    filename: str,
    source_id: str | None = None,
    max_rows: int = 1000,
) -> ParseResult:
    """Parse an arbitrary-format Excel file through the 5-stage pipeline.

    Args:
        file: File-like object (Excel .xlsx)
        filename: Original filename for reference
        source_id: Customer/vendor ID for caching
        max_rows: Maximum rows to process

    Returns:
        ParseResult with parsed rows and metadata
    """
    # Load workbook
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    # Extract all rows as lists
    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(max_row=max_rows + 1, values_only=True):
        all_rows.append([str(c) if c is not None else "" for c in row])

    if not all_rows:
        raise ValueError("File is empty")

    # Stage 1 + 2: Sniff with cache
    sample_rows = all_rows[:SAMPLE_ROWS]
    file_hash = _compute_file_hash(all_rows)

    cached = get_cached_mapping(source_id, file_hash)
    sniff_cost = 0.0

    if cached:
        mapping = ColumnMapping(
            header_row=cached.get("header_row", 0),
            mapping={int(k): v for k, v in cached.get("mapping", {}).items()},
            confidence=cached.get("confidence", 0.5),
        )
        _logger.info(f"Using cached mapping for {filename}")
    else:
        mapping = await sniff_columns(sample_rows, source_id)
        sniff_cost = estimate_cost(
            input_tokens=len(str(sample_rows)) // 4,  # Rough estimate
            output_tokens=200,
        )
        # Stage 2: Cache the mapping
        if mapping.confidence > 0.3 and mapping.mapping:
            save_mapping(source_id, file_hash, {
                "header_row": mapping.header_row,
                "mapping": mapping.mapping,
                "confidence": mapping.confidence,
            })

    if not mapping.mapping:
        raise ValueError(
            f"Could not detect column mapping from file. "
            f"Confidence: {mapping.confidence}. "
            f"Try a file with clearer headers."
        )

    # Stage 3: Apply mapping (programmatic, fast)
    data_start = mapping.header_row + 1
    data_rows = all_rows[data_start:data_start + max_rows]

    parsed_rows: list[ParsedRow] = []
    rows_needing_cleanup: list[tuple[int, ParsedRow, list[str]]] = []

    for i, raw_row in enumerate(data_rows):
        row_num = data_start + i + 1  # 1-indexed
        cells: dict[str, ParsedCell] = {}
        issues: list[str] = []

        for col_idx, field_name in mapping.mapping.items():
            if col_idx >= len(raw_row):
                continue

            raw_value = raw_row[col_idx]
            cleaned_value, confidence, was_cleaned = _apply_programmatic(field_name, raw_value)

            cells[field_name] = ParsedCell(
                value=cleaned_value,
                confidence=confidence,
                original_value=raw_value if was_cleaned else None,
                cleaned=was_cleaned,
            )

            if confidence < 0.5:
                issues.append(f"{field_name}: uncertain value '{raw_value}'")

        # Add source reference
        cells["source_row_ref"] = ParsedCell(
            value=f"{filename}:R{row_num}",
            confidence=1.0,
        )

        parsed_row = ParsedRow(
            row_number=row_num,
            cells=cells,
            source_row_ref=f"{filename}:R{row_num}",
        )

        parsed_rows.append(parsed_row)

        # Stage 4 gate: only send to LLM if multiple issues
        if len(issues) >= 2:
            row_data = {k: str(v.value) for k, v in cells.items() if k != "source_row_ref"}
            rows_needing_cleanup.append((i, parsed_row, issues))

    # Stage 4: Cleanup problematic rows via LLM (limited)
    MAX_CLEANUP_ROWS = 10  # Cost control
    for idx, parsed_row, issues in rows_needing_cleanup[:MAX_CLEANUP_ROWS]:
        try:
            row_data = {k: str(v.value) for k, v in parsed_row.cells.items() if k != "source_row_ref"}
            cleaned = await call_gemini_row_cleanup(row_data, issues)

            for field_name, cleaned_value in cleaned.items():
                if field_name in parsed_row.cells:
                    parsed_row.cells[field_name].value = cleaned_value
                    parsed_row.cells[field_name].cleaned = True
                    parsed_row.cells[field_name].confidence = 0.7
        except Exception as e:
            _logger.warning(f"Row cleanup failed for row {parsed_row.row_number}: {e}")

    # Calculate total cost
    total_cost = sniff_cost
    if rows_needing_cleanup:
        cleanup_cost = estimate_cost(
            input_tokens=len(rows_needing_cleanup) * 100,
            output_tokens=len(rows_needing_cleanup) * 50,
        )
        total_cost += cleanup_cost

    _logger.info(f"Parsed {len(parsed_rows)} rows, cost estimate: ${total_cost:.4f}")

    return ParseResult(
        column_mapping=mapping,
        rows=parsed_rows,
        total_rows=len(parsed_rows),
        cached_mapping=bool(cached),
        sniff_cost_estimate=total_cost,
    )


def _apply_programmatic(field_name: str, raw_value: str) -> tuple[Any, float, bool]:
    """Stage 3: Apply programmatic cleaning to a cell value.

    Returns:
        (cleaned_value, confidence, was_cleaned)
    """
    if not raw_value or not raw_value.strip():
        return None, 1.0, False

    value = raw_value.strip()

    if field_name == "date":
        # Try common date formats
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"]:
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime("%Y-%m-%d"), 1.0, fmt != "%Y-%m-%d"
            except ValueError:
                continue
        return value, 0.3, False

    elif field_name == "container_number":
        cleaned = value.upper().replace(" ", "").replace("-", "")
        if re.match(r"^[A-Z]{4}\d{7}$", cleaned):
            return cleaned, 1.0, cleaned != value.upper()
        return value, 0.3, False

    elif field_name == "amount":
        # Remove currency suffixes and formatting
        cleaned = re.sub(r"[^\d]", "", value)
        if cleaned and cleaned.isdigit():
            return int(cleaned), 1.0, cleaned != value
        return value, 0.3, False

    elif field_name in ("route_from", "route_to"):
        return value, 0.9, False

    elif field_name in ("customer_name", "vendor_name", "driver_name"):
        return value, 0.9, False

    elif field_name == "vehicle_plate":
        cleaned = value.upper().replace(" ", "").replace("-", "")
        return cleaned, 0.9, cleaned != value

    else:
        return value, 0.9, False
