"""Location import from Excel service.

Parses location names from port list Excel files (e.g., "DS PORT+CƯỚC TUYẾN.xlsx").
The expected structure:
- Sheet "DS PORT" with location names in columns
- First row may contain group names (can be ignored)
- Subsequent rows contain location names
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.contexts.customer_pricing.application import (
    CreateLocation,
    ListAllActiveLocations,
)
from app.contexts.customer_pricing.application.dto import LocationCreateInput
from app.contexts.customer_pricing.domain.exceptions import AlreadyExists
from app.contexts.customer_pricing.infrastructure.location_resolver import (
    normalize as _normalize,
)
from app.contexts.customer_pricing.infrastructure.repositories import (
    SqlLocationRepository,
)

_logger = logging.getLogger(__name__)


@dataclass
class LocationImportRow:
    """A single location name parsed from Excel."""

    name: str
    row: int
    column: int


@dataclass
class LocationPreviewResult:
    """Result of parsing location Excel file."""

    filename: str
    sheet_name: str
    rows: list[LocationImportRow]
    total_count: int
    duplicate_names: list[str]  # Names that appear multiple times in file
    already_exist: list[str]  # Names that already exist in DB
    new_names: list[str]  # Names that would be created


@dataclass
class LocationCommitResult:
    """Result of committing location import."""

    created: int
    skipped_existing: int
    errors: list[str]


def _score_sheet_for_locations(sheet) -> tuple[int, str]:
    """Score a sheet based on how likely it contains location data.

    Returns (score, reason) where higher score = more likely to be location sheet.

    Heuristics:
    - High number of non-empty string cells (locations are text-heavy)
    - Low number of numeric columns (not pricing/data tables)
    - Presence of common location keywords (PORT, BÃI, KHO, ICD, etc.)
    - First row looks like headers/groups, not data
    """
    score = 0
    reasons = []

    # Sample first 20 rows to analyze structure
    max_sample = min(20, sheet.max_row)
    string_cells = 0
    numeric_cells = 0
    empty_cells = 0
    total_cells = 0

    location_keywords = [
        "PORT",
        "BÃI",
        "BAY",
        "KHO",
        "ICD",
        "HẢI",
        "PHÀ",
        "BẾN",
        "CẢNG",
        "TERMINAL",
    ]

    for row in sheet.iter_rows(min_row=1, max_row=max_sample, values_only=True):
        for cell in row:
            total_cells += 1
            if cell is None:
                empty_cells += 1
                continue

            if isinstance(cell, str):
                val = cell.strip().upper()
                if val:
                    string_cells += 1
                    # Check for location keywords
                    for keyword in location_keywords:
                        if keyword in val:
                            score += 2
                            if "keyword" not in reasons:
                                reasons.append("keyword")
                            break
            elif isinstance(cell, (int, float)):
                # Check if it's a row number (small integer in first column)
                numeric_cells += 1

    # High ratio of string cells suggests location list
    if total_cells > 0:
        string_ratio = string_cells / total_cells
        if string_ratio > 0.6:
            score += 30
            reasons.append("high_string_ratio")
        elif string_ratio > 0.4:
            score += 15
            reasons.append("moderate_string_ratio")

    # Low numeric cells (excluding row numbers) suggests not a pricing table
    if numeric_cells < string_cells * 0.3:
        score += 20
        reasons.append("low_numeric")

    # Penalize if sheet looks like a pricing table (many numeric columns)
    # Pricing tables typically have row numbers in col 1 and many numeric columns
    if sheet.max_column > 5:
        # Check if row 2+ has many numeric values
        numeric_cols = 0
        for row in sheet.iter_rows(
            min_row=2, max_row=min(5, sheet.max_row), values_only=True
        ):
            row_numeric = sum(
                1 for c in row[1:] if isinstance(c, (int, float)) and c > 1000
            )
            numeric_cols += row_numeric
        if numeric_cols > 10:
            score -= 40
            reasons.append("looks_like_pricing_table")

    # Prefer sheets with more rows (location lists are usually long)
    if sheet.max_row > 30:
        score += 10
        reasons.append("many_rows")

    return score, ", ".join(reasons)


def parse_location_excel(content: bytes, filename: str) -> LocationPreviewResult:
    """Parse location names from Excel file.

    Expected format:
    - Analyzes all sheets to find the one most likely containing location data
    - Looks for sheets with high text content and location-related keywords
    - Reads all non-empty cells as location names
    - Skips header row if it looks like a group name

    Args:
        content: Excel file bytes
        filename: Original filename (for reporting)

    Returns:
        LocationPreviewResult with parsed locations
    """
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Không thể đọc file Excel: {e}")

    # Score all sheets to find the best match
    sheet_scores = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        score, reason = _score_sheet_for_locations(sheet)
        sheet_scores.append((score, sheet, sheet_name, reason))
        _logger.info(f"Sheet '{sheet_name}': score={score}, reason={reason}")

    # Sort by score descending
    sheet_scores.sort(key=lambda x: x[0], reverse=True)

    if not sheet_scores:
        raise ValueError("File không có sheet nào")

    # Use the highest-scoring sheet
    best_score, sheet, sheet_name, score_reason = sheet_scores[0]

    if best_score < 0:
        _logger.warning(f"Best sheet '{sheet_name}' has negative score: {score_reason}")

    _logger.info(
        f"Selected sheet: {sheet_name} (score: {best_score}, reason: {score_reason})"
    )

    rows: list[LocationImportRow] = []
    seen_names: dict[str, list[tuple[int, int]]] = {}  # Track duplicates

    # Determine if first row is header (contains common group name patterns)
    first_row_values = []
    for cell in sheet[1]:
        if cell.value and isinstance(cell.value, str):
            val = cell.value.strip()
            if val and val != "Unnamed: 0":
                first_row_values.append(val.upper())

    # Skip first row if it looks like a header/group name
    header_patterns = ["PORT", "BÃI", "BAY", "KHO", "ICD", "HẢI PHÒNG"]
    skip_first_row = any(p in " ".join(first_row_values) for p in header_patterns)

    start_row = 2 if skip_first_row else 1

    # Parse all cells
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=start_row, values_only=True), start=start_row
    ):
        for col_idx, cell_value in enumerate(row, start=1):
            if cell_value and isinstance(cell_value, str):
                name = cell_value.strip()
                # Skip empty values and common headers
                if name and name.lower() not in ["nan", "none", ""]:
                    # Skip first column if it's just row numbers
                    if col_idx == 1 and name.isdigit():
                        continue

                    # Normalize for duplicate detection
                    normalized = _normalize(name)

                    row_entry = LocationImportRow(
                        name=name, row=row_idx, column=col_idx
                    )
                    rows.append(row_entry)

                    # Track duplicates
                    if normalized not in seen_names:
                        seen_names[normalized] = []
                    seen_names[normalized].append((row_idx, col_idx))

    # Find duplicates (same normalized name appears more than once)
    duplicate_names = []
    for norm, positions in seen_names.items():
        if len(positions) > 1:
            # Get the original name from the first occurrence
            for r in rows:
                if _normalize(r.name) == norm:
                    duplicate_names.append(r.name)
                    break

    return LocationPreviewResult(
        filename=filename,
        sheet_name=sheet.title,
        rows=rows,
        total_count=len(rows),
        duplicate_names=duplicate_names,
        already_exist=[],
        new_names=[r.name for r in rows],
    )


async def preview_location_import(
    db: AsyncSession,
    content: bytes,
    filename: str,
) -> LocationPreviewResult:
    """Preview location import, checking for existing locations."""
    result = parse_location_excel(content, filename)

    # Check which names already exist
    list_use_case = ListAllActiveLocations(SqlLocationRepository(db))
    existing_locations = await list_use_case(limit=100000)
    existing_normalized = {_normalize(loc.name): loc.name for loc in existing_locations}

    already_exist: set[str] = set()
    new_names: list[str] = []

    for row in result.rows:
        norm = _normalize(row.name)
        if norm in existing_normalized:
            already_exist.add(row.name)
        else:
            new_names.append(row.name)

    result.already_exist = sorted(list(already_exist))
    result.new_names = sorted(set(new_names))

    return result


async def commit_location_import(
    db: AsyncSession,
    create_use_case: CreateLocation,
    names: list[str],
) -> LocationCommitResult:
    """Commit location import, creating new locations."""
    created = 0
    skipped_existing = 0
    errors: list[str] = []

    # Get existing locations to check against
    list_use_case = ListAllActiveLocations(SqlLocationRepository(db))
    existing_locations = await list_use_case(limit=100000)
    existing_normalized = {_normalize(loc.name): loc.name for loc in existing_locations}

    for name in names:
        norm = _normalize(name)

        if norm in existing_normalized:
            skipped_existing += 1
            continue

        try:
            await create_use_case(LocationCreateInput(name=name))
            created += 1
            existing_normalized[norm] = (
                name  # Add to set to avoid duplicates within batch
            )
        except AlreadyExists:
            skipped_existing += 1
        except Exception as e:
            errors.append(f"{name}: {e}")
            _logger.error(f"Error creating location '{name}': {e}")

    return LocationCommitResult(
        created=created,
        skipped_existing=skipped_existing,
        errors=errors,
    )
