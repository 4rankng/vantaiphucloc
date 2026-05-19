"""Vendor (xe ngoài) reconciliation endpoints.

Flow:
  1. POST /vendor-reconciliation/upload       — multipart, parse vendor Excel
  2. GET  /vendor-reconciliation/             — list imports (optional vendor_id filter)
  3. GET  /vendor-reconciliation/{import_id}  — import header + all rows
  4. PATCH /vendor-reconciliation/{import_id}/rows/{row_id} — update row verdict
  5. POST /vendor-reconciliation/{import_id}/apply — apply import (write back vendor_amount)
  6. DELETE /vendor-reconciliation/{import_id}     — discard import

The parser is intentionally simple: it walks the Excel rows looking for a
container-number-shaped column (pattern ABCU1234567 / 11 chars) and pulls
adjacent columns for date and amount.  Each vendor tends to have a
consistent layout so this heuristic covers most real files; the reviewer
can still correct any mismatches in the review UI before applying.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime, timezone
from typing import Any

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import require_roles
from app.database import get_db
from app.models.base import User
from app.utils.excel_utils import (
    CONTAINER_RE,
    add_template_version,
    looks_like_container,
    parse_amount,
    parse_date,
)
from app.models.domain import (
    Location,
    Client,
    Vendor,
    Vehicle,
    VehicleDriver,
    VendorReconciliationImport,
    VendorReconciliationRow,
    DeliveredTrip,
    DeliveredTripContainer,
)

_logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/vendor-reconciliation",
    tags=["vendor-reconciliation"],
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Parsing helpers are imported from app.utils.excel_utils


def _parse_vendor_excel(content: bytes, filename: str) -> list[dict]:
    """Parse a vendor Excel file into a list of raw row dicts.

    Strategy:
      1. Load the workbook via openpyxl.
      2. For each sheet, scan for the first row that contains a container-
         number-shaped cell (ABCU1234567).
      3. For that sheet, extract: container_number, trip_date, vendor_amount,
         work_type (inferred from container prefix if available), route_text
         (concatenate other string columns).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel parsing")

    import io
    from app.utils.excel_utils import get_template_version, TEMPLATE_VERSION
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    parsed: list[dict] = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Find a row containing a container number to start data extraction
        data_start = None
        for idx, row in enumerate(rows):
            if any(looks_like_container(cell) for cell in row):
                data_start = idx
                break

        if data_start is None:
            continue

        # Heuristic: try to identify column indices for container, date, amount
        # by inspecting the first data row and building a simple signature.
        for row in rows[data_start:]:
            cells = [c for c in row]
            if not any(looks_like_container(c) for c in cells):
                continue

            container = None
            trip_date = None
            vendor_amount = None
            other_texts: list[str] = []

            for c in cells:
                if c is None:
                    continue
                if looks_like_container(c) and container is None:
                    m = CONTAINER_RE.search(str(c).upper().replace(" ", ""))
                    if m:
                        container = m.group(0)
                elif isinstance(c, (date, datetime)) and trip_date is None:
                    trip_date = parse_date(c)
                elif isinstance(c, (int, float)) and vendor_amount is None:
                    if c > 0:
                        vendor_amount = int(c)
                elif isinstance(c, str):
                    stripped = c.strip()
                    # Try as date string first
                    d = parse_date(stripped)
                    if d and trip_date is None:
                        trip_date = d
                    elif stripped and stripped not in (container or ""):
                        other_texts.append(stripped)

            if container:
                route_text = " | ".join(other_texts[:3]) if other_texts else None
                # Infer work_type from container if possible (not typically in vendor files)
                parsed.append(
                    {
                        "container_number": container,
                        "work_type": None,
                        "route_text": route_text,
                        "trip_date": trip_date.isoformat() if trip_date else None,
                        "vendor_amount": vendor_amount,
                    }
                )
        break  # Use first sheet that has data

    wb.close()
    return parsed


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class VendorReconRowOut(BaseModel):
    id: int
    import_id: int
    container_number: str | None
    work_type: str | None
    route_text: str | None
    trip_date: date | None
    vendor_amount: int | None
    match_status: str
    matched_delivered_trip_id: int | None
    reviewer_note: str | None


class VendorReconImportOut(BaseModel):
    id: int
    vendor_id: int
    vendor_partner_name: str
    period_from: date
    period_to: date
    source_filename: str | None
    status: str
    totals: dict | None
    notes: str | None
    uploaded_at: datetime
    uploaded_by: int | None
    applied_at: datetime | None
    applied_by: int | None
    rows: list[VendorReconRowOut] = Field(default_factory=list)


class RowUpdateBody(BaseModel):
    match_status: str | None = None
    reviewer_note: str | None = None
    matched_delivered_trip_id: int | None = None
    vendor_amount: int | None = None


class ApplyResponse(BaseModel):
    applied: int
    skipped: int


class OurOnlyRowOut(BaseModel):
    delivered_trip_id: int
    container_number: str
    trip_date: date | None
    vehicle_plate: str | None


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.get("/export")
async def export_vendor_trips(
    vendor_id: int = Query(..., description="Vendor (nhà xe) ID"),
    date_from: date = Query(..., description="From date (YYYY-MM-DD)"),
    date_to: date = Query(..., description="To date (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_roles("accountant", "superadmin")),
):
    """Export our DeliveredTrips for a specific vendor as Excel.

    This generates a đối soát report (our trips) that can be sent to the vendor
    for review — Mode 4a flow.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.utils.text import slugify_vi

    # Validate vendor
    vendor = (
        await db.execute(select(Vendor).where(Vendor.id == vendor_id))
    ).scalar_one_or_none()
    if vendor is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhà xe.")

    # Load DeliveredTrips for this vendor in the period
    wo_result = await db.execute(
        select(DeliveredTrip).where(
            DeliveredTrip.vendor_id == vendor_id,
            DeliveredTrip.trip_date >= date_from,
            DeliveredTrip.trip_date <= date_to,
        ).order_by(DeliveredTrip.trip_date, DeliveredTrip.id)
    )
    delivered_trips = wo_result.scalars().all()
    wo_ids = [wo.id for wo in delivered_trips]

    # Load containers
    containers_map: dict[int, list] = {}
    if wo_ids:
        cont_result = await db.execute(
            select(DeliveredTripContainer).where(
                DeliveredTripContainer.delivered_trip_id.in_(wo_ids)
            )
        )
        for c in cont_result.scalars().all():
            containers_map.setdefault(c.delivered_trip_id, []).append(c)

    # Load location names
    loc_ids: set[int] = set()
    for wo in delivered_trips:
        if wo.pickup_location_id:
            loc_ids.add(wo.pickup_location_id)
        if wo.dropoff_location_id:
            loc_ids.add(wo.dropoff_location_id)
    loc_name_by_id: dict[int, str] = {}
    if loc_ids:
        loc_result = await db.execute(select(Location).where(Location.id.in_(loc_ids)))
        loc_name_by_id = {loc.id: loc.name for loc in loc_result.scalars().all()}

    # Load vehicle plates via VehicleDriver
    driver_ids = [wo.driver_id for wo in delivered_trips if wo.driver_id]
    driver_plate_map: dict[int, str] = {}
    if driver_ids:
        v_result = await db.execute(
            select(VehicleDriver.driver_id, Vehicle.plate)
            .join(Vehicle, Vehicle.id == VehicleDriver.vehicle_id)
            .where(
                VehicleDriver.driver_id.in_(driver_ids),
                VehicleDriver.is_active == True,  # noqa: E712
                Vehicle.is_active == True,  # noqa: E712
            )
        )
        driver_plate_map = dict(v_result.all())

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = vendor.name[:31]

    # Styles
    thin_side = Side(style="thin", color="BBBBBB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    _blue_dark = "1F4E79"
    _blue_header = "2E75B6"
    _white = "FFFFFF"
    _grey_row = "F5F8FC"
    _blue_light = "DEEAF1"
    _yellow_sum = "FFF2CC"

    num_cols = 12
    last_col = get_column_letter(num_cols)

    # Title rows
    ws.append(["VẬN TẢI PHÚC LỘC", *([""] * (num_cols - 1))])
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].font = Font(bold=True, size=14, color=_blue_dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    month_range = f"Từ {date_from.strftime('%d/%m/%Y')} đến {date_to.strftime('%d/%m/%Y')}"
    ws.append([f"BẢNG CHUYẾN – {vendor.name.upper()} – {month_range}", *([""] * (num_cols - 1))])
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].font = Font(bold=True, size=11, color=_blue_dark)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.append([""] * num_cols)
    ws.row_dimensions[3].height = 6

    # Header row
    _op_labels = {
        "XUAT_NHAP_TAU": "Xuất / Nhập tàu",
        "CHUYEN_BAI": "Chuyển bãi",
        "LAY_VO_HA_HANG": "Lấy vỏ hạ hàng",
        "CHAY_SA_LAN": "Chạy sà lan",
        "DONG_KHO": "Đóng kho",
    }
    WORK_TYPE_FULL = {
        "E20": "Rỗng 20ft", "E40": "Rỗng 40ft",
        "F20": "Hàng 20ft", "F40": "Hàng 40ft",
    }

    headers = [
        "STT", "Mã phiếu", "Ngày chạy", "Số cont", "Loại cont",
        "Điểm lấy", "Điểm trả", "Tác nghiệp", "Biển số xe", "Số tàu",
        "Đơn giá (VNĐ)", "Ghi chú",
    ]
    ws.append(headers)
    header_row = 4
    header_font = Font(bold=True, color=_white, size=10)
    header_fill = PatternFill(start_color=_blue_header, end_color=_blue_header, fill_type="solid")
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=thin_side, right=thin_side,
            top=Side(style="thin", color="BBBBBB"),
            bottom=Side(style="medium", color="1F4E79"),
        )
    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = ws.cell(row=5, column=1)

    # Data rows
    stt = 0
    type_count: dict[str, int] = {}
    total_amount = 0

    for row_idx, wo in enumerate(delivered_trips):
        containers = containers_map.get(wo.id, [])
        pickup = loc_name_by_id.get(wo.pickup_location_id or 0, "")
        dropoff = loc_name_by_id.get(wo.dropoff_location_id or 0, "")
        plate = driver_plate_map.get(wo.driver_id or 0, "")
        vessel = wo.vessel or ""
        op_type = _op_labels.get(wo.operation_type or "", wo.operation_type or "")
        revenue = wo.revenue or 0
        trip_date_str = wo.trip_date.strftime("%d/%m/%Y") if wo.trip_date else ""

        fill_color = _white if row_idx % 2 == 0 else _grey_row
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for c in containers:
            stt += 1
            wt_label = WORK_TYPE_FULL.get(c.work_type or "", c.work_type or "")
            type_count[wt_label] = type_count.get(wt_label, 0) + 1
            total_amount += revenue

            data_row = [
                stt, wo.id, trip_date_str, c.container_number, wt_label,
                pickup, dropoff, op_type, plate, vessel, revenue or "", "",
            ]
            ws.append(data_row)
            data_row_num = ws.max_row

            for col_num in range(1, num_cols + 1):
                cell = ws.cell(row=data_row_num, column=col_num)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if col_num in (1, 2):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_num == 11 and revenue:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[data_row_num].height = 18

    # Summary
    sum_fill = PatternFill(start_color=_blue_light, end_color=_blue_light, fill_type="solid")
    sum_font_bold = Font(bold=True, size=10, color=_blue_dark)

    ws.append([""] * num_cols)
    ws.append(["Tổng hợp theo loại container:", *([""] * (num_cols - 1))])
    label_row = ws.max_row
    ws.merge_cells(f"A{label_row}:{last_col}{label_row}")
    ws[f"A{label_row}"].font = sum_font_bold
    ws[f"A{label_row}"].fill = sum_fill

    for wt_label, count in sorted(type_count.items()):
        ws.append(["", f"  {wt_label}", count, "cont", *([""] * (num_cols - 4))])
        r = ws.max_row
        for col_num in range(1, num_cols + 1):
            ws.cell(row=r, column=col_num).fill = sum_fill
        ws.cell(row=r, column=2).font = Font(size=10)
        ws.cell(row=r, column=3).font = Font(bold=True, size=10, color=_blue_dark)

    total_fill = PatternFill(start_color=_yellow_sum, end_color=_yellow_sum, fill_type="solid")
    total_rows_data = [["Tổng số container:", *([""] * 9), stt, ""]]
    if total_amount:
        total_rows_data.append(["Tổng đơn giá:", *([""] * 9), total_amount, ""])

    for row_data in total_rows_data:
        ws.append(row_data)
        r = ws.max_row
        ws.merge_cells(f"A{r}:J{r}")
        label_cell = ws.cell(row=r, column=1)
        val_cell = ws.cell(row=r, column=11)
        label_cell.font = Font(bold=True, size=10, color=_blue_dark)
        val_cell.font = Font(bold=True, size=11, color=_blue_dark)
        val_cell.number_format = '#,##0'
        val_cell.alignment = Alignment(horizontal="right", vertical="center")
        for col_num in range(1, num_cols + 1):
            ws.cell(row=r, column=col_num).fill = total_fill

    # Column widths
    col_widths = [6, 12, 12, 16, 14, 24, 24, 14, 14, 14, 16, 24]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    add_template_version(ws, num_cols)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    slug = slugify_vi(vendor.name)
    month_str = date_from.strftime("%m-%Y")
    filename = f"DoiSoat_NhaXe_{slug}_{month_str}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/upload")
async def upload_vendor_excel(
    file: UploadFile = File(...),
    vendor_id: int = Form(...),
    period_from: date = Form(...),
    period_to: date = Form(...),
    notes: str | None = Form(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_roles("accountant", "superadmin")),
):
    """Parse a vendor Excel file and create a VendorReconciliationImport with rows."""
    if file.filename is None:
        raise HTTPException(status_code=400, detail="Tệp tải lên không có tên.")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Tệp tải lên rỗng.")

    vendor = (
        await db.execute(select(Vendor).where(Vendor.id == vendor_id))
    ).scalar_one_or_none()
    if vendor is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhà xe.")

    try:
        raw_rows = _parse_vendor_excel(content, file.filename)
    except Exception as exc:
        _logger.exception("Failed to parse vendor Excel: %s", exc)
        raise HTTPException(status_code=422, detail=f"Không thể đọc file Excel: {exc}")

    if not raw_rows:
        raise HTTPException(
            status_code=422,
            detail="Không tìm thấy dữ liệu container trong file. "
                   "Hãy kiểm tra định dạng file (cần có cột số cont dạng ABCU1234567).",
        )

    # Auto-match each row against DeliveredTrips for this vendor in this period
    wo_by_container: dict[str, int] = {}
    wo_rows = (
        await db.execute(
            select(DeliveredTripContainer.container_number, DeliveredTripContainer.delivered_trip_id)
            .join(DeliveredTrip, DeliveredTrip.id == DeliveredTripContainer.delivered_trip_id)
            .where(
                DeliveredTrip.vendor_id == vendor_id,
                DeliveredTrip.trip_date >= period_from,
                DeliveredTrip.trip_date <= period_to,
            )
        )
    ).all()
    for cont_num, wo_id in wo_rows:
        if cont_num:
            wo_by_container[cont_num.upper()] = wo_id

    # Build rows
    orm_rows: list[VendorReconciliationRow] = []
    matched_count = 0
    vendor_only_count = 0

    for r in raw_rows:
        cont = (r["container_number"] or "").upper()
        wo_id = wo_by_container.get(cont)
        if wo_id:
            status = "MATCHED"
            matched_count += 1
        else:
            status = "VENDOR_ONLY"
            vendor_only_count += 1

        trip_date_val = None
        if r["trip_date"]:
            try:
                from datetime import date as _date
                trip_date_val = _date.fromisoformat(r["trip_date"])
            except (ValueError, TypeError):
                pass

        orm_rows.append(
            VendorReconciliationRow(
                container_number=r["container_number"],
                work_type=r["work_type"],
                route_text=r["route_text"],
                trip_date=trip_date_val,
                vendor_amount=r["vendor_amount"],
                match_status=status,
                matched_delivered_trip_id=wo_id,
            )
        )

    # Find OUR_ONLY WOs (we have WO for this vendor+period, vendor didn't claim them)
    claimed_containers = {(r["container_number"] or "").upper() for r in raw_rows}
    our_only_count = 0
    for cont, wo_id in wo_by_container.items():
        if cont not in claimed_containers:
            orm_rows.append(
                VendorReconciliationRow(
                    container_number=cont,
                    work_type=None,
                    route_text=None,
                    trip_date=None,
                    vendor_amount=None,
                    match_status="OUR_ONLY",
                    matched_delivered_trip_id=wo_id,
                )
            )
            our_only_count += 1

    totals = {
        "total": len(raw_rows),
        "matched": matched_count,
        "vendor_only": vendor_only_count,
        "our_only": our_only_count,
        "disputed": 0,
    }

    imp = VendorReconciliationImport(
        vendor_id=vendor_id,
        period_from=period_from,
        period_to=period_to,
        source_filename=file.filename,
        status="PENDING_REVIEW",
        totals=totals,
        notes=notes,
        uploaded_by=user.id,
    )
    db.add(imp)
    await db.flush()

    for row in orm_rows:
        row.import_id = imp.id
        db.add(row)

    await db.commit()
    await db.refresh(imp)

    return {
        "import_id": imp.id,
        "vendor_id": imp.vendor_id,
        "status": imp.status,
        "totals": imp.totals,
        "row_count": len(orm_rows),
    }


@router.get("/")
async def list_imports(
    vendor_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_roles("accountant", "superadmin")),
):
    """List vendor reconciliation imports, optionally filtered by vendor."""
    q = select(VendorReconciliationImport)
    if vendor_id is not None:
        q = q.where(VendorReconciliationImport.vendor_id == vendor_id)
    q = q.order_by(VendorReconciliationImport.uploaded_at.desc())
    imports = (await db.execute(q)).scalars().all()

    # Fetch partner names in bulk
    client_ids = list({imp.vendor_id for imp in imports})
    partners_map: dict[int, str] = {}
    if client_ids:
        partners = (
            await db.execute(select(Vendor).where(Vendor.id.in_(client_ids)))
        ).scalars().all()
        partners_map = {p.id: p.name for p in partners}

    return [
        {
            "id": imp.id,
            "vendor_id": imp.vendor_id,
            "vendor_partner_name": partners_map.get(imp.vendor_id, ""),
            "period_from": imp.period_from,
            "period_to": imp.period_to,
            "source_filename": imp.source_filename,
            "status": imp.status,
            "totals": imp.totals,
            "notes": imp.notes,
            "uploaded_at": imp.uploaded_at,
            "uploaded_by": imp.uploaded_by,
            "applied_at": imp.applied_at,
            "applied_by": imp.applied_by,
        }
        for imp in imports
    ]


@router.get("/{import_id}")
async def get_import(
    import_id: int,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_roles("accountant", "superadmin")),
):
    """Return one import with all its rows."""
    imp = (
        await db.execute(
            select(VendorReconciliationImport).where(
                VendorReconciliationImport.id == import_id
            )
        )
    ).scalar_one_or_none()
    if imp is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy import.")

    vendor = (
        await db.execute(select(Vendor).where(Vendor.id == imp.vendor_id))
    ).scalar_one_or_none()

    rows = (
        await db.execute(
            select(VendorReconciliationRow).where(
                VendorReconciliationRow.import_id == import_id
            )
        )
    ).scalars().all()

    return {
        "id": imp.id,
        "vendor_id": imp.vendor_id,
        "vendor_partner_name": vendor.name if vendor else "",
        "period_from": imp.period_from,
        "period_to": imp.period_to,
        "source_filename": imp.source_filename,
        "status": imp.status,
        "totals": imp.totals,
        "notes": imp.notes,
        "uploaded_at": imp.uploaded_at,
        "uploaded_by": imp.uploaded_by,
        "applied_at": imp.applied_at,
        "applied_by": imp.applied_by,
        "rows": [
            {
                "id": r.id,
                "import_id": r.import_id,
                "container_number": r.container_number,
                "work_type": r.work_type,
                "route_text": r.route_text,
                "trip_date": r.trip_date,
                "vendor_amount": r.vendor_amount,
                "match_status": r.match_status,
                "matched_delivered_trip_id": r.matched_delivered_trip_id,
                "reviewer_note": r.reviewer_note,
            }
            for r in rows
        ],
    }


@router.patch("/{import_id}/rows/{row_id}")
async def update_row(
    import_id: int,
    row_id: int,
    body: RowUpdateBody = Body(...),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_roles("accountant", "superadmin")),
):
    """Update a single row's verdict (match_status, reviewer_note, etc.)."""
    row = (
        await db.execute(
            select(VendorReconciliationRow).where(
                VendorReconciliationRow.id == row_id,
                VendorReconciliationRow.import_id == import_id,
            )
        )
    ).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy dòng.")

    if body.match_status is not None:
        valid_statuses = {"MATCHED", "VENDOR_ONLY", "OUR_ONLY", "DISPUTED", "IGNORED"}
        if body.match_status not in valid_statuses:
            raise HTTPException(
                status_code=400,
                detail=f"match_status không hợp lệ. Phải là một trong: {valid_statuses}",
            )
        row.match_status = body.match_status
    if body.reviewer_note is not None:
        row.reviewer_note = body.reviewer_note
    if body.matched_delivered_trip_id is not None:
        row.matched_delivered_trip_id = body.matched_delivered_trip_id
    if body.vendor_amount is not None:
        row.vendor_amount = body.vendor_amount

    # Recompute parent totals
    all_rows = (
        await db.execute(
            select(VendorReconciliationRow).where(
                VendorReconciliationRow.import_id == import_id
            )
        )
    ).scalars().all()

    totals: dict[str, int] = {
        "total": 0,
        "matched": 0,
        "vendor_only": 0,
        "our_only": 0,
        "disputed": 0,
    }
    for r in all_rows:
        status = body.match_status if r.id == row_id else r.match_status
        totals["total"] += 1
        if status == "MATCHED":
            totals["matched"] += 1
        elif status == "VENDOR_ONLY":
            totals["vendor_only"] += 1
        elif status == "OUR_ONLY":
            totals["our_only"] += 1
        elif status == "DISPUTED":
            totals["disputed"] += 1

    imp = (
        await db.execute(
            select(VendorReconciliationImport).where(
                VendorReconciliationImport.id == import_id
            )
        )
    ).scalar_one_or_none()
    if imp:
        imp.totals = totals

    await db.commit()
    await db.refresh(row)

    return {
        "id": row.id,
        "match_status": row.match_status,
        "reviewer_note": row.reviewer_note,
        "matched_delivered_trip_id": row.matched_delivered_trip_id,
        "vendor_amount": row.vendor_amount,
    }


@router.post("/{import_id}/apply", response_model=ApplyResponse)
async def apply_import(
    import_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_roles("accountant", "superadmin")),
):
    """Apply the import: write vendor_amount back to matched DeliveredTrips.

    Only MATCHED rows with a vendor_amount are written.  Idempotent (re-apply
    overwrites the previously written amount).  Marks the import as APPLIED.
    """
    imp = (
        await db.execute(
            select(VendorReconciliationImport).where(
                VendorReconciliationImport.id == import_id
            )
        )
    ).scalar_one_or_none()
    if imp is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy import.")
    if imp.status == "DISCARDED":
        raise HTTPException(status_code=409, detail="Import đã bị huỷ, không thể áp dụng.")

    rows = (
        await db.execute(
            select(VendorReconciliationRow).where(
                VendorReconciliationRow.import_id == import_id,
                VendorReconciliationRow.match_status == "MATCHED",
            )
        )
    ).scalars().all()

    applied = 0
    skipped = 0
    for row in rows:
        if row.matched_delivered_trip_id is None or row.vendor_amount is None:
            skipped += 1
            continue
        wo = (
            await db.execute(
                select(DeliveredTrip).where(DeliveredTrip.id == row.matched_delivered_trip_id)
            )
        ).scalar_one_or_none()
        if wo is None:
            skipped += 1
            continue
        # Store vendor cost on the DeliveredTrip's driver_salary field as vendor cost proxy.
        # TODO: When a dedicated VendorInvoiceLine table exists, write there instead.
        wo.driver_salary = row.vendor_amount
        applied += 1

    imp.status = "APPLIED"
    imp.applied_at = datetime.now(timezone.utc)
    imp.applied_by = user.id

    await db.commit()
    return ApplyResponse(applied=applied, skipped=skipped)


@router.delete("/{import_id}", status_code=204)
async def discard_import(
    import_id: int,
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_roles("accountant", "superadmin")),
):
    """Mark import as DISCARDED (soft-delete)."""
    imp = (
        await db.execute(
            select(VendorReconciliationImport).where(
                VendorReconciliationImport.id == import_id
            )
        )
    ).scalar_one_or_none()
    if imp is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy import.")
    if imp.status == "APPLIED":
        raise HTTPException(
            status_code=409,
            detail="Import đã áp dụng, không thể huỷ. Liên hệ quản trị viên nếu cần điều chỉnh.",
        )
    imp.status = "DISCARDED"
    await db.commit()


# ---------------------------------------------------------------------------
# Vendor summary endpoint (mounted at /vendors/{vendor_id}/summary)
# ---------------------------------------------------------------------------

# NOTE: This router is mounted at /vendor-reconciliation, so the full path is
# /vendor-reconciliation/vendors/{vendor_id}/summary.  A separate router with
# prefix="" could be used, but co-locating is simpler for now.


@router.get("/vendors/{vendor_id}/summary")
async def vendor_summary(
    vendor_id: int,
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _user: User = Depends(require_roles("accountant", "superadmin")),
):
    vendor = (
        await db.execute(
            select(Vendor).where(
                Vendor.id == vendor_id,
            )
        )
    ).scalar_one_or_none()
    if vendor is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy nhà xe.")

    wo_q = select(DeliveredTrip).where(DeliveredTrip.vendor_id == vendor_id)
    if date_from is not None:
        wo_q = wo_q.where(DeliveredTrip.trip_date >= date_from)
    if date_to is not None:
        wo_q = wo_q.where(DeliveredTrip.trip_date <= date_to)
    delivered_trips = (await db.execute(wo_q)).scalars().all()
    wo_ids = [wo.id for wo in delivered_trips]

    container_map: dict[int, int] = {}
    if wo_ids:
        cont_rows = (
            await db.execute(
                select(
                    DeliveredTripContainer.delivered_trip_id,
                    func.count(DeliveredTripContainer.id),
                )
                .where(DeliveredTripContainer.delivered_trip_id.in_(wo_ids))
                .group_by(DeliveredTripContainer.delivered_trip_id)
            )
        ).all()
        container_map = dict(cont_rows)

    trip_count = len(delivered_trips)
    container_count = sum(container_map.values())
    total_paid = sum(wo.driver_salary or 0 for wo in delivered_trips)
    total_amount = sum(wo.revenue or 0 for wo in delivered_trips)

    plate_groups: dict[str, list[DeliveredTrip]] = {}
    for wo in delivered_trips:
        plate = None
        if not plate:
            continue
        plate_groups.setdefault(plate, []).append(wo)

    drivers = []
    for plate, wos in sorted(plate_groups.items()):
        plate_wo_ids = {wo.id for wo in wos}
        plate_containers = sum(container_map.get(wid, 0) for wid in plate_wo_ids)
        plate_paid = sum(wo.driver_salary or 0 for wo in wos)
        drivers.append(
            {
                "plate": plate,
                "tripCount": len(wos),
                "containerCount": plate_containers,
                "totalPaid": plate_paid,
            }
        )

    recon_imports = (
        await db.execute(
            select(VendorReconciliationImport)
            .where(VendorReconciliationImport.vendor_id == vendor_id)
            .order_by(VendorReconciliationImport.uploaded_at.desc())
            .limit(20)
        )
    ).scalars().all()

    reconciliations = [
        {
            "importId": imp.id,
            "periodFrom": imp.period_from,
            "periodTo": imp.period_to,
            "containerCount": (imp.totals or {}).get("total", 0),
            "status": imp.status,
        }
        for imp in recon_imports
    ]

    return {
        "vendor": {
            "id": vendor.id,
            "name": vendor.name,
            "phone": vendor.phone,
            "taxCode": vendor.tax_code,
            "address": vendor.address,
            "contactPerson": vendor.contact_person,
        },
        "stats": {
            "tripCount": trip_count,
            "containerCount": container_count,
            "totalPaid": total_paid,
            "totalAmount": total_amount,
        },
        "drivers": drivers,
        "matched_trips": reconciliations,
    }
