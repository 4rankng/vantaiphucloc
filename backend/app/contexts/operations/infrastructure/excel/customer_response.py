import logging
from io import BytesIO

_logger = logging.getLogger(__name__)


async def parse_customer_response_excel(file_content: bytes) -> list[dict]:
    """Parse a customer's response Excel file.

    Expected columns (from our export template):
    - STT, Mã chuyến (trip_id), Ngày chạy, Số cont, Loại cont,
      Điểm lấy, Điểm trả, Tác nghiệp, Biển số xe, Số tàu, Đơn giá (VNĐ),
      Xác nhận KH, Ghi chú KH

    Returns list of dicts with:
      container_number, trip_date, customer_status, customer_note, customer_amount
    """
    import openpyxl
    from datetime import date as date_type

    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active

    # Find header row — look for known headers
    header_row_idx = None
    headers = []
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        row_vals = [str(v or "").strip() for v in row]
        if "Số cont" in row_vals or "Xác nhận KH" in row_vals:
            header_row_idx = idx
            headers = row_vals
            break

    if header_row_idx is None:
        workbook.close()
        raise ValueError("Không tìm thấy header. File phải có cột 'Số cont' hoặc 'Xác nhận KH'.")

    # Map header names to column indices
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h] = i

    # Vietnamese status normalization
    _VI_STATUS = {
        "OK": "MATCHED", "KHỚP": "MATCHED", "KHOP": "MATCHED",
        "SỬA": "MATCHED", "SỬA_SỐ_TIỀN": "MATCHED",
        "TỪ CHỐI": "REJECTED", "TỪ_CHỐI": "REJECTED", "TỪCHỐI": "REJECTED",
        "KHÔNG": "REJECTED",
    }

    results: list[dict] = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue

        # Get values by header
        def _get(header_name: str) -> str | None:
            idx = col_map.get(header_name)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        container = _get("Số cont")
        trip_date_raw = _get("Ngày chạy")
        confirm = _get("Xác nhận KH")
        note = _get("Ghi chú KH")
        amount_raw = _get("Đơn giá (VNĐ)")

        # Skip rows without container or confirmation
        if not container and not confirm:
            continue

        # Parse trip_date
        trip_date = None
        if trip_date_raw:
            try:
                if isinstance(row[col_map.get("Ngày chạy", -1)], date_type):
                    trip_date = row[col_map.get("Ngày chạy", -1)]
                else:
                    trip_date = date_type.fromisoformat(trip_date_raw)
            except (ValueError, TypeError, IndexError):
                pass

        # Parse customer_status from confirmation column
        customer_status = "UNKNOWN"
        if confirm:
            confirm_upper = confirm.upper().strip()
            customer_status = _VI_STATUS.get(confirm_upper, "UNKNOWN")
            # If it's empty/unfilled, default to UNKNOWN
            if confirm_upper in ("", "-", "—"):
                customer_status = "UNKNOWN"

        # Parse amount
        customer_amount = None
        if amount_raw:
            try:
                customer_amount = int(str(amount_raw).replace(",", "").replace(".", ""))
            except ValueError:
                pass

        results.append({
            "container_number": container,
            "trip_date": trip_date,
            "customer_status": customer_status,
            "customer_note": note,
            "customer_amount": customer_amount,
        })

    workbook.close()
    return results
