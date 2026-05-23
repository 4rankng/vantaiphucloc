"""Booked trip Excel import and template generation."""

import logging
from datetime import datetime as dt
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import (
    BookedTrip,
    Client,
    PricingLine,
    Vehicle,
    VehicleDriver,
)
from app.utils.iso6346 import normalize_container_number, validate_container_number

_logger = logging.getLogger(__name__)

_TRIP_IMPORT_COLUMNS = {
    "ngay": "trip_date",
    "date": "trip_date",
    "ngay_chay": "trip_date",
    "trip_date": "trip_date",
    "ma_kh": "client_code",
    "client_code": "client_code",
    "khach_hang": "client_code",
    "ma_khach_hang": "client_code",
    "diem_lay": "pickup_location",
    "pickup": "pickup_location",
    "pickup_location": "pickup_location",
    "diem_tra": "dropoff_location",
    "dropoff": "dropoff_location",
    "dropoff_location": "dropoff_location",
    "cung_duong": "route",
    "route": "route",
    "so_cont": "container_number",
    "container": "container_number",
    "container_number": "container_number",
    "loai": "work_type",
    "work_type": "work_type",
    "loai_cont": "work_type",
    "don_gia": "revenue",
    "revenue": "revenue",
    "luong_tx": "driver_salary",
    "driver_salary": "driver_salary",
    "phu_cap": "allowance",
    "allowance": "allowance",
}


def _normalize_trip_import_header(header: str) -> str:
    """Map Vietnamese/English header to standard field name."""
    if not header:
        return ""
    from app.utils.text import slugify_vi
    normalized = slugify_vi(header)
    return _TRIP_IMPORT_COLUMNS.get(normalized, normalized)


async def parse_booked_trip_excel(file_content: bytes) -> list[dict[str, Any]]:
    """Parse Excel file for batch trip order import.

    Expected columns (Vietnamese or English):
    - trip_date (Ngày/Ngày chạy)
    - client_code (Mã KH/Mã khách hàng)
    - pickup_location (Điểm lấy) + dropoff_location (Điểm trả) or route (Cung đường)
    - container_number (Số cont/Container)
    - work_type (Loại/Loại cont) — E20/E40/F20/F40
    - revenue (Đơn giá) — optional
    - driver_salary (Lương TX) — optional
    - allowance (Phụ cấp) — optional
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel processing")

    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active

    raw_headers = [cell.value for cell in sheet[1]]
    headers = [_normalize_trip_import_header(h) for h in raw_headers]

    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = value
        results.append(row_dict)

    workbook.close()
    _logger.info(f"Parsed {len(results)} trip order rows from Excel")
    return results


async def import_booked_trips(
    db: AsyncSession,
    rows: list[dict[str, Any]],
    user_id: int,
) -> dict[str, Any]:
    """Validate and create trip orders from parsed Excel rows.

    Groups rows by (trip_date, client_code) to form trip orders.
    Each group becomes one BookedTrip (flat cont_number/cont_type).
    Returns { created: int, errors: list[str] }.
    """
    from app.contexts.customer_pricing.infrastructure.pricing_lookup import (
        find_tiered_pricing,
    )

    created = 0
    errors: list[str] = []
    warnings: list[str] = []

    # Group rows by trip key
    groups: dict[tuple, list[dict]] = {}
    for i, row in enumerate(rows):
        trip_date = row.get("trip_date")
        client_code = row.get("client_code")

        if not trip_date or not client_code:
            errors.append(f"Dòng {i + 2}: thiếu ngày hoặc mã khách hàng")
            continue

        if isinstance(trip_date, str):
            try:
                trip_date = dt.strptime(trip_date, "%Y-%m-%d").date()
            except ValueError:
                try:
                    trip_date = dt.strptime(trip_date, "%d/%m/%Y").date()
                except ValueError:
                    errors.append(f"Dòng {i + 2}: định dạng ngày không hợp lệ '{trip_date}'")
                    continue
        elif hasattr(trip_date, "date"):
            pass  # already a date

        key = (str(trip_date), str(client_code))
        groups.setdefault(key, []).append(row)

    for key, group_rows in groups.items():
        trip_date_str, client_code_str = key

        # Look up partner by code
        client_result = await db.execute(
            select(Client).where(Client.code == client_code_str)
        )
        client = client_result.scalar_one_or_none()
        if not client:
            errors.append(f"Nhóm {key}: không tìm thấy đối tác mã '{client_code_str}'")
            continue

        first_row = group_rows[0]
        pickup = first_row.get("pickup_location")
        dropoff = first_row.get("dropoff_location")
        route = first_row.get("route") or (f"{pickup} - {dropoff}" if pickup and dropoff else "")

        # Build containers (normalize numbers — strip hyphens, uppercase)
        containers_data = []
        for row in group_rows:
            cn = normalize_container_number(str(row.get("container_number", "")).strip())
            ct = str(row.get("work_type", "")).strip().upper()
            if cn:
                valid, err = validate_container_number(cn)
                if not valid:
                    errors.append(f"Nhóm {key}: Container {cn} không hợp lệ — {err}")
                    continue
                containers_data.append({"container_number": cn, "cont_type": ct or "E20"})

        if not containers_data:
            errors.append(f"Nhóm {key}: không có số container")
            continue

        cont_type = containers_data[0]["cont_type"]
        container_count = sum(1 for c in containers_data if c["cont_type"] == cont_type) or 1

        # Try auto-pricing
        revenue = int(first_row.get("revenue") or 0)
        driver_salary = int(first_row.get("driver_salary") or 0)
        allowance = int(first_row.get("allowance") or 0)
        pricing_id = None

        # Resolve pickup/dropoff strings to Location FKs via the alias
        # resolver. Auto-creates if missing.
        from app.contexts.customer_pricing.infrastructure.location_resolver import (
            LocationResolverService,
            ResolverSource,
        )
        resolver = LocationResolverService(db)
        pickup_id = None
        dropoff_id = None
        if pickup:
            r = await resolver.resolve_or_create(pickup, source=ResolverSource.MANUAL, user_id=user_id)
            pickup_id = r.location.id if r.location else None
        if dropoff:
            r = await resolver.resolve_or_create(dropoff, source=ResolverSource.MANUAL, user_id=user_id)
            dropoff_id = r.location.id if r.location else None
        if not pickup_id or not dropoff_id:
            errors.append(f"Nhóm {key}: pickup/dropoff không có")
            continue

        if not revenue:
            tiered = await find_tiered_pricing(
                db, client_id=client.id, work_type=work_type,
                quantity=container_count,
                pickup_location_id=pickup_id, dropoff_location_id=dropoff_id,
            )
            if tiered:
                revenue = tiered.revenue
                driver_salary = tiered.driver_salary
                allowance = tiered.allowance
                pricing_id = tiered.pricing.id

        # One BookedTrip per group — use first container's data
        booked_trip = BookedTrip(
            trip_date=trip_date_str,
            client_id=client.id,
            pickup_location_id=pickup_id,
            dropoff_location_id=dropoff_id,
            work_type="CHUYỂN BÃI",
            cont_number=containers_data[0]["container_number"],
            cont_type=containers_data[0]["cont_type"],
        )
        db.add(booked_trip)

        created += 1

    await db.commit()
    return {"created": created, "errors": errors, "warnings": warnings}


def generate_booked_trip_template() -> bytes:
    """Generate a blank Excel template for trip order import."""
    import openpyxl

    headers = [
        "Ngày", "Mã KH", "Điểm lấy", "Điểm trả",
        "Cung đường", "Số cont", "Loại cont", "Đơn giá",
        "Lương TX", "Phụ cấp",
    ]
    examples = [
        "01/05/2026", "KH001", "Cát lái", "KC Bình Dương",
        "Cát lái - KC Bình Dương", "TCLU1234567", "E20", "1500000",
        "500000", "100000",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Nhập chuyến"
    sheet.append(headers)
    sheet.append(examples)

    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)

    buf = BytesIO()
    workbook.save(buf)
    workbook.close()
    return buf.getvalue()
