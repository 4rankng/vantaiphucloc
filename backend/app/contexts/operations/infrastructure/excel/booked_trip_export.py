"""Booked trip and doi-soat Excel export helpers."""

import logging

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import (
    BookedTrip,
    Client,
    DeliveredTrip,
    Location,
)
from app.utils.excel_utils import (
    apply_header_style,
    auto_fit_columns,
    workbook_to_bytes,
)

_logger = logging.getLogger(__name__)


async def generate_booked_trips_excel(
    db: AsyncSession,
    date_from: str | None = None,
    date_to: str | None = None,
    client_id: int | None = None,
) -> bytes:
    """Export trip orders to Excel.

    When client_id is provided, filters to that client and includes
    match status columns for customer reconciliation.
    """
    import openpyxl

    from datetime import date as date_type

    query = select(BookedTrip).order_by(BookedTrip.id.desc())
    if client_id:
        query = query.where(BookedTrip.client_id == client_id)
    if date_from:
        query = query.where(BookedTrip.trip_date >= date_type.fromisoformat(date_from))
    if date_to:
        query = query.where(BookedTrip.trip_date <= date_type.fromisoformat(date_to))

    result = await db.execute(query)
    booked_trips = result.scalars().all()

    # Resolve display names via JOIN.
    client_ids = {to.client_id for to in booked_trips}
    loc_ids = {to.pickup_location_id for to in booked_trips} | {
        to.dropoff_location_id for to in booked_trips
    }
    loc_ids.discard(None)
    client_name_by_id = (
        {
            c.id: c.name
            for c in (await db.execute(select(Client).where(Client.id.in_(client_ids))))
            .scalars()
            .all()
        }
        if client_ids
        else {}
    )
    loc_name_by_id = (
        {
            loc.id: loc.name
            for loc in (
                await db.execute(select(Location).where(Location.id.in_(loc_ids)))
            )
            .scalars()
            .all()
        }
        if loc_ids
        else {}
    )

    # For per-partner export: determine match status and vessel from BookedTrip directly
    client_name = None
    if client_id:
        # Get partner name for filename
        p_result = await db.execute(select(Client).where(Client.id == client_id))
        client_obj = p_result.scalar_one_or_none()
        client_name = client_obj.name if client_obj else None

    wb = openpyxl.Workbook()
    ws = wb.active

    if client_id:
        ws.title = "Chuyến theo khách hàng"
        headers = [
            "STT",
            "Số container",
            "Tuyến đường",
            "Ngày chạy",
            "Biển số xe",
            "Số tàu",
            "Trạng thái khớp",
            "Đơn giá",
        ]
    else:
        ws.title = "Đơn hàng"
        headers = [
            "Mã TO",
            "Ngày chạy",
            "Khách hàng",
            "Điểm lấy",
            "Điểm trả",
            "Số cont",
            "Loại",
            "Đơn giá",
            "Lương TX",
            "Phụ cấp",
            "Trạng thái",
        ]
    ws.append(headers)

    apply_header_style(ws, 1, len(headers))

    match_labels = {True: "Đã khớp", False: "Chưa khớp"}

    # Build match status lookup: BookedTrip.id -> bool
    bt_ids = [to.id for to in booked_trips]
    matched_bt_ids: set[int] = set()
    if bt_ids:
        rows = (
            (
                await db.execute(
                    select(DeliveredTrip.booked_trip_id).where(
                        DeliveredTrip.booked_trip_id.in_(bt_ids)
                    )
                )
            )
            .scalars()
            .all()
        )
        matched_bt_ids = set(rows)

    if client_id:
        stt = 0
        for to in booked_trips:
            pickup = loc_name_by_id.get(to.pickup_location_id, "")
            dropoff = loc_name_by_id.get(to.dropoff_location_id, "")
            route = f"{pickup} → {dropoff}" if pickup and dropoff else ""
            plate = to.vehicle_plate or ""
            match_status = match_labels.get(to.id in matched_bt_ids, "Chưa khớp")
            vessel = to.vessel or ""
            stt += 1
            ws.append(
                [
                    stt,
                    to.cont_number or "",
                    route,
                    to.trip_date,
                    plate,
                    vessel,
                    match_status,
                ]
            )
    else:
        for to in booked_trips:
            ws.append(
                [
                    f"TO#{to.id}",
                    to.trip_date,
                    client_name_by_id.get(to.client_id, ""),
                    loc_name_by_id.get(to.pickup_location_id, ""),
                    loc_name_by_id.get(to.dropoff_location_id, ""),
                    to.cont_number or "",
                    to.cont_type or "",
                    match_labels.get(to.id in matched_bt_ids, "Chờ ghép"),
                ]
            )

    from app.utils.excel_utils import add_template_version as _add_ver

    auto_fit_columns(ws)
    _add_ver(ws, 13)
    return workbook_to_bytes(wb), client_name


async def generate_doi_soat_excel(
    db: AsyncSession,
    client_id: int,
    date_from: str,
    date_to: str,
) -> tuple[bytes, str]:
    """Generate reconciliation (đối soát) Excel for a specific client.

    Returns (excel_bytes, client_name) tuple.

    Unions BookedTrip (containers listed in customer's ship file) and
    DeliveredTrip (containers actually moved by drivers) for the same
    client within the date range. Each container occurrence yields its
    own row (duplicate numbers are not collapsed, so repeated trips stay
    visible). The trailing ``TRẠNG THÁI`` column reads:
    - ``Đã ghép`` when the container appears on both sides
    - ``Chưa ghép`` when it appears on only one side

    Columns: STT | NGÀY ĐI | CHỦ HÀNG | SỐ CONTAINER | F20' | F40' | E20' |
    E40' | SỐ XE CHẠY | ĐIỂM ĐI | ĐIỂM ĐẾN | TÁC NGHIỆP | CƯỚC | LƯƠNG |
    TRẠNG THÁI. Subtotal row at row 11 counts F20'/F40'/E20'/E40' and
    sums CƯỚC/LƯƠNG across all rows.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from app.utils.iso6346 import normalize_container_number

    STATUS_MATCHED = "Đã ghép"
    STATUS_UNMATCHED = "Chưa ghép"

    # -- 1. Load client --
    p_result = await db.execute(select(Client).where(Client.id == client_id))
    client = p_result.scalar_one_or_none()
    client_name = client.name if client else f"Client #{client_id}"
    client_code = (client.code or client_name) if client else f"Client #{client_id}"

    from datetime import date as date_type

    df = date_type.fromisoformat(date_from)
    dt = date_type.fromisoformat(date_to)

    # -- 2. Load BookedTrip (customer's ship file rows) --
    bt_query = (
        select(BookedTrip)
        .where(
            BookedTrip.client_id == client_id,
            BookedTrip.trip_date >= df,
            BookedTrip.trip_date <= dt,
        )
        .order_by(BookedTrip.trip_date, BookedTrip.id)
    )
    booked_trips = (await db.execute(bt_query)).scalars().all()

    # -- 3. Load DeliveredTrip (driver's actual work) --
    dt_query = (
        select(DeliveredTrip)
        .where(
            DeliveredTrip.client_id == client_id,
            DeliveredTrip.trip_date >= df,
            DeliveredTrip.trip_date <= dt,
        )
        .order_by(DeliveredTrip.trip_date, DeliveredTrip.id)
    )
    delivered_trips = (await db.execute(dt_query)).scalars().all()

    # -- 4. Bucket by normalized container number (preserve every trip) --
    # A container is NOT collapsed when it repeats: if the same number is on
    # several BookedTrips or DeliveredTrips, each trip keeps its own row so
    # duplicate trips stay visible in the reconciliation sheet (the
    # duplicate-containers endpoint exists to flag exactly these cases).
    bt_by_cont: dict[str, list[BookedTrip]] = {}
    bt_no_cont: list[BookedTrip] = []
    for bt in booked_trips:
        if bt.cont_number:
            bt_by_cont.setdefault(
                normalize_container_number(bt.cont_number), []
            ).append(bt)
        else:
            bt_no_cont.append(bt)

    dt_by_cont: dict[str, list[DeliveredTrip]] = {}
    dt_no_cont: list[DeliveredTrip] = []
    for trip in delivered_trips:
        if trip.cont_number:
            dt_by_cont.setdefault(
                normalize_container_number(trip.cont_number), []
            ).append(trip)
        else:
            dt_no_cont.append(trip)

    # -- 5. Build rows --
    # A container's rows are "Đã ghép" when the number exists on both sides,
    # otherwise "Chưa ghép". Fields prefer the DeliveredTrip (the executor's
    # actual work) and fall back to the BookedTrip for anything missing.
    def _doi_soat_row(source, fallback, status):
        fb = fallback
        return {
            "trip_date": source.trip_date or (fb.trip_date if fb else None),
            "cont_number": source.cont_number or (fb.cont_number if fb else None) or "",
            "cont_type": source.cont_type or (fb.cont_type if fb else None),
            "vehicle_plate": source.vehicle_plate
            or (fb.vehicle_plate if fb else None)
            or "",
            "pickup_location_id": source.pickup_location_id
            or (fb.pickup_location_id if fb else None),
            "dropoff_location_id": source.dropoff_location_id
            or (fb.dropoff_location_id if fb else None),
            "work_type": source.work_type or (fb.work_type if fb else None) or "",
            "revenue": getattr(source, "revenue", None),
            "driver_salary": getattr(source, "driver_salary", None),
            "status": status,
        }

    deduped_rows: list[dict] = []
    for key in set(bt_by_cont) | set(dt_by_cont):
        bts = bt_by_cont.get(key, [])
        trips = dt_by_cont.get(key, [])
        matched = bool(trips) and bool(bts)
        status = STATUS_MATCHED if matched else STATUS_UNMATCHED
        # Emit one row per delivered trip (the executor's work); pair each
        # with the corresponding booked trip for field fallback. Booked trips
        # in excess of delivered ones are emitted separately as "Chưa ghép"
        # (customer listed but the container was not moved that many times).
        primary = trips if trips else bts
        for i, source in enumerate(primary):
            fallback = bts[i] if i < len(bts) else (bts[0] if bts else None)
            deduped_rows.append(_doi_soat_row(source, fallback, status))
        if trips and len(bts) > len(trips):
            for bt in bts[len(trips) :]:
                deduped_rows.append(_doi_soat_row(bt, None, STATUS_UNMATCHED))

    for bt in bt_no_cont:
        deduped_rows.append(_doi_soat_row(bt, None, STATUS_UNMATCHED))
    for trip in dt_no_cont:
        deduped_rows.append(_doi_soat_row(trip, None, STATUS_UNMATCHED))

    # Sort by trip_date then container number for stable ordering
    deduped_rows.sort(key=lambda r: (r["trip_date"] or df, r["cont_number"] or ""))

    # -- 6. Load location names --
    loc_ids: set[int] = set()
    for row in deduped_rows:
        if row["pickup_location_id"]:
            loc_ids.add(row["pickup_location_id"])
        if row["dropoff_location_id"]:
            loc_ids.add(row["dropoff_location_id"])
    loc_name_by_id: dict[int, str] = {}
    if loc_ids:
        loc_result = await db.execute(select(Location).where(Location.id.in_(loc_ids)))
        loc_name_by_id = {loc.id: loc.name for loc in loc_result.scalars().all()}

    # -- 7. Build Excel workbook --
    wb = openpyxl.Workbook()
    ws = wb.active
    month_label = df.strftime("%m/%Y")
    ws.title = f"SL T{df.month}.{str(df.year)[2:]}"

    num_cols = 15  # A-O
    last_col = get_column_letter(num_cols)

    # -- Styles --
    _bold = Font(bold=True, size=11)
    _bold14 = Font(bold=True, size=14)
    _header_font = Font(bold=True, size=11)
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # -- Title block (rows 1-8) --
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "CÔNG TY TNHH AMT PHÚC LỘC"
    ws["A1"].font = _bold

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Địa chỉ: Số 56B/97 đường Đoàn Kết, P. Hải An, TP Hải Phòng"
    ws["A2"].font = _bold

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = "MST: 0201965047"

    # Row 4 empty
    ws.merge_cells(f"A5:{last_col}5")
    ws["A5"] = f"BẢNG KÊ QUYẾT TOÁN CƯỚC VẬN CHUYỂN THÁNG {month_label}"
    ws["A5"].font = _bold14
    ws["A5"].alignment = center

    ws.merge_cells(f"A6:{last_col}6")
    ws["A6"] = f"Từ ngày {df.strftime('%d/%m/%Y')} đến ngày {dt.strftime('%d/%m/%Y')}"
    ws["A6"].font = _bold
    ws["A6"].alignment = center

    ws["A7"] = "KHÁCH HÀNG:"
    ws.merge_cells(f"C7:{last_col}7")
    ws["C7"] = client_name
    ws["C7"].font = _bold

    ws.merge_cells(f"A8:{last_col}8")
    ws["A8"] = (
        "Công ty TNHH AMT Phúc Lộc xin gửi tới Quý Công ty bảng kê quyết toán vận tải như sau:"
    )

    # -- Header row 10 --
    headers = [
        "STT",
        "NGÀY ĐI",
        "CHỦ HÀNG",
        "SỐ CONTAINER",
        "F20'",
        "F40'",
        "E20'",
        "E40'",
        "SỐ XE CHẠY",
        "ĐIỂM ĐI",
        "ĐIỂM ĐẾN",
        "TÁC NGHIỆP",
        "CƯỚC",
        "LƯƠNG",
        "TRẠNG THÁI",
    ]
    ws.append([])  # row 9 empty
    ws.append(headers)  # row 10
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=10, column=col_num)
        cell.font = _header_font
        cell.alignment = center
        cell.border = thin_border

    # -- Subtotal row 11 --
    # Data rows run 12..(11 + N); use the correct last row up front so the
    # subtotal formula doesn't need to be rewritten after the loop.
    last_data_row = max(len(deduped_rows) + 11, 12)
    ws.append(
        [
            "",
            "",
            "",
            "",
            f"=SUBTOTAL(9,E12:E{last_data_row})",
            f"=SUBTOTAL(9,F12:F{last_data_row})",
            f"=SUBTOTAL(9,G12:G{last_data_row})",
            f"=SUBTOTAL(9,H12:H{last_data_row})",
            "",
            "",
            "",
            "",
            f"=SUBTOTAL(9,M12:M{last_data_row})",
            f"=SUBTOTAL(9,N12:N{last_data_row})",
            "",
        ]
    )
    for col_num in [5, 6, 7, 8, 13, 14]:
        ws.cell(row=11, column=col_num).font = _bold
        ws.cell(row=11, column=col_num).alignment = (
            right if col_num in (13, 14) else center
        )
        if col_num in (13, 14):
            ws.cell(row=11, column=col_num).number_format = "#,##0"

    # -- Data rows (12+) -- one row per unique container
    stt = 0

    for row in deduped_rows:
        pickup = loc_name_by_id.get(row["pickup_location_id"] or 0, "")
        dropoff = loc_name_by_id.get(row["dropoff_location_id"] or 0, "")
        trip_date_str = (
            row["trip_date"].strftime("%d/%m/%Y") if row["trip_date"] else ""
        )

        ct = (row["cont_type"] or "").upper()

        # Cont type flags
        f20 = 1 if ct == "F20" else None
        f40 = 1 if ct == "F40" else None
        e20 = 1 if ct == "E20" else None
        e40 = 1 if ct == "E40" else None

        stt += 1
        row_num = 11 + stt
        ws.append(
            [
                stt,
                trip_date_str,
                client_code,
                row["cont_number"],
                f20,
                f40,
                e20,
                e40,
                row["vehicle_plate"],
                pickup,
                dropoff,
                row["work_type"],
                row["revenue"],
                row["driver_salary"],
                row["status"],
            ]
        )

        # Styling
        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in (1, 2, 3, 5, 6, 7, 8, 9, 10, 11, 12, 15):
                cell.alignment = center
            elif col_num == 4:
                cell.alignment = left
            elif col_num in (13, 14):
                cell.alignment = right
                cell.number_format = "#,##0"

    # -- Column widths --
    col_widths = [6, 12, 12, 18, 6, 6, 6, 6, 14, 20, 20, 16, 14, 14, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header
    ws.freeze_panes = "A12"

    return workbook_to_bytes(wb), client_name
