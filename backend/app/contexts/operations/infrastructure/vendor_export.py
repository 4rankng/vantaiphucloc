"""Vendor reconciliation Excel export generation."""

from __future__ import annotations

import io
from datetime import date

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import (
    Location,
    Vendor,
    DeliveredTrip,
)
from app.utils.excel_utils import add_template_version
from app.utils.text import slugify_vi


async def generate_vendor_export_excel(
    db: AsyncSession,
    vendor: Vendor,
    date_from: date,
    date_to: date,
) -> tuple[bytes, str]:
    """Generate vendor trips Excel export. Returns (excel_bytes, filename)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    vendor_id = vendor.id

    # Load DeliveredTrips for this vendor in the period
    wo_result = await db.execute(
        select(DeliveredTrip).where(
            DeliveredTrip.vendor_id == vendor_id,
            DeliveredTrip.trip_date >= date_from,
            DeliveredTrip.trip_date <= date_to,
        ).order_by(DeliveredTrip.trip_date, DeliveredTrip.id)
    )
    delivered_trips = wo_result.scalars().all()

    # Load location names
    loc_ids: set[int] = set()
    for wo in delivered_trips:
        if wo.pickup_location_id:
            loc_ids.add(wo.pickup_location_id)
        if wo.dropoff_location_id:
            loc_ids.add(wo.dropoff_location_id)
    loc_name_by_id: dict[int, str] = {}
    if loc_ids:
        loc_result = await db.execute(select(Location).where(Location.id.in_(loc_ids)))
        loc_name_by_id = {loc.id: loc.name for loc in loc_result.scalars().all()}

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = vendor.name[:31]

    # Styles
    thin_side = Side(style="thin", color="BBBBBB")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    _blue_dark = "1F4E79"
    _blue_header = "2E75B6"
    _white = "FFFFFF"
    _grey_row = "F5F8FC"
    _blue_light = "DEEAF1"
    _yellow_sum = "FFF2CC"

    num_cols = 12
    last_col = get_column_letter(num_cols)

    # Title rows
    ws.append(["VẬN TẢI PHÚC LỘC", *([""] * (num_cols - 1))])
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].font = Font(bold=True, size=14, color=_blue_dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    month_range = f"Từ {date_from.strftime('%d/%m/%Y')} đến {date_to.strftime('%d/%m/%Y')}"
    ws.append([f"BẢNG CHUYẾN – {vendor.name.upper()} – {month_range}", *([""] * (num_cols - 1))])
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].font = Font(bold=True, size=11, color=_blue_dark)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.append([""] * num_cols)
    ws.row_dimensions[3].height = 6

    # Header row
    _op_labels = {
        "XUẤT/NHẬP TÀU": "Xuất / Nhập tàu",
        "CHUYỂN BÃI": "Chuyển bãi",
        "LẤY VỎ HẠ HÀNG": "Lấy vỏ hạ hàng",
        "CHẠY SÀ LAN": "Chạy sà lan",
        "ĐÓNG KHO": "Đóng kho",
    }
    WORK_TYPE_FULL = {
        "E20": "Rỗng 20ft", "E40": "Rỗng 40ft",
        "F20": "Hàng 20ft", "F40": "Hàng 40ft",
    }

    headers = [
        "STT", "Mã phiếu", "Ngày chạy", "Số cont", "Loại cont",
        "Điểm lấy", "Điểm trả", "Tác nghiệp", "Biển số xe", "Số tàu",
        "Đơn giá (VNĐ)", "Ghi chú",
    ]
    ws.append(headers)
    header_row = 4
    header_font = Font(bold=True, color=_white, size=10)
    header_fill = PatternFill(start_color=_blue_header, end_color=_blue_header, fill_type="solid")
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=thin_side, right=thin_side,
            top=Side(style="thin", color="BBBBBB"),
            bottom=Side(style="medium", color="1F4E79"),
        )
    ws.row_dimensions[header_row].height = 32
    ws.freeze_panes = ws.cell(row=5, column=1)

    # Data rows — one row per DeliveredTrip (flat cont_number/cont_type)
    stt = 0
    type_count: dict[str, int] = {}
    total_amount = 0

    for row_idx, wo in enumerate(delivered_trips):
        pickup = loc_name_by_id.get(wo.pickup_location_id or 0, "")
        dropoff = loc_name_by_id.get(wo.dropoff_location_id or 0, "")
        plate = wo.vehicle_plate or ""
        vessel = wo.vessel or ""
        op_type = ""  # operation_type removed; work_type carries equivalent meaning
        revenue = wo.revenue or 0
        trip_date_str = wo.trip_date.strftime("%d/%m/%Y") if wo.trip_date else ""

        fill_color = _white if row_idx % 2 == 0 else _grey_row
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        stt += 1
        cont_label = WORK_TYPE_FULL.get(wo.cont_type or "", wo.cont_type or wo.work_type or "")
        type_count[cont_label] = type_count.get(cont_label, 0) + 1
        total_amount += revenue

        data_row = [
            stt, wo.id, trip_date_str, wo.cont_number or "", cont_label,
            pickup, dropoff, op_type, plate, vessel, revenue or "", "",
        ]
        ws.append(data_row)
        data_row_num = ws.max_row

        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=data_row_num, column=col_num)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_num in (1, 2):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_num == 11 and revenue:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[data_row_num].height = 18

    # Summary
    sum_fill = PatternFill(start_color=_blue_light, end_color=_blue_light, fill_type="solid")
    sum_font_bold = Font(bold=True, size=10, color=_blue_dark)

    ws.append([""] * num_cols)
    ws.append(["Tổng hợp theo loại container:", *([""] * (num_cols - 1))])
    label_row = ws.max_row
    ws.merge_cells(f"A{label_row}:{last_col}{label_row}")
    ws[f"A{label_row}"].font = sum_font_bold
    ws[f"A{label_row}"].fill = sum_fill

    for wt_label, count in sorted(type_count.items()):
        ws.append(["", f"  {wt_label}", count, "cont", *([""] * (num_cols - 4))])
        r = ws.max_row
        for col_num in range(1, num_cols + 1):
            ws.cell(row=r, column=col_num).fill = sum_fill
        ws.cell(row=r, column=2).font = Font(size=10)
        ws.cell(row=r, column=3).font = Font(bold=True, size=10, color=_blue_dark)

    total_fill = PatternFill(start_color=_yellow_sum, end_color=_yellow_sum, fill_type="solid")
    total_rows_data = [["Tổng số container:", *([""] * 9), stt, ""]]
    if total_amount:
        total_rows_data.append(["Tổng đơn giá:", *([""] * 9), total_amount, ""])

    for row_data in total_rows_data:
        ws.append(row_data)
        r = ws.max_row
        ws.merge_cells(f"A{r}:J{r}")
        label_cell = ws.cell(row=r, column=1)
        val_cell = ws.cell(row=r, column=11)
        label_cell.font = Font(bold=True, size=10, color=_blue_dark)
        val_cell.font = Font(bold=True, size=11, color=_blue_dark)
        val_cell.number_format = '#,##0'
        val_cell.alignment = Alignment(horizontal="right", vertical="center")
        for col_num in range(1, num_cols + 1):
            ws.cell(row=r, column=col_num).fill = total_fill

    # Column widths
    col_widths = [6, 12, 12, 16, 14, 24, 24, 14, 14, 14, 16, 24]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    add_template_version(ws, num_cols)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    slug = slugify_vi(vendor.name)
    month_str = date_from.strftime("%m-%Y")
    filename = f"DoiSoat_NhaXe_{slug}_{month_str}.xlsx"

    return buf.getvalue(), filename
