"""Format-agnostic workbook reader.

Unifies `.xlsx` (openpyxl) and `.xls` (xlrd) behind a tiny interface so the
rest of the pipeline doesn't care about the file format.

Cells are returned as Python primitives (str | int | float | datetime |
None). Empty cells become None.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any


@dataclass
class SheetView:
    name: str
    state: str           # "visible" | "hidden" | "veryHidden"
    n_rows: int
    n_cols: int
    rows: list[list[Any]]   # rows[r][c]; r,c are 0-indexed


def load_workbook(content: bytes, filename: str) -> list[SheetView]:
    """Return one `SheetView` per sheet, in workbook order."""
    name = (filename or "").lower()
    if name.endswith(".xls"):
        return _load_xls(content)
    return _load_xlsx(content)


# ---------------------------------------------------------------------------
# .xlsx via openpyxl
# ---------------------------------------------------------------------------

def _load_xlsx(content: bytes) -> list[SheetView]:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=False)
    out: list[SheetView] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            state = ws.sheet_state or "visible"
            # Skip hidden sheets entirely
            if state in ("hidden", "veryHidden"):
                continue
            n_rows = ws.max_row or 0
            n_cols = ws.max_column or 0
            # Detect hidden columns
            hidden_cols: set[int] = set()
            for col_idx in range(1, n_cols + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                dim = ws.column_dimensions.get(col_letter)
                if dim and dim.hidden:
                    hidden_cols.add(col_idx - 1)  # 0-based index
            # Cap for very large sheets
            cap_rows = min(n_rows, 5000)
            rows: list[list[Any]] = []
            if cap_rows > 0 and n_cols > 0:
                for row in ws.iter_rows(min_row=1, max_row=cap_rows, max_col=n_cols, values_only=True):
                    if hidden_cols:
                        rows.append([v for i, v in enumerate(row) if i not in hidden_cols])
                    else:
                        rows.append(list(row))
            visible_cols = n_cols - len(hidden_cols)
            out.append(SheetView(
                name=sheet_name, state=state,
                n_rows=cap_rows, n_cols=visible_cols, rows=rows,
            ))
    finally:
        wb.close()
    return out


# ---------------------------------------------------------------------------
# .xls via xlrd
# ---------------------------------------------------------------------------

def _load_xls(content: bytes) -> list[SheetView]:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content, formatting_info=False)
    out: list[SheetView] = []
    for idx in range(wb.nsheets):
        # Skip hidden sheets (0=visible, 1=hidden, 2=very hidden)
        vis = wb._sheet_visibility
        if vis and idx < len(vis) and vis[idx] != 0:
            continue
        ws = wb.sheet_by_index(idx)
        n_rows = ws.nrows
        n_cols = ws.ncols
        # Detect hidden columns from colinfo_map
        hidden_cols: set[int] = set()
        if hasattr(ws, 'colinfo_map') and ws.colinfo_map:
            for col_idx, info in ws.colinfo_map.items():
                if getattr(info, 'is_hidden', False):
                    hidden_cols.add(col_idx)
        cap_rows = min(n_rows, 5000)
        rows: list[list[Any]] = []
        for r in range(cap_rows):
            row: list[Any] = []
            for c in range(n_cols):
                if c in hidden_cols:
                    continue
                cell = ws.cell(r, c)
                row.append(_xls_cell_value(cell, wb))
            rows.append(row)
        visible_cols = n_cols - len(hidden_cols)
        out.append(SheetView(
            name=ws.name, state="visible",
            n_rows=cap_rows, n_cols=visible_cols, rows=rows,
        ))
    return out


def _xls_cell_value(cell: Any, wb: Any) -> Any:
    import xlrd

    if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            tup = xlrd.xldate_as_tuple(cell.value, wb.datemode)
            return datetime(*tup) if tup[0] else None
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        f = float(cell.value)
        if f.is_integer():
            return int(f)
        return f
    return cell.value
