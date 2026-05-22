"""Excel file processing for customer reconciliation uploads and trip order imports."""

import logging
import re
from datetime import date as date_type, datetime as dt
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import DeliveredTrip, DeliveredTripContainer, BookedTrip, BookedTripContainer, Client, PricingLine, Vehicle, VehicleDriver
from app.utils.iso6346 import normalize_container_number, validate_container_number
from app.utils.excel_utils import parse_operation_type

_logger = logging.getLogger(__name__)


class ReconciliationResult:
    """Result of comparing customer Excel data with system records."""

    def __init__(
        self,
        container_number: str,
        normalized_number: str,
        delivered_trip_id: int | None = None,
        booked_trip_id: int | None = None,
        status: str = "pending",
        match_type: str = "none",
    ):
        self.container_number = container_number
        self.normalized_number = normalized_number
        self.delivered_trip_id = delivered_trip_id
        self.booked_trip_id = booked_trip_id
        self.status = status  # "confirmed" | "pending" | "rejected"
        self.is_duplicate = False
        self.match_type = match_type  # "exact" | "partial" | "none"

    def to_dict(self) -> dict[str, Any]:
        return {
            "container_number": self.container_number,
            "normalized_number": self.normalized_number,
            "delivered_trip_id": self.delivered_trip_id,
            "booked_trip_id": self.booked_trip_id,
            "status": self.status,
            "is_duplicate": self.is_duplicate,
            "match_type": self.match_type,
        }


async def parse_customer_excel(
    file_content: bytes,
    client_id: int | None = None,
) -> list[dict[str, Any]]:
    """
    Parse Excel file uploaded by customer.

    Expected columns (to be confirmed with customer):
    - Container Number (required)
    - Date (optional)
    - Pickup Location (optional)
    - Dropoff Location (optional)

    Returns list of dicts with parsed data.
    """
    try:
        import openpyxl
    except ImportError:
        _logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise ImportError("openpyxl is required for Excel processing")

    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active

    # Get header row
    headers = [cell.value for cell in sheet[1]]
    _logger.info(f"Excel headers: {headers}")

    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue

        row_dict = dict(zip(headers, row))
        results.append(row_dict)

    workbook.close()
    _logger.info(f"Parsed {len(results)} rows from Excel file")
    return results


async def compare_with_system_records(
    db: AsyncSession,
    client_id: int,
    excel_data: list[dict[str, Any]],
    date_from: str | None = None,
    date_to: str | None = None,
) -> list[ReconciliationResult]:
    """
    Compare customer Excel data with system work orders and trip orders.

    Returns list of ReconciliationResult objects with duplicate detection.
    """
    results: list[ReconciliationResult] = []

    # Extract and normalize container numbers from Excel
    excel_containers = []
    for row in excel_data:
        container_number = row.get("Container Number") or row.get("Số Container") or row.get("container_number")
        if container_number:
            normalized = normalize_container_number(str(container_number))
            excel_containers.append({
                "original": container_number,
                "normalized": normalized,
                "row_data": row,
            })

    # Query work orders for this partner
    wo_query = select(DeliveredTrip).where(DeliveredTrip.client_id == client_id)
    if date_from:
        wo_query = wo_query.where(DeliveredTrip.created_at >= date_from)
    if date_to:
        wo_query = wo_query.where(DeliveredTrip.created_at <= date_to)

    wo_result = await db.execute(wo_query)
    delivered_trips = wo_result.scalars().all()

    # Query trip orders for this partner
    to_query = select(BookedTrip).where(BookedTrip.client_id == client_id)
    if date_from:
        to_query = to_query.where(BookedTrip.trip_date >= date_from)
    if date_to:
        to_query = to_query.where(BookedTrip.trip_date <= date_to)

    to_result = await db.execute(to_query)
    booked_trips = to_result.scalars().all()

    # Load containers for all work orders
    wo_ids = [wo.id for wo in delivered_trips]
    if wo_ids:
        wo_cont_result = await db.execute(
            select(DeliveredTripContainer).where(
                DeliveredTripContainer.delivered_trip_id.in_(wo_ids)
            )
        )
        wo_containers = wo_cont_result.scalars().all()

        # Build lookup: normalized container number -> list of work order IDs
        wo_container_map: dict[str, list[int]] = {}
        for wc in wo_containers:
            normalized = normalize_container_number(wc.container_number)
            if normalized not in wo_container_map:
                wo_container_map[normalized] = []
            wo_container_map[normalized].append(wc.delivered_trip_id)

    # Load containers for all trip orders
    to_ids = [to.id for to in booked_trips]
    if to_ids:
        to_cont_result = await db.execute(
            select(BookedTripContainer).where(
                BookedTripContainer.booked_trip_id.in_(to_ids)
            )
        )
        to_containers = to_cont_result.scalars().all()

        # Build lookup: normalized container number -> list of trip order IDs
        to_container_map: dict[str, list[int]] = {}
        for tc in to_containers:
            normalized = normalize_container_number(tc.container_number)
            if normalized not in to_container_map:
                to_container_map[normalized] = []
            to_container_map[normalized].append(tc.booked_trip_id)

    # Compare Excel containers with system records
    # Build digits-only lookup for partial matching
    def _digits_only(s: str) -> str:
        return re.sub(r'[^0-9]', '', s)

    wo_digits_map: dict[str, list[int]] = {}
    for cn, ids in wo_container_map.items():
        digits = _digits_only(cn)
        if digits:
            wo_digits_map.setdefault(digits, []).extend(ids)

    to_digits_map: dict[str, list[int]] = {}
    for cn, ids in to_container_map.items():
        digits = _digits_only(cn)
        if digits:
            to_digits_map.setdefault(digits, []).extend(ids)

    for excel_cont in excel_containers:
        normalized = excel_cont["normalized"]

        # Check for exact matches
        wo_ids = wo_container_map.get(normalized, [])
        to_ids = to_container_map.get(normalized, [])

        match_type = "none"

        if wo_ids or to_ids:
            match_type = "exact"
        else:
            # Try digits-only partial match
            digits = _digits_only(normalized)
            if digits:
                wo_ids = list(set(wo_digits_map.get(digits, [])))
                to_ids = list(set(to_digits_map.get(digits, [])))
                if wo_ids or to_ids:
                    match_type = "partial"

        result = ReconciliationResult(
            container_number=excel_cont["original"],
            normalized_number=normalized,
            status="pending",
            match_type=match_type,
        )

        if wo_ids or to_ids:
            result.is_duplicate = True
            result.delivered_trip_id = wo_ids[0] if wo_ids else None
            result.booked_trip_id = to_ids[0] if to_ids else None
            if wo_ids and to_ids:
                result.status = "confirmed"

        results.append(result)

    _logger.info(
        f"Comparison complete: {len(results)} containers, "
        f"{sum(1 for r in results if r.is_duplicate)} duplicates found"
    )

    return results


async def generate_reconciliation_excel(
    db: AsyncSession,
    client_id: int,
    date_from: str | None = None,
    date_to: str | None = None,
) -> bytes:
    """
    Generate Excel file with reconciliation data for export.

    Returns Excel file content as bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        _logger.error("openpyxl not installed")
        raise ImportError("openpyxl is required for Excel generation")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đối soát"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    headers = [
        "Số Container",
        "Số cont chuẩn hóa",
        "Mã WO",
        "Mã TO",
        "Trạng thái",
        "Đã chốt",
        "Ngày chạy",
    ]
    ws.append(headers)

    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Query data
    wo_query = select(DeliveredTrip).where(DeliveredTrip.client_id == client_id)
    if date_from:
        wo_query = wo_query.where(DeliveredTrip.created_at >= date_from)
    if date_to:
        wo_query = wo_query.where(DeliveredTrip.created_at <= date_to)

    wo_result = await db.execute(wo_query)
    delivered_trips = wo_result.scalars().all()

    wo_ids = [wo.id for wo in delivered_trips]
    if wo_ids:
        wo_cont_result = await db.execute(
            select(DeliveredTripContainer).where(
                DeliveredTripContainer.delivered_trip_id.in_(wo_ids)
            )
        )
        wo_containers = wo_cont_result.scalars().all()

        # Group containers by work order
        wo_cont_map: dict[int, list[DeliveredTripContainer]] = {}
        for wc in wo_containers:
            if wc.delivered_trip_id not in wo_cont_map:
                wo_cont_map[wc.delivered_trip_id] = []
            wo_cont_map[wc.delivered_trip_id].append(wc)

        # Check for matched trip orders
        to_result = await db.execute(
            select(BookedTrip).where(BookedTrip.client_id == client_id)
        )
        booked_trips = to_result.scalars().all()

        to_map: dict[int, BookedTrip] = {to.id: to for to in booked_trips}

        # Query Reconciliation table for matches
        from app.models.domain import Reconciliation as ReconciliationModel
        join_result = await db.execute(
            select(ReconciliationModel).where(
                ReconciliationModel.delivered_trip_id.in_(wo_ids),
                ReconciliationModel.is_active == True,  # noqa: E712
            )
        )
        joins = join_result.scalars().all()

        wo_to_map: dict[int, int] = {j.delivered_trip_id: j.booked_trip_id for j in joins}

        # Add data rows
        row_num = 2
        for wo in delivered_trips:
            containers = wo_cont_map.get(wo.id, [])
            for wc in containers:
                to_id = wo_to_map.get(wo.id)
                to = to_map.get(to_id) if to_id else None

                status = "Chưa khớp"
                if to:
                    status = "Đã khớp"

                ws.append([
                    wc.container_number,
                    normalize_container_number(wc.container_number),
                    f"WO#{wo.id}",
                    f"TO#{to_id}" if to_id else "",
                    status,
                    "✓" if to else "",
                    wo.created_at.date() if wo.created_at else "",
                ])

                # Highlight duplicate containers
                if to:
                    cell = ws.cell(row=row_num, column=1)
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                row_num += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    content = excel_buffer.getvalue()
    wb.close()

    _logger.info(f"Generated Excel file with {row_num - 1} data rows")
    return content


# ---------------------------------------------------------------------------
# Trip Order Excel Import
# ---------------------------------------------------------------------------

# Vietnamese/English column name mappings
_TRIP_IMPORT_COLUMNS = {
    "ngay": "trip_date",
    "date": "trip_date",
    "ngay_chay": "trip_date",
    "trip_date": "trip_date",
    "ma_kh": "client_code",
    "client_code": "client_code",
    "khach_hang": "client_code",
    "ma_khach_hang": "client_code",
    "diem_lay": "pickup_location",
    "pickup": "pickup_location",
    "pickup_location": "pickup_location",
    "diem_tra": "dropoff_location",
    "dropoff": "dropoff_location",
    "dropoff_location": "dropoff_location",
    "cung_duong": "route",
    "route": "route",
    "so_cont": "container_number",
    "container": "container_number",
    "container_number": "container_number",
    "loai": "work_type",
    "work_type": "work_type",
    "loai_cont": "work_type",
    "don_gia": "revenue",
    "revenue": "revenue",
    "luong_tx": "driver_salary",
    "driver_salary": "driver_salary",
    "phu_cap": "allowance",
    "allowance": "allowance",
    "tac_nghiep": "operation_type",
    "operation_type": "operation_type",
    "operation": "operation_type",
}


def _normalize_trip_import_header(header: str) -> str:
    """Map Vietnamese/English header to standard field name."""
    if not header:
        return ""
    from app.utils.text import slugify_vi
    normalized = slugify_vi(header)
    return _TRIP_IMPORT_COLUMNS.get(normalized, normalized)


async def parse_booked_trip_excel(file_content: bytes) -> list[dict[str, Any]]:
    """Parse Excel file for batch trip order import.

    Expected columns (Vietnamese or English):
    - trip_date (Ngày/Ngày chạy)
    - client_code (Mã KH/Mã khách hàng)
    - pickup_location (Điểm lấy) + dropoff_location (Điểm trả) or route (Cung đường)
    - container_number (Số cont/Container)
    - work_type (Loại/Loại cont) — E20/E40/F20/F40
    - revenue (Đơn giá) — optional
    - driver_salary (Lương TX) — optional
    - allowance (Phụ cấp) — optional
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel processing")

    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active

    raw_headers = [cell.value for cell in sheet[1]]
    headers = [_normalize_trip_import_header(h) for h in raw_headers]

    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = value
        results.append(row_dict)

    workbook.close()
    _logger.info(f"Parsed {len(results)} trip order rows from Excel")
    return results


async def import_booked_trips(
    db: AsyncSession,
    rows: list[dict[str, Any]],
    user_id: int,
) -> dict[str, Any]:
    """Validate and create trip orders from parsed Excel rows.

    Groups rows by (trip_date, client_code) to form trip orders.
    Returns { created: int, errors: list[str] }.
    """
    from app.contexts.customer_pricing.infrastructure.pricing_lookup import (
        find_tiered_pricing,
    )

    created = 0
    errors: list[str] = []
    warnings: list[str] = []

    # Group rows by trip key
    groups: dict[tuple, list[dict]] = {}
    for i, row in enumerate(rows):
        trip_date = row.get("trip_date")
        client_code = row.get("client_code")

        if not trip_date or not client_code:
            errors.append(f"Dòng {i + 2}: thiếu ngày hoặc mã khách hàng")
            continue

        if isinstance(trip_date, str):
            try:
                trip_date = dt.strptime(trip_date, "%Y-%m-%d").date()
            except ValueError:
                try:
                    trip_date = dt.strptime(trip_date, "%d/%m/%Y").date()
                except ValueError:
                    errors.append(f"Dòng {i + 2}: định dạng ngày không hợp lệ '{trip_date}'")
                    continue
        elif hasattr(trip_date, "date"):
            pass  # already a date

        key = (str(trip_date), str(client_code))
        groups.setdefault(key, []).append(row)

    for key, group_rows in groups.items():
        trip_date, client_code_str = key

        # Look up partner by code
        client_result = await db.execute(
            select(Client).where(Client.code == client_code_str)
        )
        client = client_result.scalar_one_or_none()
        if not client:
            errors.append(f"Nhóm {key}: không tìm thấy đối tác mã '{client_code_str}'")
            continue

        first_row = group_rows[0]
        pickup = first_row.get("pickup_location")
        dropoff = first_row.get("dropoff_location")
        route = first_row.get("route") or (f"{pickup} - {dropoff}" if pickup and dropoff else "")

        # Build containers (normalize numbers — strip hyphens, uppercase)
        containers_data = []
        for row in group_rows:
            cn = normalize_container_number(str(row.get("container_number", "")).strip())
            wt = str(row.get("work_type", "")).strip().upper()
            if cn:
                valid, err = validate_container_number(cn)
                if not valid:
                    errors.append(f"Nhóm {key}: Container {cn} không hợp lệ — {err}")
                    continue
                containers_data.append({"container_number": cn, "work_type": wt or "E20"})

        if not containers_data:
            errors.append(f"Nhóm {key}: không có số container")
            continue

        work_type = containers_data[0]["work_type"]
        container_count = sum(1 for c in containers_data if c["work_type"] == work_type) or 1

        # Try auto-pricing
        revenue = int(first_row.get("revenue") or 0)
        driver_salary = int(first_row.get("driver_salary") or 0)
        allowance = int(first_row.get("allowance") or 0)
        pricing_id = None

        # Resolve pickup/dropoff strings to Location FKs via the alias
        # resolver. Auto-creates if missing.
        from app.contexts.customer_pricing.infrastructure.location_resolver import (
            LocationResolverService,
            ResolverSource,
        )
        resolver = LocationResolverService(db)
        pickup_id = None
        dropoff_id = None
        if pickup:
            r = await resolver.resolve_or_create(pickup, source=ResolverSource.MANUAL, user_id=user_id)
            pickup_id = r.location.id if r.location else None
        if dropoff:
            r = await resolver.resolve_or_create(dropoff, source=ResolverSource.MANUAL, user_id=user_id)
            dropoff_id = r.location.id if r.location else None
        if not pickup_id or not dropoff_id:
            errors.append(f"Nhóm {key}: pickup/dropoff không có")
            continue

        if not revenue:
            tiered = await find_tiered_pricing(
                db, client_id=client.id, work_type=work_type,
                quantity=container_count,
                pickup_location_id=pickup_id, dropoff_location_id=dropoff_id,
            )
            if tiered:
                revenue = tiered.revenue
                driver_salary = tiered.driver_salary
                allowance = tiered.allowance
                pricing_id = tiered.pricing.id

        operation_type_raw = first_row.get("operation_type")
        operation_type = parse_operation_type(operation_type_raw) if operation_type_raw else None

        booked_trip = BookedTrip(
            trip_date=trip_date_val,
            client_id=client.id,
            pickup_location_id=pickup_id,
            dropoff_location_id=dropoff_id,
            pricing_id=pricing_id,
            operation_type=operation_type,
            revenue=revenue,
            driver_salary=driver_salary,
            allowance=allowance,
            status="PENDING",
        )
        db.add(booked_trip)
        await db.flush()

        for c in containers_data:
            db.add(BookedTripContainer(
                booked_trip_id=booked_trip.id,
                container_number=c["container_number"],
                work_type=c["work_type"],
            ))

        created += 1

    await db.commit()
    return {"created": created, "errors": errors, "warnings": warnings}


def generate_booked_trip_template() -> bytes:
    """Generate a blank Excel template for trip order import."""
    headers = [
        "Ngày", "Mã KH", "Điểm lấy", "Điểm trả",
        "Cung đường", "Số cont", "Loại cont", "Tác nghiệp", "Đơn giá",
        "Lương TX", "Phụ cấp",
    ]
    examples = [
        "01/05/2026", "KH001", "Cát lái", "KC Bình Dương",
        "Cát lái - KC Bình Dương", "TCLU1234567", "E20", "Xuất / Nhập tàu", "1500000",
        "500000", "100000",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Nhập chuyến"
    sheet.append(headers)
    sheet.append(examples)

    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)

    buf = io.BytesIO()
    workbook.save(buf)
    workbook.close()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel Export for Work Orders, Trip Orders, Salary
# ---------------------------------------------------------------------------

async def generate_delivered_trips_excel(
    db: AsyncSession,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = None,
) -> bytes:
    """Export work orders to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = select(DeliveredTrip).order_by(DeliveredTrip.id.desc())
    if date_from:
        query = query.where(DeliveredTrip.created_at >= date_from)
    if date_to:
        query = query.where(DeliveredTrip.created_at <= date_to)
    if status:
        query = query.where(DeliveredTrip.status == status)

    result = await db.execute(query)
    delivered_trips = result.scalars().all()

    wo_ids = [wo.id for wo in delivered_trips]
    containers_map: dict[int, list[DeliveredTripContainer]] = {}
    if wo_ids:
        cont_result = await db.execute(
            select(DeliveredTripContainer).where(DeliveredTripContainer.delivered_trip_id.in_(wo_ids))
        )
        for c in cont_result.scalars().all():
            containers_map.setdefault(c.delivered_trip_id, []).append(c)

    # Resolve display names via JOIN (denormalized cols dropped).
    from app.models.domain import Client, Location, Vehicle
    from app.models.base import User as _User
    client_ids = {wo.client_id for wo in delivered_trips}
    driver_ids = {wo.driver_id for wo in delivered_trips}
    loc_ids = {wo.pickup_location_id for wo in delivered_trips} | {wo.dropoff_location_id for wo in delivered_trips}
    loc_ids.discard(None)
    client_name_by_id = {c.id: c.name for c in (await db.execute(select(Client).where(Client.id.in_(client_ids)))).scalars().all()} if client_ids else {}
    driver_name_by_id = {u.id: (u.full_name or u.username) for u in (await db.execute(select(_User).where(_User.id.in_(driver_ids)))).scalars().all()} if driver_ids else {}
    loc_name_by_id = {l.id: l.name for l in (await db.execute(select(Location).where(Location.id.in_(loc_ids)))).scalars().all()} if loc_ids else {}
    vehicle_ids = {wo.vehicle_id for wo in delivered_trips if wo.vehicle_id}
    vehicle_by_id = {v.id: v for v in (await db.execute(select(Vehicle).where(Vehicle.id.in_(vehicle_ids)))).scalars().all()} if vehicle_ids else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phiếu làm việc"

    headers = ["Mã WO", "Khách hàng", "Điểm lấy", "Điểm trả", "Tài xế", "Biển số", "Số tàu", "Số cont", "Loại", "Lương TX", "Phụ cấp", "Thu nhập", "Trạng thái", "Ngày tạo"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_labels = {"PENDING": "Chờ ghép", "MATCHED": "Đã đối soát"}
    for wo in delivered_trips:
        containers = containers_map.get(wo.id, [])
        plate = vehicle_by_id.get(wo.vehicle_id).plate if wo.vehicle_id and wo.vehicle_id in vehicle_by_id else ""
        for c in containers:
            ws.append([
                f"WO#{wo.id}", client_name_by_id.get(wo.client_id, ""),
                loc_name_by_id.get(wo.pickup_location_id, ""),
                loc_name_by_id.get(wo.dropoff_location_id, ""),
                driver_name_by_id.get(wo.driver_id, ""), plate,
                wo.vessel or "",
                c.container_number, c.cont_type,
                wo.driver_salary, wo.allowance, wo.driver_salary + wo.allowance,
                status_labels.get(wo.status, wo.status),
                wo.created_at.date() if wo.created_at else "",
            ])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


async def generate_booked_trips_excel(
    db: AsyncSession,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = None,
    client_id: int | None = None,
) -> bytes:
    """Export trip orders to Excel.

    When client_id is provided, filters to that client and includes
    match status columns for customer reconciliation.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = select(BookedTrip).order_by(BookedTrip.id.desc())
    if client_id:
        query = query.where(BookedTrip.client_id == client_id)
    if date_from:
        query = query.where(BookedTrip.trip_date >= date_from)
    if date_to:
        query = query.where(BookedTrip.trip_date <= date_to)
    if status:
        query = query.where(BookedTrip.status == status)

    result = await db.execute(query)
    booked_trips = result.scalars().all()

    to_ids = [to.id for to in booked_trips]
    containers_map: dict[int, list[BookedTripContainer]] = {}
    if to_ids:
        cont_result = await db.execute(
            select(BookedTripContainer).where(BookedTripContainer.booked_trip_id.in_(to_ids))
        )
        for c in cont_result.scalars().all():
            containers_map.setdefault(c.booked_trip_id, []).append(c)

    # Resolve display names via JOIN.
    from app.models.domain import Client, Location
    client_ids = {to.client_id for to in booked_trips}
    loc_ids = {to.pickup_location_id for to in booked_trips} | {to.dropoff_location_id for to in booked_trips}
    loc_ids.discard(None)
    client_name_by_id = {c.id: c.name for c in (await db.execute(select(Client).where(Client.id.in_(client_ids)))).scalars().all()} if client_ids else {}
    loc_name_by_id = {l.id: l.name for l in (await db.execute(select(Location).where(Location.id.in_(loc_ids)))).scalars().all()} if loc_ids else {}

    # For per-partner export: load match status and vehicle plates
    match_map: dict[int, str] = {}  # to_id -> "Đã khớp" / "Chưa khớp"
    plate_map: dict[int, str] = {}  # to_id -> plate
    vessel_map: dict[int, str] = {}  # to_id -> vessel
    client_name = None
    if client_id:
        # Get partner name for filename
        p_result = await db.execute(select(Client).where(Client.id == client_id))
        client_obj = p_result.scalar_one_or_none()
        client_name = client_obj.name if client_obj else None

        # Load reconciliation links to determine match status
        from app.models.domain import Reconciliation
        if to_ids:
            recon_result = await db.execute(
                select(Reconciliation.booked_trip_id).where(
                    Reconciliation.booked_trip_id.in_(to_ids),
                    Reconciliation.is_active == True,  # noqa: E712
                )
            )
            matched_to_ids = {r for (r,) in recon_result.all()}
            for to_id in to_ids:
                match_map[to_id] = "Đã khớp" if to_id in matched_to_ids else "Chưa khớp"

        # Load driver plates via work orders linked to these TOs
        from app.models.base import User as _User
        from app.models.domain import Vehicle
        if to_ids:
            # Get driver_ids from work orders linked via reconciliation
            recon_wo_result = await db.execute(
                select(Reconciliation.delivered_trip_id, Reconciliation.booked_trip_id).where(
                    Reconciliation.booked_trip_id.in_(to_ids),
                    Reconciliation.is_active == True,  # noqa: E712
                )
            )
            wo_to_pairs = recon_wo_result.all()
            wo_ids = list({r[0] for r in wo_to_pairs})

            if wo_ids:
                wo_result = await db.execute(
                    select(DeliveredTrip.id, DeliveredTrip.driver_id, DeliveredTrip.vessel).where(DeliveredTrip.id.in_(wo_ids))
                )
                wo_driver_map = {r[0]: r[1] for r in wo_result.all()}
                wo_vessel_map = {r[0]: (r[2] or "") for r in wo_result.all()}
                driver_ids = list(set(wo_driver_map.values()))

                if driver_ids:
                    v_result = await db.execute(
                        select(VehicleDriver.driver_id, Vehicle.plate)
                        .join(Vehicle, Vehicle.id == VehicleDriver.vehicle_id)
                        .where(
                            VehicleDriver.driver_id.in_(driver_ids),
                            VehicleDriver.is_active == True,  # noqa: E712
                            Vehicle.is_active == True,  # noqa: E712
                        )
                    )
                    driver_plate_map = dict(v_result.all())

                    for wo_id, to_id in wo_to_pairs:
                        driver_id = wo_driver_map.get(wo_id)
                        if driver_id and driver_id in driver_plate_map:
                            plate_map[to_id] = driver_plate_map[driver_id]
                        vessel = wo_vessel_map.get(wo_id, "")
                        if vessel:
                            vessel_map[to_id] = vessel

    wb = openpyxl.Workbook()
    ws = wb.active

    if client_id:
        ws.title = "Chuyến theo khách hàng"
        headers = ["STT", "Số container", "Tuyến đường", "Ngày chạy", "Biển số xe", "Số tàu", "Trạng thái khớp", "Đơn giá"]
    else:
        ws.title = "Đơn hàng"
        headers = ["Mã TO", "Ngày chạy", "Khách hàng", "Điểm lấy", "Điểm trả", "Số cont", "Loại", "Đơn giá", "Lương TX", "Phụ cấp", "Trạng thái"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_labels = {"PENDING": "Chờ ghép", "MATCHED": "Đã đối soát"}

    if client_id:
        stt = 0
        for to in booked_trips:
            containers = containers_map.get(to.id, [])
            pickup = loc_name_by_id.get(to.pickup_location_id, "")
            dropoff = loc_name_by_id.get(to.dropoff_location_id, "")
            route = f"{pickup} → {dropoff}" if pickup and dropoff else ""
            plate = plate_map.get(to.id, "")
            match_status = match_map.get(to.id, "Chưa khớp")
            for c in containers:
                stt += 1
                ws.append([
                    stt,
                    c.container_number,
                    route,
                    to.trip_date,
                    plate,
                    vessel_map.get(to.id, ""),
                    match_status,
                    to.revenue or "",
                ])
    else:
        for to in booked_trips:
            containers = containers_map.get(to.id, [])
            for c in containers:
                ws.append([
                    f"TO#{to.id}", to.trip_date,
                    client_name_by_id.get(to.client_id, ""),
                    loc_name_by_id.get(to.pickup_location_id, ""),
                    loc_name_by_id.get(to.dropoff_location_id, ""),
                    c.container_number, c.cont_type,
                    to.revenue, to.revenue, 0,
                    status_labels.get(to.status, to.status),
                ])

    from app.utils.excel_utils import add_template_version as _add_ver

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    _add_ver(ws, 13)
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue(), client_name


async def generate_doi_soat_excel(
    db: AsyncSession,
    client_id: int,
    date_from: str,
    date_to: str,
) -> tuple[bytes, str]:
    """Generate reconciliation (đối soát) Excel for a specific client.

    Returns (excel_bytes, client_name) tuple.
    Only includes MATCHED trip orders within the date range.
    Columns: STT | Ngày chạy | Số cont | Loại | Điểm lấy | Điểm trả | Tác nghiệp | Biển số xe | Số tàu | Đơn giá
    Summary row at the bottom: count per work type + total amount.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.domain import Client, Location, Reconciliation, Vehicle
    from app.models.base import User as _User

    # ── 1. Load partner ───────────────────────────────────────────────────────
    p_result = await db.execute(select(Client).where(Client.id == client_id))
    client = p_result.scalar_one_or_none()
    client_name = client.name if client else f"Client #{client_id}"

    from datetime import date as date_type

    df = date_type.fromisoformat(date_from)
    dt = date_type.fromisoformat(date_to)

    # ── 2. Load MATCHED trip orders ───────────────────────────────────────────
    to_query = select(BookedTrip).where(
        BookedTrip.client_id == client_id,
        BookedTrip.trip_date >= df,
        BookedTrip.trip_date <= dt,
        BookedTrip.status == "MATCHED",
    ).order_by(BookedTrip.trip_date, BookedTrip.id)
    to_result = await db.execute(to_query)
    booked_trips = to_result.scalars().all()

    to_ids = [to.id for to in booked_trips]

    # ── 3. Load containers ───────────────────────────────────────────────────
    containers_map: dict[int, list[BookedTripContainer]] = {}
    if to_ids:
        cont_result = await db.execute(
            select(BookedTripContainer).where(BookedTripContainer.booked_trip_id.in_(to_ids))
        )
        for c in cont_result.scalars().all():
            containers_map.setdefault(c.booked_trip_id, []).append(c)

    # ── 4. Load location names ───────────────────────────────────────────────
    loc_ids: set[int] = set()
    for to in booked_trips:
        if to.pickup_location_id:
            loc_ids.add(to.pickup_location_id)
        if to.dropoff_location_id:
            loc_ids.add(to.dropoff_location_id)
    loc_name_by_id: dict[int, str] = {}
    if loc_ids:
        loc_result = await db.execute(select(Location).where(Location.id.in_(loc_ids)))
        loc_name_by_id = {loc.id: loc.name for loc in loc_result.scalars().all()}

    # ── 5. Load plate + vessel + operation_type via Reconciliation → DeliveredTrip ─
    plate_map: dict[int, str] = {}   # to_id -> plate string
    vessel_map: dict[int, str] = {}  # to_id -> vessel string
    op_type_map: dict[int, str] = {} # to_id -> operation_type label
    _op_labels = {
        "XUAT_NHAP_TAU": "Xuất / Nhập tàu",
        "CHUYEN_BAI": "Chuyển bãi",
        "LAY_VO_HA_HANG": "Lấy vỏ hạ hàng",
        "CHAY_SA_LAN": "Chạy sà lan",
        "DONG_KHO": "Đóng kho",
        "TIEN_LUAT": "Tiền luật",
    }

    if to_ids:
        recon_result = await db.execute(
            select(Reconciliation.delivered_trip_id, Reconciliation.booked_trip_id).where(
                Reconciliation.booked_trip_id.in_(to_ids),
                Reconciliation.is_active == True,  # noqa: E712
            )
        )
        wo_to_pairs = recon_result.all()
        wo_ids = list({r[0] for r in wo_to_pairs})

        if wo_ids:
            wo_result = await db.execute(
                select(
                    DeliveredTrip.id, DeliveredTrip.driver_id, DeliveredTrip.vessel,
                    DeliveredTrip.vendor_id,
                    DeliveredTrip.operation_type,
                ).where(DeliveredTrip.id.in_(wo_ids))
            )
            wo_rows = wo_result.all()
            wo_driver_map   = {r[0]: r[1] for r in wo_rows}
            wo_vessel_map   = {r[0]: (r[2] or "") for r in wo_rows}
            wo_ext_plate_map = {r[0]: r[4] for r in wo_rows if r[3]}
            wo_op_type_map  = {r[0]: (r[4] or "") for r in wo_rows}
            driver_ids = [d for d in wo_driver_map.values() if d is not None]

            driver_plate_map: dict[int, str] = {}
            if driver_ids:
                v_result = await db.execute(
                    select(VehicleDriver.driver_id, Vehicle.plate)
                    .join(Vehicle, Vehicle.id == VehicleDriver.vehicle_id)
                    .where(
                        VehicleDriver.driver_id.in_(driver_ids),
                        VehicleDriver.is_active == True,  # noqa: E712
                        Vehicle.is_active == True,  # noqa: E712
                    )
                )
                driver_plate_map = dict(v_result.all())

            for wo_id, to_id in wo_to_pairs:
                if wo_id in wo_ext_plate_map and wo_ext_plate_map[wo_id]:
                    plate_map[to_id] = wo_ext_plate_map[wo_id]
                else:
                    d_id = wo_driver_map.get(wo_id)
                    if d_id and d_id in driver_plate_map:
                        plate_map[to_id] = driver_plate_map[d_id]
                vessel = wo_vessel_map.get(wo_id, "")
                if vessel:
                    vessel_map[to_id] = vessel
                op = wo_op_type_map.get(wo_id, "")
                if op:
                    op_type_map[to_id] = _op_labels.get(op, op)

    # ── 6. Build Excel workbook ───────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    month_label = df.strftime("%m/%Y")
    ws.title = f"SL T{df.month}.{str(df.year)[2:]}"

    num_cols = 15  # A–O
    last_col = get_column_letter(num_cols)

    # ── Styles ────────────────────────────────────────────────────────────────
    _bold = Font(bold=True, size=11)
    _bold14 = Font(bold=True, size=14)
    _header_font = Font(bold=True, size=11)
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # ── Title block (rows 1–8) ────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "CÔNG TY TNHH AMT PHÚC LỘC"
    ws["A1"].font = _bold

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Địa chỉ: Số 56B/97 đường Đoàn Kết, P. Hải An, TP Hải Phòng"
    ws["A2"].font = _bold

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = "MST: 0201965047"

    # Row 4 empty
    ws.merge_cells(f"A5:{last_col}5")
    ws["A5"] = f"BẢNG KÊ QUYẾT TOÁN CƯỚC VẬN CHUYỂN THÁNG {month_label}"
    ws["A5"].font = _bold14
    ws["A5"].alignment = center

    ws.merge_cells(f"A6:{last_col}6")
    ws["A6"] = f"Từ ngày {df.strftime('%d/%m/%Y')} đến ngày {dt.strftime('%d/%m/%Y')}"
    ws["A6"].font = _bold
    ws["A6"].alignment = center

    ws["A7"] = "KHÁCH HÀNG:"
    ws.merge_cells(f"C7:{last_col}7")
    ws["C7"] = client_name
    ws["C7"].font = _bold

    ws.merge_cells(f"A8:{last_col}8")
    ws["A8"] = "Công ty TNHH AMT Phúc Lộc xin gửi tới Quý Công ty bảng kê quyết toán vận tải như sau:"

    # ── Header row 10 ─────────────────────────────────────────────────────────
    headers = [
        "STT", "NGÀY ĐI", "CHỦ HÀNG", "SỐ CONTAINER",
        "F20'", "F40'", "E20'", "E40'",
        "SỐ XE CHẠY", "ĐIỂM ĐI", "ĐIỂM ĐẾN",
        "SỐ CHUYẾN", "CƯỚC CHUYẾN", "TỔNG TT", "TÁC NGHIỆP",
    ]
    ws.append([])  # row 9 empty
    ws.append(headers)  # row 10
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=10, column=col_num)
        cell.font = _header_font
        cell.alignment = center
        cell.border = thin_border

    # ── Subtotal row 11 ───────────────────────────────────────────────────────
    last_data_row = max(len(booked_trips) * 3 + 11, 12)  # rough estimate
    ws.append([
        "", "", "", "",
        f"=SUBTOTAL(9,E12:E{last_data_row})",
        f"=SUBTOTAL(9,F12:F{last_data_row})",
        f"=SUBTOTAL(9,G12:G{last_data_row})",
        f"=SUBTOTAL(9,H12:H{last_data_row})",
        "", "", "",
        f"=SUBTOTAL(9,L12:L{last_data_row})",
        "",
        f"=SUBTOTAL(9,N12:N{last_data_row})",
        "",
    ])
    for col_num in [5, 6, 7, 8, 12, 14]:
        ws.cell(row=11, column=col_num).font = _bold
        ws.cell(row=11, column=col_num).alignment = center

    # ── Data rows (12+) ──────────────────────────────────────────────────────
    stt = 0
    client_code = client.code if client else ""

    for to in booked_trips:
        containers = containers_map.get(to.id, [])
        pickup = loc_name_by_id.get(to.pickup_location_id or 0, "")
        dropoff = loc_name_by_id.get(to.dropoff_location_id or 0, "")
        plate = plate_map.get(to.id, "")
        op_type = op_type_map.get(to.id, "")
        revenue = to.revenue or 0
        trip_date_str = to.trip_date.strftime("%d/%m/%Y") if to.trip_date else ""

        for c in containers:
            stt += 1
            ct = (c.cont_type or "").upper()

            # Cont type flags
            f20 = 1 if ct == "F20" else None
            f40 = 1 if ct == "F40" else None
            e20 = 1 if ct == "E20" else None
            e40 = 1 if ct == "E40" else None

            row_num = 11 + stt
            ws.append([
                stt, trip_date_str, client_code, c.container_number,
                f20, f40, e20, e40,
                plate, pickup, dropoff,
                f"=E{row_num}+F{row_num}+G{row_num}+H{row_num}",
                revenue if revenue else None,
                f"=L{row_num}*M{row_num}",
                op_type,
            ])

            # Styling
            for col_num in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if col_num in (1, 2, 3, 5, 6, 7, 8, 9, 12, 15):
                    cell.alignment = center
                elif col_num == 4:
                    cell.alignment = left
                elif col_num in (10, 11):
                    cell.alignment = center
                elif col_num in (13, 14):
                    cell.alignment = right_align
                    cell.number_format = '#,##0'
                if col_num == 14:
                    cell.font = _bold

    # Update subtotal range to actual last row
    actual_last = 11 + stt if stt > 0 else 12
    for col_num, col_letter in [(5, "E"), (6, "F"), (7, "G"), (8, "H"), (12, "L"), (14, "N")]:
        ws.cell(row=11, column=col_num).value = f"=SUBTOTAL(9,{col_letter}12:{col_letter}{actual_last})"

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [6, 12, 12, 18, 6, 6, 6, 6, 14, 20, 20, 12, 14, 14, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header
    ws.freeze_panes = "A12"

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue(), client_name


async def parse_customer_response_excel(file_content: bytes) -> list[dict]:
    """Parse a customer's response Excel file.

    Expected columns (from our export template):
    - STT, Mã chuyến (trip_id), Ngày chạy, Số cont, Loại cont,
      Điểm lấy, Điểm trả, Tác nghiệp, Biển số xe, Số tàu, Đơn giá (VNĐ),
      Xác nhận KH, Ghi chú KH

    Returns list of dicts with:
      container_number, trip_date, customer_status, customer_note, customer_amount
    """
    import openpyxl
    from datetime import date as date_type
    from app.utils.iso6346 import normalize_container_number

    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active

    # Find header row — look for known headers
    header_row_idx = None
    headers = []
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        row_vals = [str(v or "").strip() for v in row]
        if "Số cont" in row_vals or "Xác nhận KH" in row_vals:
            header_row_idx = idx
            headers = row_vals
            break

    if header_row_idx is None:
        workbook.close()
        raise ValueError("Không tìm thấy header. File phải có cột 'Số cont' hoặc 'Xác nhận KH'.")

    # Map header names to column indices
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h] = i

    # Vietnamese status normalization
    _VI_STATUS = {
        "OK": "MATCHED", "KHỚP": "MATCHED", "KHOP": "MATCHED",
        "SỬA": "MATCHED", "SỬA_SỐ_TIỀN": "MATCHED",
        "TỪ CHỐI": "REJECTED", "TỪ_CHỐI": "REJECTED", "TỪCHỐI": "REJECTED",
        "KHÔNG": "REJECTED",
    }

    results: list[dict] = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue

        # Get values by header
        def _get(header_name: str) -> str | None:
            idx = col_map.get(header_name)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        container = _get("Số cont")
        trip_date_raw = _get("Ngày chạy")
        confirm = _get("Xác nhận KH")
        note = _get("Ghi chú KH")
        amount_raw = _get("Đơn giá (VNĐ)")

        # Skip rows without container or confirmation
        if not container and not confirm:
            continue

        # Parse trip_date
        trip_date = None
        if trip_date_raw:
            try:
                if isinstance(row[col_map.get("Ngày chạy", -1)], date_type):
                    trip_date = row[col_map.get("Ngày chạy", -1)]
                else:
                    trip_date = date_type.fromisoformat(trip_date_raw)
            except (ValueError, TypeError, IndexError):
                pass

        # Parse customer_status from confirmation column
        customer_status = "UNKNOWN"
        if confirm:
            confirm_upper = confirm.upper().strip()
            customer_status = _VI_STATUS.get(confirm_upper, "UNKNOWN")
            # If it's empty/unfilled, default to UNKNOWN
            if confirm_upper in ("", "-", "—"):
                customer_status = "UNKNOWN"

        # Parse amount
        customer_amount = None
        if amount_raw:
            try:
                customer_amount = int(str(amount_raw).replace(",", "").replace(".", ""))
            except ValueError:
                pass

        results.append({
            "container_number": container,
            "trip_date": trip_date,
            "customer_status": customer_status,
            "customer_note": note,
            "customer_amount": customer_amount,
        })

    workbook.close()
    return results


async def generate_salary_excel(
    db: AsyncSession,
    start_date: str,
    end_date: str,
) -> bytes:
    """Export driver earnings breakdown to Excel, computed on-the-fly from matched work orders."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models.domain import Reconciliation
    from datetime import date as _date

    from app.models.base import User as _User

    start_dt = _date.fromisoformat(start_date)
    end_dt = _date.fromisoformat(end_date)

    # Find matched work orders in the date range
    result = await db.execute(
        select(DeliveredTrip).where(
            DeliveredTrip.status == "MATCHED",
        ).order_by(DeliveredTrip.driver_id)
    )
    all_matched = result.scalars().all()

    # Filter by reconciliations and trip order date range
    # Get work order IDs that have active reconciliations with booked_trips in the date range
    matched_wo_ids = {wo.id for wo in all_matched}
    driver_earnings: dict[int, dict] = {}

    if matched_wo_ids:
        recon_result = await db.execute(
            select(Reconciliation, BookedTrip).join(
                BookedTrip, BookedTrip.id == Reconciliation.booked_trip_id
            ).where(
                Reconciliation.delivered_trip_id.in_(matched_wo_ids),
                Reconciliation.is_active == True,  # noqa: E712
                BookedTrip.trip_date >= start_dt,
                BookedTrip.trip_date <= end_dt,
            )
        )
        for recon, trip in recon_result.all():
            wo = next((w for w in all_matched if w.id == recon.delivered_trip_id), None)
            if wo is None:
                continue
            if wo.driver_id not in driver_earnings:
                driver_earnings[wo.driver_id] = {
                    "order_count": 0,
                    "total_salary": 0,
                    "total_allowance": 0,
                }
            driver_earnings[wo.driver_id]["order_count"] += 1
            driver_earnings[wo.driver_id]["total_salary"] += wo.driver_salary
            driver_earnings[wo.driver_id]["total_allowance"] += wo.allowance

    driver_ids = set(driver_earnings.keys())
    driver_name_by_id = (
        {u.id: (u.full_name or u.username) for u in
         (await db.execute(select(_User).where(_User.id.in_(driver_ids)))).scalars().all()}
        if driver_ids else {}
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng lương"

    headers = ["Tài xế", "Kỳ lương", "Số chuyến", "Tổng lương", "Phụ cấp", "Tổng thu nhập"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for driver_id, data in sorted(driver_earnings.items()):
        ws.append([
            driver_name_by_id.get(driver_id, ""),
            f"{start_date} → {end_date}",
            data["order_count"],
            data["total_salary"],
            data["total_allowance"],
            data["total_salary"] + data["total_allowance"],
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()
