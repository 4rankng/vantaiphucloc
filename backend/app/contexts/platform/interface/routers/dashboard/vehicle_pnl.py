"""Dashboard per-vehicle P&L endpoints (JSON + Excel export)."""

from typing import Optional

from fastapi import APIRouter, Depends, Query
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import (
    Vehicle,
    VehicleDriver,
    VehicleExpense,
    DeliveredTrip,
)
from app.models.base import User
from app.core.deps import require_permission
from app.database import get_db
from app.schemas.domain import (
    VehicleExpenseSummary,
    VehiclePnLResponse,
    VehiclePnLRow,
)

router = APIRouter()


@router.get("/vehicle-pnl", response_model=VehiclePnLResponse)
async def get_vehicle_pnl(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    vehicle_id: Optional[int] = Query(None, description="Filter to a single vehicle"),
    _current_user: User = Depends(require_permission("read", "Salary")),
    db: AsyncSession = Depends(get_db),
):
    """Per-vehicle P&L: Doanh thu − Chi phí = Lợi nhuận.

    For each vehicle returns:
      - Doanh thu: 0 (reconciliation table dropped; revenue-to-vehicle
        mapping not yet available).
      - CP Xe: vehicle_expenses subtotals (XANG_DAU, SUA_CHUA, TIEN_LUAT, KHAC).
      - CP Lương sản lượng: SUM(DeliveredTrip.driver_salary) for WOs
        on this vehicle.
      - CP Lương cơ bản: effective base salary × period for drivers attached to
        this vehicle via vehicle_drivers.
    """
    from datetime import datetime as _dt
    from app.contexts.payroll.infrastructure.repositories import (
        SqlDriverSalaryConfigRepository,
    )
    from app.contexts.payroll.domain.base_salary import effective_base_salary

    try:
        df = _dt.strptime(date_from, "%Y-%m-%d").date()
        dt = _dt.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=422, detail="Invalid date format. Use YYYY-MM-DD."
        )

    # ── 1. Build vehicle set ─────────────────────────────────────────────────
    veh_q = select(Vehicle.id, Vehicle.plate, Vehicle.vendor_id).where(
        Vehicle.is_active == True  # noqa: E712
    )
    if vehicle_id is not None:
        veh_q = veh_q.where(Vehicle.id == vehicle_id)
    veh_rows = (await db.execute(veh_q)).all()
    vehicles: dict[int, str] = {r[0]: r[1] for r in veh_rows}
    vendor_id_by_vehicle: dict[int, int | None] = {r[0]: r[2] for r in veh_rows}

    # Load vendor names for xe ngoai
    all_vendor_ids = list({r[2] for r in veh_rows if r[2] is not None})
    vendor_name_by_id: dict[int, str] = {}
    if all_vendor_ids:
        from app.models.domain import Vendor

        vnd_rows = (
            await db.execute(
                select(Vendor.id, Vendor.name).where(Vendor.id.in_(all_vendor_ids))
            )
        ).all()
        vendor_name_by_id = {r[0]: r[1] for r in vnd_rows}

    # ── 2. Revenue per vehicle from DeliveredTrip.revenue (matched trips) ────
    from app.core.pricing_lookup import TripPriceInfo, lookup_vendor_prices

    trip_detail_rows = (
        await db.execute(
            select(
                DeliveredTrip.id,
                DeliveredTrip.vendor_id,
                DeliveredTrip.pickup_location_id,
                DeliveredTrip.dropoff_location_id,
                DeliveredTrip.work_type,
                DeliveredTrip.cont_type,
                DeliveredTrip.revenue,
                Vehicle.id,
            )
            .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
            .where(
                DeliveredTrip.booked_trip_id.isnot(None),
                Vehicle.id.in_(list(vehicles.keys())),
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                >= df,
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                <= dt,
            )
        )
    ).all()

    vendor_trips = [
        TripPriceInfo(
            id=r[0],
            partner_id=r[1],
            pickup_location_id=r[2],
            dropoff_location_id=r[3],
            work_type=r[4],
            cont_type=r[5],
        )
        for r in trip_detail_rows
        if r[1] is not None
    ]

    vendor_prices = await lookup_vendor_prices(db, vendor_trips)

    revenue_by_vehicle: dict[int, int] = {}
    vendor_cost_by_vehicle: dict[int, int] = {}
    for r in trip_detail_rows:
        trip_id, vid, trip_rev = r[0], r[7], int(r[6] or 0)
        if vid:
            revenue_by_vehicle[vid] = revenue_by_vehicle.get(vid, 0) + trip_rev
            vendor_cost_by_vehicle[vid] = vendor_cost_by_vehicle.get(
                vid, 0
            ) + vendor_prices.get(trip_id, 0)

    # ── 3. CP Lương sản lượng per vehicle ───────────────────────────────────
    # vehicle_id FK removed; join via vehicle_plate
    {plate: vid for vid, plate in vehicles.items()}
    wo_salary_rows = (
        await db.execute(
            select(
                Vehicle.id,
                func.coalesce(func.sum(DeliveredTrip.driver_salary), 0),
            )
            .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
            .where(
                DeliveredTrip.booked_trip_id.isnot(None),
                Vehicle.id.in_(list(vehicles.keys())),
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                >= df,
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                <= dt,
            )
            .group_by(Vehicle.id)
        )
    ).all()

    salary_by_vehicle: dict[int, int] = {}
    for vid, sal in wo_salary_rows:
        if vid:
            salary_by_vehicle[vid] = int(sal or 0)

    # ── 3b. Trip count per vehicle for CP Chung allocation ─────────
    wo_count_rows = (
        await db.execute(
            select(
                Vehicle.id,
                func.count(DeliveredTrip.id),
            )
            .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
            .where(
                DeliveredTrip.booked_trip_id.isnot(None),
                Vehicle.id.in_(list(vehicles.keys())),
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                >= df,
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                <= dt,
            )
            .group_by(Vehicle.id)
        )
    ).all()

    trip_count_by_vehicle: dict[int, int] = {}
    total_trips = 0
    for vid, cnt in wo_count_rows:
        if vid:
            trip_count_by_vehicle[vid] = int(cnt)
            total_trips += int(cnt)

    # ── 4. CP Xe (vehicle expenses) per vehicle ──────────────────────────────
    expense_rows = (
        await db.execute(
            select(
                VehicleExpense.vehicle_id,
                VehicleExpense.category,
                func.coalesce(func.sum(VehicleExpense.amount), 0),
            )
            .where(
                VehicleExpense.expense_date >= df,
                VehicleExpense.expense_date <= dt,
                VehicleExpense.category.in_(
                    ["XANG_DAU", "SUA_CHUA", "TIEN_LUAT", "KHAC"]
                ),
            )
            .group_by(VehicleExpense.vehicle_id, VehicleExpense.category)
        )
    ).all()

    cp_xe_by_vehicle: dict[int, dict[str, int]] = {}
    for vid, cat, total_amt in expense_rows:
        amt = int(total_amt or 0)
        if vid and vid in vehicles:
            slot = cp_xe_by_vehicle.setdefault(
                vid, {"XANG_DAU": 0, "SUA_CHUA": 0, "TIEN_LUAT": 0, "KHAC": 0}
            )
            slot[cat] = slot.get(cat, 0) + amt

    # ── 5. CP Lương cơ bản: drivers attached to each vehicle via vehicle_drivers
    vd_rows = (
        await db.execute(
            select(VehicleDriver.vehicle_id, VehicleDriver.driver_id).where(
                VehicleDriver.vehicle_id.in_(list(vehicles.keys())),
                VehicleDriver.is_active == True,  # noqa: E712
                VehicleDriver.effective_from <= dt,
            )
        )
    ).all()

    vehicle_driver_map: dict[int, list[int]] = {}
    for vid, did in vd_rows:
        vehicle_driver_map.setdefault(vid, []).append(did)

    all_driver_ids = list(
        {d for drivers in vehicle_driver_map.values() for d in drivers}
    )
    base_salary_repo = SqlDriverSalaryConfigRepository(db)
    history_by_driver = (
        await base_salary_repo.list_history_for_drivers(all_driver_ids)
        if all_driver_ids
        else {}
    )

    base_salary_by_vehicle: dict[int, int] = {}
    for vid, driver_ids in vehicle_driver_map.items():
        total_base = sum(
            effective_base_salary(history_by_driver.get(did, []), dt)
            for did in driver_ids
        )
        base_salary_by_vehicle[vid] = total_base

    # ── 6. Assemble rows ─────────────────────────────────────────────────────
    rows: list[VehiclePnLRow] = []
    total_revenue = 0
    sum_row_profits = 0

    for vid, plate in sorted(vehicles.items(), key=lambda x: x[1]):
        rev = revenue_by_vehicle.get(vid, 0)
        sal = salary_by_vehicle.get(vid, 0)
        base = base_salary_by_vehicle.get(vid, 0)
        vcost = vendor_cost_by_vehicle.get(vid, 0)
        xe_cats = cp_xe_by_vehicle.get(vid, {})
        xe_summary = VehicleExpenseSummary(
            xang_dau=xe_cats.get("XANG_DAU", 0),
            sua_chua=xe_cats.get("SUA_CHUA", 0),
            tien_luat=xe_cats.get("TIEN_LUAT", 0),
            khac=xe_cats.get("KHAC", 0),
            total=sum(xe_cats.values()),
        )
        loi_nhuan = rev - (xe_summary.total + sal + base + vcost)
        vnd_id = vendor_id_by_vehicle.get(vid)
        rows.append(
            VehiclePnLRow(
                vehicle_id=vid,
                plate=plate,
                is_vendor=vnd_id is not None,
                vendor_name=vendor_name_by_id.get(vnd_id)
                if vnd_id is not None
                else None,
                revenue=rev,
                cp_xe=xe_summary,
                cp_luong_san_luong=sal,
                cp_luong_co_ban=base,
                cp_vendor=vcost,
                loi_nhuan=loi_nhuan,
            )
        )
        total_revenue += rev
        sum_row_profits += loi_nhuan

    total_profit = sum_row_profits

    return VehiclePnLResponse(
        date_from=df,
        date_to=dt,
        rows=rows,
        total_revenue=total_revenue,
        total_profit=total_profit,
    )


@router.get("/vehicle-pnl/export")
async def export_vehicle_pnl(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    vehicle_id: Optional[int] = Query(None, description="Filter to a single vehicle"),
    _current_user: User = Depends(require_permission("read", "Salary")),
    db: AsyncSession = Depends(get_db),
):
    """Export per-vehicle P&L as an Excel (.xlsx) file."""
    from fastapi import HTTPException
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    from app.contexts.payroll.infrastructure.repositories import (
        SqlDriverSalaryConfigRepository,
    )
    from app.contexts.payroll.domain.base_salary import effective_base_salary
    from app.core.pricing_lookup import TripPriceInfo, lookup_vendor_prices
    from app.utils.excel_utils import workbook_to_bytes

    try:
        df = _dt.strptime(date_from, "%Y-%m-%d").date()
        dt = _dt.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(
            status_code=422, detail="Invalid date format. Use YYYY-MM-DD."
        )

    # ── Re-use the same data logic as vehicle-pnl ────────────────────────────
    veh_q = select(Vehicle.id, Vehicle.plate).where(Vehicle.is_active == True)  # noqa: E712
    if vehicle_id is not None:
        veh_q = veh_q.where(Vehicle.id == vehicle_id)
    veh_rows = (await db.execute(veh_q)).all()
    vehicles: dict[int, str] = {r[0]: r[1] for r in veh_rows}

    trip_detail_rows = (
        await db.execute(
            select(
                DeliveredTrip.id,
                DeliveredTrip.vendor_id,
                DeliveredTrip.pickup_location_id,
                DeliveredTrip.dropoff_location_id,
                DeliveredTrip.work_type,
                DeliveredTrip.cont_type,
                DeliveredTrip.revenue,
                Vehicle.id,
            )
            .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
            .where(
                DeliveredTrip.booked_trip_id.isnot(None),
                Vehicle.id.in_(list(vehicles.keys())),
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                >= df,
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                <= dt,
            )
        )
    ).all()

    vendor_trips = [
        TripPriceInfo(
            id=r[0],
            partner_id=r[1],
            pickup_location_id=r[2],
            dropoff_location_id=r[3],
            work_type=r[4],
            cont_type=r[5],
        )
        for r in trip_detail_rows
        if r[1] is not None
    ]
    vendor_prices = await lookup_vendor_prices(db, vendor_trips)

    revenue_by_vehicle: dict[int, int] = {}
    vendor_cost_by_vehicle: dict[int, int] = {}
    for r in trip_detail_rows:
        trip_id, vid, trip_rev = r[0], r[7], int(r[6] or 0)
        if vid:
            revenue_by_vehicle[vid] = revenue_by_vehicle.get(vid, 0) + trip_rev
            vendor_cost_by_vehicle[vid] = vendor_cost_by_vehicle.get(
                vid, 0
            ) + vendor_prices.get(trip_id, 0)

    wo_salary_rows = (
        await db.execute(
            select(
                Vehicle.id,
                func.coalesce(func.sum(DeliveredTrip.driver_salary), 0),
            )
            .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
            .where(
                DeliveredTrip.booked_trip_id.isnot(None),
                Vehicle.id.in_(list(vehicles.keys())),
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                >= df,
                func.coalesce(
                    DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)
                )
                <= dt,
            )
            .group_by(Vehicle.id)
        )
    ).all()
    salary_by_vehicle: dict[int, int] = {
        vid: int(sal or 0) for vid, sal in wo_salary_rows if vid
    }

    expense_rows = (
        await db.execute(
            select(
                VehicleExpense.vehicle_id,
                VehicleExpense.category,
                func.coalesce(func.sum(VehicleExpense.amount), 0),
            )
            .where(
                VehicleExpense.expense_date >= df,
                VehicleExpense.expense_date <= dt,
                VehicleExpense.category.in_(
                    ["XANG_DAU", "SUA_CHUA", "TIEN_LUAT", "KHAC"]
                ),
            )
            .group_by(VehicleExpense.vehicle_id, VehicleExpense.category)
        )
    ).all()

    cp_xe_by_vehicle: dict[int, dict[str, int]] = {}
    for vid, cat, total_amt in expense_rows:
        if vid and vid in vehicles:
            slot = cp_xe_by_vehicle.setdefault(
                vid, {"XANG_DAU": 0, "SUA_CHUA": 0, "TIEN_LUAT": 0, "KHAC": 0}
            )
            slot[cat] = slot.get(cat, 0) + int(total_amt or 0)

    vd_rows = (
        await db.execute(
            select(VehicleDriver.vehicle_id, VehicleDriver.driver_id).where(
                VehicleDriver.vehicle_id.in_(list(vehicles.keys())),
                VehicleDriver.is_active == True,  # noqa: E712
                VehicleDriver.effective_from <= dt,
            )
        )
    ).all()
    vehicle_driver_map: dict[int, list[int]] = {}
    for vid, did in vd_rows:
        vehicle_driver_map.setdefault(vid, []).append(did)

    all_driver_ids = list(
        {d for drivers in vehicle_driver_map.values() for d in drivers}
    )
    base_salary_repo = SqlDriverSalaryConfigRepository(db)
    history_by_driver = (
        await base_salary_repo.list_history_for_drivers(all_driver_ids)
        if all_driver_ids
        else {}
    )

    base_salary_by_vehicle: dict[int, int] = {
        vid: sum(
            effective_base_salary(history_by_driver.get(did, []), dt)
            for did in driver_ids
        )
        for vid, driver_ids in vehicle_driver_map.items()
    }

    # ── Build Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo P&L"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill("solid", fgColor="E8F0FE")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    profit_pos_fill = PatternFill("solid", fgColor="D1FAE5")
    profit_neg_fill = PatternFill("solid", fgColor="FEE2E2")
    profit_pos_font = Font(bold=True, color="065F46", size=11)
    profit_neg_font = Font(bold=True, color="991B1B", size=11)

    # Title row
    period_label = f"{df.strftime('%d/%m/%Y')} – {dt.strftime('%d/%m/%Y')}"
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"BÁO CÁO LỢI NHUẬN THEO XE  |  Kỳ: {period_label}"
    title_cell.font = Font(bold=True, size=13, color="1E3A5F")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    ws.append([])  # blank row 2

    # Header row (row 3)
    headers = [
        "Biển số",
        "Doanh thu",
        "CP Xăng dầu",
        "CP Sửa chữa",
        "CP Tiền luật",
        "CP Khác",
        "Lương LX",
        "Tổng CP",
        "Lợi nhuận",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 22

    # Number format: thousand-separator, no decimals, ₫ suffix
    VND_FMT = "#,##0\\ [$₫-vi-VN]"

    def set_num(cell, value: int) -> None:
        """Write a real integer and apply VND display format."""
        cell.value = value
        cell.number_format = VND_FMT
        cell.alignment = right

    # Data rows
    data_start_row = 4
    total_rev = total_luong = total_xang = total_sua = total_luat = total_khac = (
        total_xe
    ) = total_profit = 0

    sorted_vehicles = sorted(vehicles.items(), key=lambda x: x[1])
    for row_idx, (vid, plate) in enumerate(sorted_vehicles, start=data_start_row):
        rev = revenue_by_vehicle.get(vid, 0)
        sal = salary_by_vehicle.get(vid, 0)
        base = base_salary_by_vehicle.get(vid, 0)
        vcost = vendor_cost_by_vehicle.get(vid, 0)
        xe_cats = cp_xe_by_vehicle.get(vid, {})
        xang = xe_cats.get("XANG_DAU", 0)
        sua = xe_cats.get("SUA_CHUA", 0)
        luat = xe_cats.get("TIEN_LUAT", 0)
        khac = xe_cats.get("KHAC", 0)
        xe_total = xang + sua + luat + khac
        luong = sal + base
        total_cp = xe_total + luong + vcost
        profit = rev - total_cp

        total_rev += rev
        total_luong += luong
        total_xang += xang
        total_sua += sua
        total_luat += luat
        total_khac += khac
        total_xe += xe_total
        total_profit += profit

        # Append plate first so row exists, then fill numeric cells
        ws.append([plate])
        num_values = [rev, xang, sua, luat, khac, luong, total_cp, profit]
        for col_offset, val in enumerate(num_values, start=2):
            set_num(ws.cell(row=row_idx, column=col_offset), val)

        for col_idx in range(1, len(num_values) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 1:
                cell.alignment = center
                cell.font = Font(bold=True)
            elif col_idx == len(num_values) + 1:
                # Profit column — colour by sign
                if profit >= 0:
                    cell.fill = profit_pos_fill
                    cell.font = profit_pos_font
                else:
                    cell.fill = profit_neg_fill
                    cell.font = profit_neg_font

        ws.row_dimensions[row_idx].height = 18

    # Total row
    total_row_idx = data_start_row + len(sorted_vehicles)
    total_cp_all = total_xe + total_luong
    ws.append(["TỔNG"])
    total_num_values = [
        total_rev,
        total_xang,
        total_sua,
        total_luat,
        total_khac,
        total_luong,
        total_cp_all,
        total_profit,
    ]
    for col_offset, val in enumerate(total_num_values, start=2):
        set_num(ws.cell(row=total_row_idx, column=col_offset), val)

    for col_idx in range(1, len(total_num_values) + 2):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = right if col_idx > 1 else center
    ws.row_dimensions[total_row_idx].height = 20

    # Column widths
    col_widths = [14, 16, 15, 15, 15, 13, 15, 15, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Footer
    footer_row = total_row_idx + 2
    ws.cell(
        row=footer_row, column=1
    ).value = f"Xuất ngày: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="888888", size=9)

    # Stream response
    from io import BytesIO

    buf = BytesIO(workbook_to_bytes(wb))
    filename = f"PnL_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
