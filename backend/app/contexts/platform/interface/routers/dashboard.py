"""
Dashboard API — aggregated summary using SQL, not client-side computation.
"""

import json
import logging
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import Client, BookedTrip, Vehicle, VehicleDriver, VehicleExpense, DeliveredTrip, Vendor, Location
from app.models.base import User
from app.core.deps import get_current_user, require_permission
from app.core.worker import get_arq_pool
from app.database import get_db
from app.schemas.domain import (
    DashboardSummaryOut,
    KpiTrendDeltas,
    KpiTrendsOut,
    TripDailyStatsOut,
    VehicleExpenseSummary,
    VehiclePnLResponse,
    VehiclePnLRow,
    VehiclePnLGroup,
    DirectorDashboardOut,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/dashboard", tags=["dashboard"])


@router.get("/summary", response_model=DashboardSummaryOut)
async def get_dashboard_summary(
    _current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD"),
):
    parsed_from = None
    parsed_to = None
    if date_from:
        try:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            pass
    if date_to:
        try:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            pass

    # Revenue: sum of revenue from matched DeliveredTrips
    revenue_query = select(func.coalesce(func.sum(DeliveredTrip.revenue), 0)).where(
        DeliveredTrip.booked_trip_id.isnot(None)
    )
    if parsed_from:
        revenue_query = revenue_query.where(
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= parsed_from
        )
    if parsed_to:
        revenue_query = revenue_query.where(
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) < parsed_to + timedelta(days=1)
        )
    revenue_q = await db.execute(revenue_query)
    total_revenue = revenue_q.scalar() or 0

    # Expense: sum of driver_salary per work order
    expense_query = select(func.coalesce(func.sum(DeliveredTrip.driver_salary), 0))
    if parsed_from:
        expense_query = expense_query.where(DeliveredTrip.created_at >= parsed_from)
    if parsed_to:
        expense_query = expense_query.where(DeliveredTrip.created_at < parsed_to + timedelta(days=1))
    expense_q = await db.execute(expense_query)
    total_expense = expense_q.scalar() or 0

    _wo_day = func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at))

    trip_count_query = select(func.count(DeliveredTrip.id))
    if parsed_from:
        trip_count_query = trip_count_query.where(_wo_day >= parsed_from)
    if parsed_to:
        trip_count_query = trip_count_query.where(_wo_day < parsed_to + timedelta(days=1))
    trip_count_q = await db.execute(trip_count_query)
    trip_count = trip_count_q.scalar() or 0

    active_query = select(func.count(DeliveredTrip.id)).where(
        DeliveredTrip.booked_trip_id.isnot(None)
    )
    if parsed_from:
        active_query = active_query.where(_wo_day >= parsed_from)
    if parsed_to:
        active_query = active_query.where(_wo_day < parsed_to + timedelta(days=1))
    active_q = await db.execute(active_query)
    active_trips = active_q.scalar() or 0

    # outstanding_debt: confirmed receivables (matched DeliveredTrips)
    debt_query = select(func.coalesce(func.sum(DeliveredTrip.revenue), 0)).where(
        DeliveredTrip.booked_trip_id.isnot(None)
    )
    if parsed_from:
        debt_query = debt_query.where(
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= parsed_from
        )
    if parsed_to:
        debt_query = debt_query.where(
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) < parsed_to + timedelta(days=1)
        )
    outstanding_debt = (await db.execute(debt_query)).scalar() or 0

    # Use vehicle_drivers join for plate lookup so multi-driver vehicles
    # show the correct plate per driver. Falls back to NULL plate gracefully
    # when no vehicle_drivers row exists for the driver.
    driver_salary_q = await db.execute(
        select(
            User.id.label("driver_id"),
            User.username.label("driver_name"),
            Vehicle.plate.label("vehicle_plate"),
            func.count(DeliveredTrip.id).label("total_jobs"),
            func.coalesce(func.sum(DeliveredTrip.driver_salary), 0).label("total_salary"),
        )
        .join(
            VehicleDriver,
            (VehicleDriver.driver_id == User.id) & (VehicleDriver.is_active == True),  # noqa: E712
            isouter=True,
        )
        .join(Vehicle, Vehicle.id == VehicleDriver.vehicle_id, isouter=True)
        .join(DeliveredTrip, DeliveredTrip.driver_id == User.id)
        .where(DeliveredTrip.booked_trip_id.isnot(None))
        .group_by(User.id, User.username, Vehicle.plate)
    )
    driver_salary_summary = [
        {
            "driver_id": row.driver_id,
            "driver_name": row.driver_name,
            "vehicle_plate": row.vehicle_plate,
            "total_jobs": row.total_jobs,
            "total_salary": row.total_salary,
        }
        for row in driver_salary_q.all()
    ]

    unmatched_q = await db.execute(
        select(func.count(DeliveredTrip.id)).where(
            DeliveredTrip.booked_trip_id.is_(None)
        )
    )
    unmatched_delivered_trip_count = unmatched_q.scalar() or 0

    pending_q = await db.execute(
        select(func.count(DeliveredTrip.id)).where(
            DeliveredTrip.booked_trip_id.is_(None)
        )
    )
    pending_trip_count = pending_q.scalar() or 0

    return DashboardSummaryOut(
        total_revenue=total_revenue,
        total_expense=total_expense,
        trip_count=trip_count,
        active_trips=active_trips,
        outstanding_debt=outstanding_debt,
        driver_salary_summary=driver_salary_summary,
        unmatched_delivered_trip_count=unmatched_delivered_trip_count,
        pending_trip_count=pending_trip_count,
    )


@router.get("/kpi-trends", response_model=KpiTrendsOut)
async def get_kpi_trends(
    days: int = Query(12, ge=2, le=90, description="Number of trailing days, including end_date"),
    end_date: Optional[str] = Query(None, description="YYYY-MM-DD; defaults to today (UTC)"),
    _current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Daily activity time-series powering accountant KPI sparklines.

    Returns parallel arrays (length == ``days``) for unmatched work orders,
    pending trips, driver salary expense, and revenue. Also returns a
    second-half-vs-first-half percent delta per series for the trend pill.
    """
    from datetime import date as _date, timezone as _tz
    from fastapi import HTTPException

    # ── 1. Resolve window ────────────────────────────────────────────────────
    if end_date:
        try:
            parsed_end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=422, detail="Invalid end_date. Use YYYY-MM-DD.")
    else:
        parsed_end = datetime.now(_tz.utc).date()

    start_date = parsed_end - timedelta(days=days - 1)
    labels: list[_date] = [start_date + timedelta(days=i) for i in range(days)]

    def _normalize(d):
        """Coerce SQL date/datetime/string to Python date."""
        if d is None:
            return None
        if isinstance(d, str):
            try:
                return datetime.strptime(d[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
        if isinstance(d, datetime):
            return d.date()
        return d  # already date

    # ── 2. DeliveredTrip per-day expression ──────────────────────────────────────
    # Use explicit trip_date when set, else fall back to created_at's date.
    wo_day = func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at))

    # ── 3. Unmatched work orders (PENDING, dated in window) ──────────────────
    wo_pending_rows = (await db.execute(
        select(wo_day.label("d"), func.count(DeliveredTrip.id))
        .where(
            wo_day >= start_date,
            wo_day <= parsed_end,
            DeliveredTrip.booked_trip_id.is_(None),
        )
        .group_by(wo_day)
    )).all()
    wo_pending_map: dict = {_normalize(r[0]): int(r[1]) for r in wo_pending_rows}

    # ── 4. Pending trips (UNMATCHED DeliveredTrips in window) ──────────────────
    trip_pending_rows = (await db.execute(
        select(wo_day.label("d"), func.count(DeliveredTrip.id))
        .where(
            wo_day >= start_date,
            wo_day <= parsed_end,
            DeliveredTrip.booked_trip_id.is_(None),
        )
        .group_by(wo_day)
    )).all()
    trip_pending_map: dict = {_normalize(r[0]): int(r[1]) for r in trip_pending_rows}

    # ── 5. Driver salary per day (MATCHED WOs) ───────────────────────────────
    salary_rows = (await db.execute(
        select(
            wo_day.label("d"),
            func.coalesce(func.sum(DeliveredTrip.driver_salary), 0),
        )
        .where(
            wo_day >= start_date,
            wo_day <= parsed_end,
            DeliveredTrip.booked_trip_id.isnot(None),
        )
        .group_by(wo_day)
    )).all()
    salary_map: dict = {_normalize(r[0]): int(r[1] or 0) for r in salary_rows}

    # ── 6. Revenue per day (matched DeliveredTrips dated in window) ───────────
    revenue_rows = (await db.execute(
        select(
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)).label("d"),
            func.coalesce(func.sum(DeliveredTrip.revenue), 0),
        )
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),  # noqa: E712
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= start_date,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= parsed_end,
        )
        .group_by(
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at))
        )
    )).all()
    revenue_map: dict = {_normalize(r[0]): int(r[1] or 0) for r in revenue_rows}

    def _fill(m: dict) -> list[int]:
        return [m.get(d, 0) for d in labels]

    unmatched_series = _fill(wo_pending_map)
    pending_series = _fill(trip_pending_map)
    salary_series = _fill(salary_map)
    revenue_series = _fill(revenue_map)

    def _delta_pct(series: list[int]) -> float:
        """Trend % = (mean of 2nd half − mean of 1st half) / mean of 1st half × 100."""
        if len(series) < 2:
            return 0.0
        mid = len(series) // 2
        first_half = series[:mid]
        last_half = series[mid:]
        first_mean = sum(first_half) / max(len(first_half), 1)
        last_mean = sum(last_half) / max(len(last_half), 1)
        if first_mean == 0:
            return 0.0 if last_mean == 0 else 100.0
        return round(((last_mean - first_mean) / first_mean) * 100, 1)

    return KpiTrendsOut(
        end_date=parsed_end,
        days=days,
        labels=[d.isoformat() for d in labels],
        unmatched_delivered_trips=unmatched_series,
        pending_trips=pending_series,
        driver_salary=salary_series,
        revenue=revenue_series,
        deltas=KpiTrendDeltas(
            unmatched_delivered_trips=_delta_pct(unmatched_series),
            pending_trips=_delta_pct(pending_series),
            driver_salary=_delta_pct(salary_series),
            revenue=_delta_pct(revenue_series),
        ),
    )


@router.get("/vehicle-pnl", response_model=VehiclePnLResponse)
async def get_vehicle_pnl(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    vehicle_id: Optional[int] = Query(None, description="Filter to a single vehicle"),
    _current_user: User = Depends(require_permission("read", "Salary")),
    db: AsyncSession = Depends(get_db),
):
    """Per-vehicle P&L: Doanh thu − Chi phí = Lợi nhuận.

    For each vehicle returns:
      - Doanh thu: 0 (reconciliation table dropped; revenue-to-vehicle
        mapping not yet available).
      - CP Xe: vehicle_expenses subtotals (XANG_DAU, SUA_CHUA, TIEN_LUAT, KHAC).
      - CP Lương sản lượng: SUM(DeliveredTrip.driver_salary) for WOs
        on this vehicle.
      - CP Lương cơ bản: effective base salary × period for drivers attached to
        this vehicle via vehicle_drivers.
    """
    from datetime import datetime as _dt
    from app.contexts.payroll.infrastructure.repositories import SqlDriverSalaryConfigRepository
    from app.contexts.payroll.domain.base_salary import effective_base_salary

    try:
        df = _dt.strptime(date_from, "%Y-%m-%d").date()
        dt = _dt.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        from fastapi import HTTPException
        raise HTTPException(status_code=422, detail="Invalid date format. Use YYYY-MM-DD.")

    # ── 1. Build vehicle set ─────────────────────────────────────────────────
    veh_q = select(Vehicle.id, Vehicle.plate, Vehicle.vendor_id).where(Vehicle.is_active == True)  # noqa: E712
    if vehicle_id is not None:
        veh_q = veh_q.where(Vehicle.id == vehicle_id)
    veh_rows = (await db.execute(veh_q)).all()
    vehicles: dict[int, str] = {r[0]: r[1] for r in veh_rows}
    vendor_id_by_vehicle: dict[int, int | None] = {r[0]: r[2] for r in veh_rows}

    # Load vendor names for xe ngoai
    all_vendor_ids = list({r[2] for r in veh_rows if r[2] is not None})
    vendor_name_by_id: dict[int, str] = {}
    if all_vendor_ids:
        vnd_rows = (await db.execute(
            select(Vendor.id, Vendor.name).where(Vendor.id.in_(all_vendor_ids))
        )).all()
        vendor_name_by_id = {r[0]: r[1] for r in vnd_rows}

    # ── 2. Revenue per vehicle from DeliveredTrip.revenue (matched trips) ────
    from app.core.pricing_lookup import TripPriceInfo, lookup_vendor_prices

    trip_detail_rows = (await db.execute(
        select(
            DeliveredTrip.id,
            DeliveredTrip.vendor_id,
            DeliveredTrip.pickup_location_id,
            DeliveredTrip.dropoff_location_id,
            DeliveredTrip.work_type,
            DeliveredTrip.cont_type,
            DeliveredTrip.revenue,
            Vehicle.id,
        )
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
    )).all()

    vendor_trips = [
        TripPriceInfo(id=r[0], partner_id=r[1], pickup_location_id=r[2], dropoff_location_id=r[3], work_type=r[4], cont_type=r[5])
        for r in trip_detail_rows if r[1] is not None
    ]

    vendor_prices = await lookup_vendor_prices(db, vendor_trips)

    revenue_by_vehicle: dict[int, int] = {}
    vendor_cost_by_vehicle: dict[int, int] = {}
    for r in trip_detail_rows:
        trip_id, vid, trip_rev = r[0], r[7], int(r[6] or 0)
        if vid:
            revenue_by_vehicle[vid] = revenue_by_vehicle.get(vid, 0) + trip_rev
            vendor_cost_by_vehicle[vid] = vendor_cost_by_vehicle.get(vid, 0) + vendor_prices.get(trip_id, 0)

    # ── 3. CP Lương sản lượng per vehicle ───────────────────────────────────
    # vehicle_id FK removed; join via vehicle_plate
    plate_to_vid = {plate: vid for vid, plate in vehicles.items()}
    wo_salary_rows = (await db.execute(
        select(
            Vehicle.id,
            func.coalesce(func.sum(DeliveredTrip.driver_salary), 0),
        )
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
        .group_by(Vehicle.id)
    )).all()

    salary_by_vehicle: dict[int, int] = {}
    for vid, sal in wo_salary_rows:
        if vid:
            salary_by_vehicle[vid] = int(sal or 0)

    # ── 3b. Trip count per vehicle for CP Chung allocation ─────────
    wo_count_rows = (await db.execute(
        select(
            Vehicle.id,
            func.count(DeliveredTrip.id),
        )
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
        .group_by(Vehicle.id)
    )).all()

    trip_count_by_vehicle: dict[int, int] = {}
    total_trips = 0
    for vid, cnt in wo_count_rows:
        if vid:
            trip_count_by_vehicle[vid] = int(cnt)
            total_trips += int(cnt)

    # ── 4. CP Xe (vehicle expenses) per vehicle ──────────────────────────────
    expense_rows = (await db.execute(
        select(
            VehicleExpense.vehicle_id,
            VehicleExpense.category,
            func.coalesce(func.sum(VehicleExpense.amount), 0),
        )
        .where(
            VehicleExpense.expense_date >= df,
            VehicleExpense.expense_date <= dt,
            VehicleExpense.category.in_(["XANG_DAU", "SUA_CHUA", "TIEN_LUAT", "KHAC"]),
        )
        .group_by(VehicleExpense.vehicle_id, VehicleExpense.category)
    )).all()

    cp_xe_by_vehicle: dict[int, dict[str, int]] = {}
    for vid, cat, total_amt in expense_rows:
        amt = int(total_amt or 0)
        if vid and vid in vehicles:
            slot = cp_xe_by_vehicle.setdefault(vid, {"XANG_DAU": 0, "SUA_CHUA": 0, "TIEN_LUAT": 0, "KHAC": 0})
            slot[cat] = slot.get(cat, 0) + amt

    # ── 5. CP Lương cơ bản: drivers attached to each vehicle via vehicle_drivers
    vd_rows = (await db.execute(
        select(VehicleDriver.vehicle_id, VehicleDriver.driver_id)
        .where(
            VehicleDriver.vehicle_id.in_(list(vehicles.keys())),
            VehicleDriver.is_active == True,  # noqa: E712
            VehicleDriver.effective_from <= dt,
        )
    )).all()

    vehicle_driver_map: dict[int, list[int]] = {}
    for vid, did in vd_rows:
        vehicle_driver_map.setdefault(vid, []).append(did)

    all_driver_ids = list({d for drivers in vehicle_driver_map.values() for d in drivers})
    base_salary_repo = SqlDriverSalaryConfigRepository(db)
    history_by_driver = await base_salary_repo.list_history_for_drivers(all_driver_ids) if all_driver_ids else {}

    base_salary_by_vehicle: dict[int, int] = {}
    for vid, driver_ids in vehicle_driver_map.items():
        total_base = sum(
            effective_base_salary(history_by_driver.get(did, []), dt)
            for did in driver_ids
        )
        base_salary_by_vehicle[vid] = total_base

    # ── 6. Assemble rows ─────────────────────────────────────────────────────
    rows: list[VehiclePnLRow] = []
    total_revenue = 0
    sum_row_profits = 0

    for vid, plate in sorted(vehicles.items(), key=lambda x: x[1]):
        rev = revenue_by_vehicle.get(vid, 0)
        sal = salary_by_vehicle.get(vid, 0)
        base = base_salary_by_vehicle.get(vid, 0)
        vcost = vendor_cost_by_vehicle.get(vid, 0)
        xe_cats = cp_xe_by_vehicle.get(vid, {})
        xe_summary = VehicleExpenseSummary(
            xang_dau=xe_cats.get("XANG_DAU", 0),
            sua_chua=xe_cats.get("SUA_CHUA", 0),
            tien_luat=xe_cats.get("TIEN_LUAT", 0),
            khac=xe_cats.get("KHAC", 0),
            total=sum(xe_cats.values()),
        )
        loi_nhuan = rev - (xe_summary.total + sal + base + vcost)
        vnd_id = vendor_id_by_vehicle.get(vid)
        rows.append(VehiclePnLRow(
            vehicle_id=vid,
            plate=plate,
            is_vendor=vnd_id is not None,
            vendor_name=vendor_name_by_id.get(vnd_id) if vnd_id is not None else None,
            revenue=rev,
            cp_xe=xe_summary,
            cp_luong_san_luong=sal,
            cp_luong_co_ban=base,
            cp_vendor=vcost,
            loi_nhuan=loi_nhuan,
        ))
        total_revenue += rev
        sum_row_profits += loi_nhuan

    total_profit = sum_row_profits

    return VehiclePnLResponse(
        date_from=df,
        date_to=dt,
        rows=rows,
        total_revenue=total_revenue,
        total_profit=total_profit,
    )


@router.get("/vehicle-pnl/export")
async def export_vehicle_pnl(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    vehicle_id: Optional[int] = Query(None, description="Filter to a single vehicle"),
    _current_user: User = Depends(require_permission("read", "Salary")),
    db: AsyncSession = Depends(get_db),
):
    """Export per-vehicle P&L as an Excel (.xlsx) file."""
    import io
    from fastapi import HTTPException
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    from app.contexts.payroll.infrastructure.repositories import SqlDriverSalaryConfigRepository
    from app.contexts.payroll.domain.base_salary import effective_base_salary
    from app.core.pricing_lookup import TripPriceInfo, lookup_vendor_prices

    try:
        df = _dt.strptime(date_from, "%Y-%m-%d").date()
        dt = _dt.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=422, detail="Invalid date format. Use YYYY-MM-DD.")

    # ── Re-use the same data logic as vehicle-pnl ────────────────────────────
    veh_q = select(Vehicle.id, Vehicle.plate).where(Vehicle.is_active == True)  # noqa: E712
    if vehicle_id is not None:
        veh_q = veh_q.where(Vehicle.id == vehicle_id)
    veh_rows = (await db.execute(veh_q)).all()
    vehicles: dict[int, str] = {r[0]: r[1] for r in veh_rows}

    trip_detail_rows = (await db.execute(
        select(
            DeliveredTrip.id,
            DeliveredTrip.vendor_id,
            DeliveredTrip.pickup_location_id,
            DeliveredTrip.dropoff_location_id,
            DeliveredTrip.work_type,
            DeliveredTrip.cont_type,
            DeliveredTrip.revenue,
            Vehicle.id,
        )
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
    )).all()

    vendor_trips = [
        TripPriceInfo(id=r[0], partner_id=r[1], pickup_location_id=r[2], dropoff_location_id=r[3], work_type=r[4], cont_type=r[5])
        for r in trip_detail_rows if r[1] is not None
    ]
    vendor_prices = await lookup_vendor_prices(db, vendor_trips)

    revenue_by_vehicle: dict[int, int] = {}
    vendor_cost_by_vehicle: dict[int, int] = {}
    for r in trip_detail_rows:
        trip_id, vid, trip_rev = r[0], r[7], int(r[6] or 0)
        if vid:
            revenue_by_vehicle[vid] = revenue_by_vehicle.get(vid, 0) + trip_rev
            vendor_cost_by_vehicle[vid] = vendor_cost_by_vehicle.get(vid, 0) + vendor_prices.get(trip_id, 0)

    wo_salary_rows = (await db.execute(
        select(
            Vehicle.id,
            func.coalesce(func.sum(DeliveredTrip.driver_salary), 0),
        )
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
        .group_by(Vehicle.id)
    )).all()
    salary_by_vehicle: dict[int, int] = {vid: int(sal or 0) for vid, sal in wo_salary_rows if vid}

    expense_rows = (await db.execute(
        select(
            VehicleExpense.vehicle_id,
            VehicleExpense.category,
            func.coalesce(func.sum(VehicleExpense.amount), 0),
        )
        .where(
            VehicleExpense.expense_date >= df,
            VehicleExpense.expense_date <= dt,
            VehicleExpense.category.in_(["XANG_DAU", "SUA_CHUA", "TIEN_LUAT", "KHAC"]),
        )
        .group_by(VehicleExpense.vehicle_id, VehicleExpense.category)
    )).all()

    cp_xe_by_vehicle: dict[int, dict[str, int]] = {}
    for vid, cat, total_amt in expense_rows:
        if vid and vid in vehicles:
            slot = cp_xe_by_vehicle.setdefault(vid, {"XANG_DAU": 0, "SUA_CHUA": 0, "TIEN_LUAT": 0, "KHAC": 0})
            slot[cat] = slot.get(cat, 0) + int(total_amt or 0)

    vd_rows = (await db.execute(
        select(VehicleDriver.vehicle_id, VehicleDriver.driver_id)
        .where(
            VehicleDriver.vehicle_id.in_(list(vehicles.keys())),
            VehicleDriver.is_active == True,  # noqa: E712
            VehicleDriver.effective_from <= dt,
        )
    )).all()
    vehicle_driver_map: dict[int, list[int]] = {}
    for vid, did in vd_rows:
        vehicle_driver_map.setdefault(vid, []).append(did)

    all_driver_ids = list({d for drivers in vehicle_driver_map.values() for d in drivers})
    base_salary_repo = SqlDriverSalaryConfigRepository(db)
    history_by_driver = await base_salary_repo.list_history_for_drivers(all_driver_ids) if all_driver_ids else {}

    base_salary_by_vehicle: dict[int, int] = {
        vid: sum(effective_base_salary(history_by_driver.get(did, []), dt) for did in driver_ids)
        for vid, driver_ids in vehicle_driver_map.items()
    }

    # ── Build Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo P&L"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill("solid", fgColor="E8F0FE")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    profit_pos_fill = PatternFill("solid", fgColor="D1FAE5")
    profit_neg_fill = PatternFill("solid", fgColor="FEE2E2")
    profit_pos_font = Font(bold=True, color="065F46", size=11)
    profit_neg_font = Font(bold=True, color="991B1B", size=11)

    # Title row
    period_label = f"{df.strftime('%d/%m/%Y')} – {dt.strftime('%d/%m/%Y')}"
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"BÁO CÁO LỢI NHUẬN THEO XE  |  Kỳ: {period_label}"
    title_cell.font = Font(bold=True, size=13, color="1E3A5F")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    ws.append([])  # blank row 2

    # Header row (row 3)
    headers = [
        "Biển số",
        "Doanh thu",
        "CP Xăng dầu",
        "CP Sửa chữa",
        "CP Tiền luật",
        "CP Khác",
        "Lương LX",
        "Tổng CP",
        "Lợi nhuận",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 22

    # Number format: thousand-separator, no decimals, ₫ suffix
    VND_FMT = '#,##0\\ [$₫-vi-VN]'

    def set_num(cell, value: int) -> None:
        """Write a real integer and apply VND display format."""
        cell.value = value
        cell.number_format = VND_FMT
        cell.alignment = right

    # Data rows
    data_start_row = 4
    total_rev = total_luong = total_xang = total_sua = total_luat = total_khac = total_xe = total_profit = 0

    sorted_vehicles = sorted(vehicles.items(), key=lambda x: x[1])
    for row_idx, (vid, plate) in enumerate(sorted_vehicles, start=data_start_row):
        rev = revenue_by_vehicle.get(vid, 0)
        sal = salary_by_vehicle.get(vid, 0)
        base = base_salary_by_vehicle.get(vid, 0)
        vcost = vendor_cost_by_vehicle.get(vid, 0)
        xe_cats = cp_xe_by_vehicle.get(vid, {})
        xang = xe_cats.get("XANG_DAU", 0)
        sua = xe_cats.get("SUA_CHUA", 0)
        luat = xe_cats.get("TIEN_LUAT", 0)
        khac = xe_cats.get("KHAC", 0)
        xe_total = xang + sua + luat + khac
        luong = sal + base
        total_cp = xe_total + luong + vcost
        profit = rev - total_cp

        total_rev += rev
        total_luong += luong
        total_xang += xang
        total_sua += sua
        total_luat += luat
        total_khac += khac
        total_xe += xe_total
        total_profit += profit

        # Append plate first so row exists, then fill numeric cells
        ws.append([plate])
        num_values = [rev, xang, sua, luat, khac, luong, total_cp, profit]
        for col_offset, val in enumerate(num_values, start=2):
            set_num(ws.cell(row=row_idx, column=col_offset), val)

        for col_idx in range(1, len(num_values) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 1:
                cell.alignment = center
                cell.font = Font(bold=True)
            elif col_idx == len(num_values) + 1:
                # Profit column — colour by sign
                if profit >= 0:
                    cell.fill = profit_pos_fill
                    cell.font = profit_pos_font
                else:
                    cell.fill = profit_neg_fill
                    cell.font = profit_neg_font

        ws.row_dimensions[row_idx].height = 18

    # Total row
    total_row_idx = data_start_row + len(sorted_vehicles)
    total_cp_all = total_xe + total_luong
    ws.append(["TỔNG"])
    total_num_values = [total_rev, total_xang, total_sua, total_luat, total_khac, total_luong, total_cp_all, total_profit]
    for col_offset, val in enumerate(total_num_values, start=2):
        set_num(ws.cell(row=total_row_idx, column=col_offset), val)

    for col_idx in range(1, len(total_num_values) + 2):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = right if col_idx > 1 else center
    ws.row_dimensions[total_row_idx].height = 20

    # Column widths
    col_widths = [14, 16, 15, 15, 15, 13, 15, 15, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Footer
    footer_row = total_row_idx + 2
    ws.cell(row=footer_row, column=1).value = f"Xuất ngày: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="888888", size=9)

    # Stream response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"PnL_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/trip-daily-stats", response_model=TripDailyStatsOut)
async def get_trip_daily_stats(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    client_id: int | None = None,
    _current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Lightweight daily trip aggregation for the dashboard bar chart.

    Returns matched/pending counts per day without fetching
    full booked-trip objects.  ~10x faster than fetching all trips.
    """
    try:
        df = datetime.strptime(date_from, "%Y-%m-%d").date()
        dt = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        from fastapi import HTTPException
        raise HTTPException(status_code=422, detail="Invalid date format. Use YYYY-MM-DD.")

    stmt = select(
        DeliveredTrip.trip_date,
        DeliveredTrip.booked_trip_id,
        func.count(DeliveredTrip.id),
        func.coalesce(func.sum(DeliveredTrip.revenue), 0),
    ).where(
        DeliveredTrip.trip_date >= df,
        DeliveredTrip.trip_date <= dt,
    )

    if client_id is not None:
        stmt = stmt.where(DeliveredTrip.client_id == client_id)

    rows = (await db.execute(
        stmt.group_by(DeliveredTrip.trip_date, DeliveredTrip.booked_trip_id)
    )).all()

    date_map: dict[str, dict[str, int]] = {}
    total = 0
    matched = 0
    pending = 0
    internal_count = 0
    vendor_count = 0
    total_revenue = 0
    for trip_date, booked_id, cnt, rev in rows:
        ds = str(trip_date) if not hasattr(trip_date, 'isoformat') else trip_date.isoformat()
        bucket = date_map.setdefault(ds, {"matched": 0, "pending": 0})
        total += cnt
        if booked_id is not None:
            bucket["matched"] += cnt
            matched += cnt
            total_revenue += int(rev)
        else:
            bucket["pending"] += cnt
            pending += cnt
            total_revenue += int(rev)

    # Build one bucket per day across the full date range
    from datetime import timedelta as _td
    buckets = []
    cur = df
    idx = 0
    while cur <= dt:
        ds = cur.isoformat()
        b = date_map.get(ds, {"matched": 0, "pending": 0})
        buckets.append({
            "day": idx + 1,
            "date": ds,
            "matched": b["matched"],
            "pending": b["pending"],
        })
        cur += _td(days=1)
        idx += 1

    # Distinct vehicle counts (not trip counts)
    base_where = [
        DeliveredTrip.trip_date >= df,
        DeliveredTrip.trip_date <= dt,
    ]
    if client_id is not None:
        base_where.append(DeliveredTrip.client_id == client_id)

    internal_count = (await db.execute(
        select(func.count(func.distinct(DeliveredTrip.driver_id)))
        .where(DeliveredTrip.vendor_id.is_(None), DeliveredTrip.driver_id.isnot(None), *base_where)
    )).scalar() or 0

    vendor_count = (await db.execute(
        select(func.count(func.distinct(DeliveredTrip.vendor_id)))
        .where(DeliveredTrip.vendor_id.isnot(None), *base_where)
    )).scalar() or 0

    return TripDailyStatsOut(
        date_from=df,
        date_to=dt,
        total=total,
        matched=matched,
        pending=pending,
        internal_count=internal_count,
        vendor_count=vendor_count,
        total_revenue=total_revenue,
        match_rate=round(matched / total * 100) if total > 0 else None,
        buckets=buckets,
    )


async def _compute_vehicle_pnl_rows(db: AsyncSession, df, dt) -> list[VehiclePnLRow]:
    """Compute per-vehicle PnL rows for a date range.

    Reused by both the /vehicle-pnl endpoint and the /director dashboard so the
    two stay in sync. Mirrors the logic in get_vehicle_pnl (rev, cp_xe, salary,
    base salary, vendor cost) but without HTTP plumbing.
    """
    from app.contexts.payroll.infrastructure.repositories import SqlDriverSalaryConfigRepository
    from app.contexts.payroll.domain.base_salary import effective_base_salary
    from app.core.pricing_lookup import TripPriceInfo, lookup_vendor_prices

    veh_rows = (await db.execute(
        select(Vehicle.id, Vehicle.plate, Vehicle.vendor_id).where(Vehicle.is_active == True)  # noqa: E712
    )).all()
    vehicles: dict[int, str] = {r[0]: r[1] for r in veh_rows}
    vendor_id_by_vehicle: dict[int, int | None] = {r[0]: r[2] for r in veh_rows}

    all_vendor_ids = list({r[2] for r in veh_rows if r[2] is not None})
    vendor_name_by_id: dict[int, str] = {}
    if all_vendor_ids:
        vnd_rows = (await db.execute(
            select(Vendor.id, Vendor.name).where(Vendor.id.in_(all_vendor_ids))
        )).all()
        vendor_name_by_id = {r[0]: r[1] for r in vnd_rows}

    if not vehicles:
        return []

    trip_detail_rows = (await db.execute(
        select(
            DeliveredTrip.id,
            DeliveredTrip.vendor_id,
            DeliveredTrip.pickup_location_id,
            DeliveredTrip.dropoff_location_id,
            DeliveredTrip.work_type,
            DeliveredTrip.cont_type,
            DeliveredTrip.revenue,
            Vehicle.id,
        )
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
    )).all()

    vendor_trips = [
        TripPriceInfo(id=r[0], partner_id=r[1], pickup_location_id=r[2], dropoff_location_id=r[3], work_type=r[4], cont_type=r[5])
        for r in trip_detail_rows if r[1] is not None
    ]
    vendor_prices = await lookup_vendor_prices(db, vendor_trips)

    revenue_by_vehicle: dict[int, int] = {}
    vendor_cost_by_vehicle: dict[int, int] = {}
    for r in trip_detail_rows:
        trip_id, vid, trip_rev = r[0], r[7], int(r[6] or 0)
        if vid:
            revenue_by_vehicle[vid] = revenue_by_vehicle.get(vid, 0) + trip_rev
            vendor_cost_by_vehicle[vid] = vendor_cost_by_vehicle.get(vid, 0) + vendor_prices.get(trip_id, 0)

    wo_salary_rows = (await db.execute(
        select(
            Vehicle.id,
            func.coalesce(func.sum(DeliveredTrip.driver_salary), 0),
        )
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
        .group_by(Vehicle.id)
    )).all()
    salary_by_vehicle: dict[int, int] = {vid: int(sal or 0) for vid, sal in wo_salary_rows if vid}

    # Trip count per vehicle for ordering / context
    wo_count_rows = (await db.execute(
        select(Vehicle.id, func.count(DeliveredTrip.id))
        .join(Vehicle, Vehicle.plate == DeliveredTrip.vehicle_plate)
        .where(
            DeliveredTrip.booked_trip_id.isnot(None),
            Vehicle.id.in_(list(vehicles.keys())),
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) >= df,
            func.coalesce(DeliveredTrip.trip_date, func.date(DeliveredTrip.created_at)) <= dt,
        )
        .group_by(Vehicle.id)
    )).all()
    trip_count_by_vehicle: dict[int, int] = {vid: int(cnt) for vid, cnt in wo_count_rows if vid}

    expense_rows = (await db.execute(
        select(
            VehicleExpense.vehicle_id,
            VehicleExpense.category,
            func.coalesce(func.sum(VehicleExpense.amount), 0),
        )
        .where(
            VehicleExpense.expense_date >= df,
            VehicleExpense.expense_date <= dt,
            VehicleExpense.category.in_(["XANG_DAU", "SUA_CHUA", "TIEN_LUAT", "KHAC"]),
        )
        .group_by(VehicleExpense.vehicle_id, VehicleExpense.category)
    )).all()
    cp_xe_by_vehicle: dict[int, dict[str, int]] = {}
    for vid, cat, total_amt in expense_rows:
        if vid and vid in vehicles:
            slot = cp_xe_by_vehicle.setdefault(vid, {"XANG_DAU": 0, "SUA_CHUA": 0, "TIEN_LUAT": 0, "KHAC": 0})
            slot[cat] = slot.get(cat, 0) + int(total_amt or 0)

    vd_rows = (await db.execute(
        select(VehicleDriver.vehicle_id, VehicleDriver.driver_id)
        .where(
            VehicleDriver.vehicle_id.in_(list(vehicles.keys())),
            VehicleDriver.is_active == True,  # noqa: E712
            VehicleDriver.effective_from <= dt,
        )
    )).all()
    vehicle_driver_map: dict[int, list[int]] = {}
    for vid, did in vd_rows:
        vehicle_driver_map.setdefault(vid, []).append(did)

    all_driver_ids = list({d for drivers in vehicle_driver_map.values() for d in drivers})
    base_salary_repo = SqlDriverSalaryConfigRepository(db)
    history_by_driver = await base_salary_repo.list_history_for_drivers(all_driver_ids) if all_driver_ids else {}

    base_salary_by_vehicle: dict[int, int] = {
        vid: sum(effective_base_salary(history_by_driver.get(did, []), dt) for did in driver_ids)
        for vid, driver_ids in vehicle_driver_map.items()
    }

    rows: list[VehiclePnLRow] = []
    for vid, plate in sorted(vehicles.items(), key=lambda x: x[1]):
        rev = revenue_by_vehicle.get(vid, 0)
        sal = salary_by_vehicle.get(vid, 0)
        base = base_salary_by_vehicle.get(vid, 0)
        vcost = vendor_cost_by_vehicle.get(vid, 0)
        xe_cats = cp_xe_by_vehicle.get(vid, {})
        xe_summary = VehicleExpenseSummary(
            xang_dau=xe_cats.get("XANG_DAU", 0),
            sua_chua=xe_cats.get("SUA_CHUA", 0),
            tien_luat=xe_cats.get("TIEN_LUAT", 0),
            khac=xe_cats.get("KHAC", 0),
            total=sum(xe_cats.values()),
        )
        # Skip rows that had no activity AND no expenses for the period.
        if (
            rev == 0 and sal == 0 and base == 0 and vcost == 0
            and xe_summary.total == 0 and trip_count_by_vehicle.get(vid, 0) == 0
        ):
            continue
        loi_nhuan = rev - (xe_summary.total + sal + base + vcost)
        vnd_id = vendor_id_by_vehicle.get(vid)
        rows.append(VehiclePnLRow(
            vehicle_id=vid,
            plate=plate,
            is_vendor=vnd_id is not None,
            vendor_name=vendor_name_by_id.get(vnd_id) if vnd_id is not None else None,
            revenue=rev,
            cp_xe=xe_summary,
            cp_luong_san_luong=sal,
            cp_luong_co_ban=base,
            cp_vendor=vcost,
            loi_nhuan=loi_nhuan,
        ))
    return rows


def _row_cost(row: VehiclePnLRow) -> int:
    return int(row.cp_xe.total + row.cp_luong_san_luong + row.cp_luong_co_ban + row.cp_vendor)


@router.get("/director", response_model=DirectorDashboardOut)
async def get_director_dashboard(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    _current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Director dashboard: KPIs, trends, top routes/drivers, daily buckets.

    Computes current period stats and compares with the previous calendar month
    to produce delta percentages.
    """
    try:
        df = datetime.strptime(date_from, "%Y-%m-%d").date()
        dt = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        from fastapi import HTTPException
        raise HTTPException(status_code=422, detail="Invalid date format. Use YYYY-MM-DD.")

    # Previous period: same-length window ending the day before df
    span = (dt - df).days + 1
    prev_dt = df - timedelta(days=1)
    prev_df = prev_dt - timedelta(days=span - 1)

    async def _period_stats(start, end):
        """Return (total, matched, pending, revenue, buckets, top_routes, top_drivers)."""
        rows = (await db.execute(
            select(
                DeliveredTrip.trip_date,
                DeliveredTrip.booked_trip_id,
                func.count(DeliveredTrip.id),
                func.coalesce(func.sum(DeliveredTrip.revenue), 0),
            )
            .where(DeliveredTrip.trip_date >= start, DeliveredTrip.trip_date <= end)
            .group_by(DeliveredTrip.trip_date, DeliveredTrip.booked_trip_id)
        )).all()

        date_map = {}
        total = matched = pending = revenue = 0
        for trip_date, booked_id, cnt, rev in rows:
            ds = trip_date.isoformat()
            bucket = date_map.setdefault(ds, {"matched": 0, "pending": 0})
            total += cnt
            if booked_id is not None:
                bucket["matched"] += cnt
                matched += cnt
                revenue += int(rev)
            else:
                bucket["pending"] += cnt
                pending += cnt

        buckets = []
        cur, idx = start, 0
        while cur <= end:
            ds = cur.isoformat()
            b = date_map.get(ds, {"matched": 0, "pending": 0})
            buckets.append({"day": idx + 1, "date": ds, "matched": b["matched"], "pending": b["pending"]})
            cur += timedelta(days=1)
            idx += 1

        # Top routes
        from sqlalchemy.orm import aliased
        PickupLoc = aliased(Location)
        DropLoc = aliased(Location)
        route_rows = (await db.execute(
            select(
                func.concat(PickupLoc.name, " → ", DropLoc.name).label("route"),
                func.count(DeliveredTrip.id).label("cnt"),
            )
            .join(PickupLoc, PickupLoc.id == DeliveredTrip.pickup_location_id)
            .join(DropLoc, DropLoc.id == DeliveredTrip.dropoff_location_id)
            .where(DeliveredTrip.trip_date >= start, DeliveredTrip.trip_date <= end)
            .group_by(PickupLoc.name, DropLoc.name)
            .order_by(func.count(DeliveredTrip.id).desc())
            .limit(5)
        )).all()
        top_routes = [{"name": r.route, "count": r.cnt} for r in route_rows]

        # Top drivers
        driver_rows = (await db.execute(
            select(
                User.full_name.label("name"),
                DeliveredTrip.vehicle_plate.label("plate"),
                func.count(DeliveredTrip.id).label("cnt"),
            )
            .join(User, User.id == DeliveredTrip.driver_id)
            .where(DeliveredTrip.trip_date >= start, DeliveredTrip.trip_date <= end)
            .group_by(User.full_name, DeliveredTrip.vehicle_plate)
            .order_by(func.count(DeliveredTrip.id).desc())
            .limit(5)
        )).all()
        top_drivers = [{"name": r.name, "plate": r.plate or "", "trip_count": r.cnt} for r in driver_rows]

        return total, matched, pending, revenue, buckets, top_routes, top_drivers

    total, matched, pending, revenue, buckets, top_routes, top_drivers = await _period_stats(df, dt)
    prev_total, prev_matched, prev_pending, prev_revenue, _, _, _ = await _period_stats(prev_df, prev_dt)

    # ── Per-vehicle PnL, split own-fleet vs vendor ──────────────────────────
    pnl_rows = await _compute_vehicle_pnl_rows(db, df, dt)
    prev_pnl_rows = await _compute_vehicle_pnl_rows(db, prev_df, prev_dt)

    own_rows = [r for r in pnl_rows if not r.is_vendor]
    vendor_rows = [r for r in pnl_rows if r.is_vendor]

    def _group(rows: list[VehiclePnLRow]) -> VehiclePnLGroup:
        return VehiclePnLGroup(
            rows=rows,
            total_revenue=sum(r.revenue for r in rows),
            total_cost=sum(_row_cost(r) for r in rows),
            total_profit=sum(r.loi_nhuan for r in rows),
            trip_count=0,
        )

    own_group = _group(own_rows)
    vendor_group = _group(vendor_rows)

    # Aggregate totals — revenue, cost & profit all derived from PnL rows
    # (matched DeliveredTrips) so the KPI cards stay consistent with the
    # per-vehicle breakdown tables below them.
    pnl_revenue = sum(r.revenue for r in pnl_rows)
    total_cost = sum(_row_cost(r) for r in pnl_rows)
    profit = sum(r.loi_nhuan for r in pnl_rows)
    prev_pnl_revenue = sum(r.revenue for r in prev_pnl_rows)
    prev_total_cost = sum(_row_cost(r) for r in prev_pnl_rows)
    prev_profit = sum(r.loi_nhuan for r in prev_pnl_rows)

    def delta(curr, prev):
        if not prev:
            return None
        return round(((curr - prev) / abs(prev)) * 100, 1)

    return DirectorDashboardOut(
        total=total,
        matched=matched,
        pending=pending,
        match_rate=round(matched / total * 100) if total > 0 else None,
        revenue=pnl_revenue,
        avg_revenue_per_trip=round(pnl_revenue / matched) if matched > 0 else 0,
        total_cost=total_cost,
        profit=profit,
        total_delta=delta(total, prev_total),
        matched_delta=delta(matched, prev_matched),
        pending_delta=delta(pending, prev_pending),
        revenue_delta=delta(pnl_revenue, prev_pnl_revenue),
        cost_delta=delta(total_cost, prev_total_cost),
        profit_delta=delta(profit, prev_profit),
        buckets=buckets,
        top_routes=top_routes,
        top_drivers=top_drivers,
        own_fleet_pnl=own_group,
        vendor_pnl=vendor_group,
    )


@router.get("/notifications")
async def get_notifications(
    current_user: User = Depends(get_current_user),
):
    try:
        redis = get_arq_pool()
        key = f"notifications:user:{current_user.id}"
        raw_items = await redis.zrevrange(key, 0, 49)
        notifications = []
        for i, raw in enumerate(raw_items):
            try:
                data = json.loads(raw)
                notifications.append({
                    "id": str(i),
                    "type": data.get("channel", "general"),
                    "title": data.get("title", ""),
                    "message": data.get("message", ""),
                    "time": data.get("created_at", ""),
                    "read": data.get("read", False),
                })
            except (json.JSONDecodeError, KeyError):
                logger.warning("Malformed notification entry for user %s", current_user.id)
                continue
        return notifications
    except RuntimeError:
        logger.warning("arq pool unavailable, returning empty notifications")
        return []
