"""Render the customer-settlement workbook (BKTT + SL sheets).

Mirrors the layout of `PAN_BK_SL_T04.26_HD.xlsx` so the customer's accountant
can drop our file in beside theirs and reconcile column-for-column.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.contexts.billing.domain.entities import SettlementStatement
from app.utils.number_to_words_vi import number_to_vietnamese_words


COMPANY_NAME = "CÔNG TY TNHH AMT PHÚC LỘC"
COMPANY_ADDRESS = "Số 56B/97 đường Đoàn Kết, Phường Hải An, TP Hải Phòng"
COMPANY_TAX_CODE = "0201965047"


_THIN = Side(style="thin", color="000000")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_BOLD = Font(bold=True)
_BOLD_LARGE = Font(bold=True, size=14)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _money_fmt(cell):
    cell.number_format = "#,##0;-#,##0;\"\""
    cell.alignment = _RIGHT


def _date_fmt(cell):
    cell.number_format = "dd/mm/yyyy"
    cell.alignment = _CENTER


def _apply_border(ws, row_start: int, row_end: int, col_start: int, col_end: int) -> None:
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = _BORDER_ALL


def _write_sl_sheet(wb: openpyxl.Workbook, data: SettlementStatement, sheet_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    p_start = data.period.start
    p_end = data.period.end
    period_month = p_end.month
    period_year = p_end.year

    ws.cell(row=1, column=1, value=COMPANY_NAME).font = _BOLD
    ws.cell(row=2, column=1, value=COMPANY_ADDRESS)
    ws.cell(row=3, column=1, value=f"MST: {COMPANY_TAX_CODE}")
    title_cell = ws.cell(
        row=5, column=1,
        value=f"BẢNG KÊ QUYẾT TOÁN CƯỚC VẬN CHUYỂN THÁNG {period_month:02d}/{period_year}",
    )
    title_cell.font = _BOLD_LARGE
    title_cell.alignment = _CENTER
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=16)
    period_cell = ws.cell(
        row=6, column=1,
        value=f"Từ ngày {p_start.strftime('%d/%m/%Y')} đến ngày {p_end.strftime('%d/%m/%Y')}",
    )
    period_cell.alignment = _CENTER
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=16)
    ws.cell(row=7, column=1, value="KHÁCH HÀNG:")
    ws.cell(row=7, column=3, value=data.client.name).font = _BOLD
    ws.cell(
        row=8, column=1,
        value="Công ty TNHH AMT Phúc Lộc xin gửi tới Quý Công ty bảng kê khối lượng vận chuyển trong kỳ.",
    )

    headers = [
        ("STT", 6),
        ("NGÀY ĐI", 12),
        ("CHỦ HÀNG", 10),
        ("SỐ CONTAINER", 16),
        ("F20'", 6),
        ("F40'", 6),
        ("E20'", 6),
        ("E40'", 6),
        ("SỐ XE CHẠY", 14),
        ("SỐ TÀU", 14),
        ("ĐIỂM ĐI", 16),
        ("ĐIỂM ĐẾN", 16),
        ("CƯỚC CHUYẾN", 14),
        ("TỔNG TT", 14),
        ("TÁC NGHIỆP", 12),
        ("GHI CHÚ", 12),
    ]
    header_row = 10
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    first_data_row = header_row + 1
    row = first_data_row
    for stt, line in enumerate(data.trip_lines, start=1):
        ws.cell(row=row, column=1, value=stt).alignment = _CENTER
        d = ws.cell(row=row, column=2, value=line.trip_date)
        _date_fmt(d)
        ws.cell(row=row, column=3, value=line.client_code).alignment = _CENTER
        ws.cell(row=row, column=4, value=line.container_number).alignment = _CENTER

        ct = line.cont_type
        ws.cell(row=row, column=5, value=1 if ct == "F20" else None).alignment = _CENTER
        ws.cell(row=row, column=6, value=1 if ct == "F40" else None).alignment = _CENTER
        ws.cell(row=row, column=7, value=1 if ct == "E20" else None).alignment = _CENTER
        ws.cell(row=row, column=8, value=1 if ct == "E40" else None).alignment = _CENTER

        ws.cell(row=row, column=9, value=line.tractor_plate).alignment = _CENTER
        ws.cell(row=row, column=10, value=line.vessel).alignment = _CENTER
        ws.cell(row=row, column=11, value=line.pickup_location).alignment = _LEFT
        ws.cell(row=row, column=12, value=line.dropoff_location).alignment = _LEFT
        cuoc = ws.cell(row=row, column=13, value=line.unit_price)
        _money_fmt(cuoc)
        tt = ws.cell(row=row, column=14, value=f"=M{row}")
        _money_fmt(tt)
        ws.cell(row=row, column=15, value=line.work_type).alignment = _CENTER
        ws.cell(row=row, column=16, value="").alignment = _CENTER
        row += 1

    last_data_row = row - 1 if data.trip_lines else first_data_row - 1

    if data.trip_lines:
        _apply_border(ws, header_row, last_data_row, 1, 16)
        total_row = last_data_row + 1
        ws.cell(row=total_row, column=4, value="TỔNG CỘNG").font = _BOLD
        ws.cell(row=total_row, column=4).alignment = _RIGHT
        sum_cell = ws.cell(row=total_row, column=14, value=f"=SUM(N{first_data_row}:N{last_data_row})")
        _money_fmt(sum_cell)
        sum_cell.font = _BOLD
        sum_cell.fill = _TOTAL_FILL

        vat_row = total_row + 1
        ws.cell(row=vat_row, column=4, value="VAT 8%").font = _BOLD
        ws.cell(row=vat_row, column=4).alignment = _RIGHT
        vat_cell = ws.cell(row=vat_row, column=14, value=f"=ROUND(N{total_row}*0.08,0)")
        _money_fmt(vat_cell)

        grand_row = vat_row + 1
        ws.cell(row=grand_row, column=4, value="TỔNG THANH TOÁN").font = _BOLD
        ws.cell(row=grand_row, column=4).alignment = _RIGHT
        grand_cell = ws.cell(row=grand_row, column=14, value=f"=N{total_row}+N{vat_row}")
        _money_fmt(grand_cell)
        grand_cell.font = _BOLD
        grand_cell.fill = _TOTAL_FILL

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _write_bktt_sheet(wb: openpyxl.Workbook, data: SettlementStatement, sheet_name: str) -> None:
    ws = wb.create_sheet(sheet_name)
    p_end = data.period.end
    period_month = p_end.month
    period_year = p_end.year

    ws.cell(row=1, column=1, value=COMPANY_NAME).font = _BOLD
    ws.cell(row=2, column=1, value=f"Địa chỉ: {COMPANY_ADDRESS}")
    title = ws.cell(
        row=4, column=1,
        value=f"BẢNG QUYẾT TOÁN PHÍ VẬN CHUYỂN THÁNG {period_month:02d}/{period_year}",
    )
    title.font = _BOLD_LARGE
    title.alignment = _CENTER
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=12)

    ws.cell(row=5, column=2, value=f"KHÁCH HÀNG:  {data.client.name}").font = _BOLD
    ws.cell(row=6, column=2, value="ĐỊA CHỈ:")
    ws.cell(row=6, column=3, value=data.client.address or "")
    ws.cell(row=7, column=2, value="MST :")
    ws.cell(row=7, column=3, value=data.client.tax_code or "")

    header_top = 9
    header_bottom = 10

    headers_top = [
        (1, "STT"),
        (2, "Tuyến vận tải"),
        (4, "Đơn vị tính"),
        (5, "Loại cont"),
        (9, " Cộng Số lượng"),
        (10, "Đơn giá"),
        (11, "Thành tiền"),
        (12, "Ghi chú"),
    ]
    for col, label in headers_top:
        c = ws.cell(row=header_top, column=col, value=label)
        c.font = _BOLD
        c.alignment = _CENTER
        c.fill = _HEADER_FILL

    ws.merge_cells(start_row=header_top, start_column=2, end_row=header_top, end_column=3)
    ws.merge_cells(start_row=header_top, start_column=5, end_row=header_top, end_column=8)
    for col in (1, 4, 9, 10, 11, 12):
        ws.merge_cells(start_row=header_top, start_column=col, end_row=header_bottom, end_column=col)

    subheaders = [
        (2, "Điểm đi"),
        (3, "Điểm đến"),
        (5, "Hàng 20'"),
        (6, "Hàng 40'"),
        (7, "Vỏ"),
        (8, ""),
    ]
    for col, label in subheaders:
        c = ws.cell(row=header_bottom, column=col, value=label)
        c.font = _BOLD
        c.alignment = _CENTER
        c.fill = _HEADER_FILL

    column_widths = {1: 6, 2: 18, 3: 18, 4: 12, 5: 10, 6: 10, 7: 10, 8: 8, 9: 12, 10: 14, 11: 16, 12: 16}
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    first_data_row = header_bottom + 1
    row = first_data_row
    for stt, summary in enumerate(data.route_summary, start=1):
        ws.cell(row=row, column=1, value=stt).alignment = _CENTER
        ws.cell(row=row, column=2, value=summary.pickup_location).alignment = _LEFT
        ws.cell(row=row, column=3, value=summary.dropoff_location).alignment = _LEFT
        ws.cell(row=row, column=4, value="Chuyến").alignment = _CENTER
        ws.cell(row=row, column=5, value=summary.f20_count or None).alignment = _CENTER
        ws.cell(row=row, column=6, value=summary.f40_count or None).alignment = _CENTER
        ws.cell(row=row, column=7, value=summary.empty_count or None).alignment = _CENTER
        ws.cell(row=row, column=8, value=None).alignment = _CENTER
        ws.cell(row=row, column=9, value=f"=E{row}+F{row}+G{row}+H{row}").alignment = _CENTER
        unit_cell = ws.cell(row=row, column=10, value=f"=IFERROR(K{row}/I{row},0)")
        _money_fmt(unit_cell)
        amount_cell = ws.cell(row=row, column=11, value=summary.total_amount)
        _money_fmt(amount_cell)
        ws.cell(row=row, column=12, value="")
        row += 1

    last_data_row = row - 1 if data.route_summary else first_data_row - 1

    if data.route_summary:
        _apply_border(ws, header_top, last_data_row, 1, 12)
        total_row = last_data_row + 2
        ws.cell(row=total_row, column=1, value="TỔNG CƯỚC CHƯA VAT").font = _BOLD
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
        for col in (5, 6, 7, 8):
            cell = ws.cell(
                row=total_row,
                column=col,
                value=f"=SUM({get_column_letter(col)}{first_data_row}:{get_column_letter(col)}{last_data_row})",
            )
            cell.font = _BOLD
            cell.alignment = _CENTER
        ws.cell(row=total_row, column=9, value=f"=E{total_row}+F{total_row}+G{total_row}+H{total_row}").font = _BOLD
        ws.cell(row=total_row, column=9).alignment = _CENTER
        sum_cell = ws.cell(
            row=total_row, column=11,
            value=f"=SUM(K{first_data_row}:K{last_data_row})",
        )
        _money_fmt(sum_cell)
        sum_cell.font = _BOLD
        sum_cell.fill = _TOTAL_FILL

        vat_row = total_row + 1
        ws.cell(row=vat_row, column=1, value="THUẾ GTGT 8%").font = _BOLD
        ws.merge_cells(start_row=vat_row, start_column=1, end_row=vat_row, end_column=10)
        vat_cell = ws.cell(row=vat_row, column=11, value=f"=ROUND(K{total_row}*0.08,0)")
        _money_fmt(vat_cell)

        grand_row = vat_row + 1
        ws.cell(row=grand_row, column=1, value="TỔNG CƯỚC ĐÃ CÓ VAT").font = _BOLD
        ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=10)
        grand_cell = ws.cell(row=grand_row, column=11, value=f"=K{total_row}+K{vat_row}")
        _money_fmt(grand_cell)
        grand_cell.font = _BOLD
        grand_cell.fill = _TOTAL_FILL

        words_row = grand_row + 1
        ws.cell(row=words_row, column=2, value="Bằng chữ:").font = _BOLD
        ws.cell(
            row=words_row, column=3,
            value=number_to_vietnamese_words(data.total_with_vat),
        )
        ws.merge_cells(start_row=words_row, start_column=3, end_row=words_row, end_column=12)

        sign_row = words_row + 2
        date_text = f"Hải Phòng, ngày {p_end.strftime('%d')} tháng {p_end.strftime('%m')} năm {p_end.year}"
        c = ws.cell(row=sign_row, column=10, value=date_text)
        c.alignment = _CENTER
        ws.merge_cells(start_row=sign_row, start_column=10, end_row=sign_row, end_column=12)

        sign_label_row = sign_row + 1
        c1 = ws.cell(row=sign_label_row, column=2, value="XÁC NHẬN BÊN THUÊ VẬN CHUYỂN")
        c1.font = _BOLD
        c1.alignment = _CENTER
        ws.merge_cells(start_row=sign_label_row, start_column=2, end_row=sign_label_row, end_column=4)
        c2 = ws.cell(row=sign_label_row, column=10, value="XÁC NHẬN BÊN VẬN CHUYỂN")
        c2.font = _BOLD
        c2.alignment = _CENTER
        ws.merge_cells(start_row=sign_label_row, start_column=10, end_row=sign_label_row, end_column=12)

    ws.freeze_panes = ws.cell(row=header_bottom + 1, column=1)


def generate_pan_bk_sl_workbook(data: SettlementStatement) -> bytes:
    """Render the workbook to bytes. Sheets: BKTT first, then SL."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    mm = data.period.end.month
    yy = data.period.end.year % 100
    bktt_name = f"BKTT T{mm}.{yy:02d}"
    sl_name = f"SL T{mm}.{yy:02d}"

    _write_bktt_sheet(wb, data, bktt_name)
    _write_sl_sheet(wb, data, sl_name)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def settlement_filename(data: SettlementStatement) -> str:
    code = (data.client.code or data.client.name or "KH").strip().upper().replace(" ", "_")
    mm = data.period.end.month
    yy = data.period.end.year % 100
    return f"{code}_BK_SL_T{mm:02d}.{yy:02d}_HD.xlsx"
