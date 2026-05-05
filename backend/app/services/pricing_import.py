"""Customer-tariff (bảng giá) Excel import.

Wraps the per-format parsers from `scripts/seeds/seed_pricing_from_files.py`
behind a preview/commit API surface that mirrors the customer-Excel orders
import. Each tariff layout is structurally different — PAN's `Trucking (HD)`
sheet, HAP's `CUOC`, NEWWAY's settlement-style data — so we dispatch by
named format rather than running a column-mapping pipeline.

The preview returns parsed `TariffRow`s + per-route location resolutions
(same shape as orders import) so the accountant can review and override
before committing.

The commit upserts Pricing + PricingLine rows. Idempotent on
`(client_id, work_type, pickup_location_id, dropoff_location_id)` for the
header row, and on `(pricing_id, quantity)` for the line. Re-running with
the same file is a no-op; running with a changed unit price updates the
existing PricingLine.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Iterable, Sequence

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import Client, Pricing, PricingLine
from app.contexts.customer_pricing.application.location_resolver import (
    LocationResolverService,
    ResolverSource,
)


# Vietnamese-facing aggregator-row keywords we should skip.
_AGGREGATE_TOKENS = ("TỔNG", "TOTAL", "GHI CHU", "GHI CHÚ", "CỘNG", "GRAND TOTAL")

SUPPORTED_FORMATS = ("pan", "hap", "newway")


@dataclass
class TariffRow:
    pickup_raw: str
    dropoff_raw: str
    work_type: str          # F20 | F40 | E20 | E40
    unit_price: int
    quantity: int = 1
    driver_salary: int = 0
    allowance: int = 0
    note: str = ""

    def to_dict(self) -> dict:
        return {
            "pickup_location": self.pickup_raw,
            "dropoff_location": self.dropoff_raw,
            "work_type": self.work_type,
            "unit_price": self.unit_price,
            "quantity": self.quantity,
            "driver_salary": self.driver_salary,
            "allowance": self.allowance,
            "note": self.note,
        }


@dataclass
class PricingPreview:
    format: str
    sheet_name: str
    rows: list[TariffRow]
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "format": self.format,
            "sheet_name": self.sheet_name,
            "rows": [r.to_dict() for r in self.rows],
            "warnings": list(self.warnings),
            "stats": {
                "row_count": len(self.rows),
                "unique_routes": len(
                    {(r.pickup_raw, r.dropoff_raw) for r in self.rows}
                ),
            },
        }


# ── Layout detection ───────────────────────────────────────────────


def detect_format(filename: str) -> str | None:
    """Filename heuristic. Returns None if no match — caller should pick."""
    n = (filename or "").lower()
    if "pan" in n and ("bk" in n or "sl" in n):
        return "pan"
    if "hap" in n or "shipside" in n:
        return "hap"
    if "newway" in n or "hechun" in n:
        return "newway"
    return None


# ── Per-format parsers ─────────────────────────────────────────────


def _split_route(route: str) -> tuple[str, str]:
    for sep in (" – ", " — ", " - ", "–", "—", "-"):
        if sep in route:
            parts = [p.strip() for p in route.split(sep, 1)]
            if len(parts) == 2 and all(parts):
                return parts[0], parts[1]
    return route.strip(), ""


def _is_aggregate_row(value: str) -> bool:
    upper = value.upper()
    return any(tok in upper for tok in _AGGREGATE_TOKENS)


def parse_pan_bytes(content: bytes) -> PricingPreview:
    """PAN tariff. Sheet `Trucking (HD)`, route in col B starting row 7,
    cols AJ/AK/AL (36/37/38) carry vỏ / hàng-20 / hàng-40 prices."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if "Trucking (HD)" not in wb.sheetnames:
        wb.close()
        return PricingPreview(
            format="pan",
            sheet_name="",
            rows=[],
            warnings=["Không tìm thấy sheet 'Trucking (HD)' trong tệp PAN."],
        )
    ws = wb["Trucking (HD)"]
    rows: list[TariffRow] = []
    for r in range(7, ws.max_row + 1):
        route = ws.cell(r, 2).value
        if not route or not isinstance(route, str):
            continue
        route = route.strip()
        if _is_aggregate_row(route):
            continue
        empty_p = ws.cell(r, 36).value
        full20 = ws.cell(r, 37).value
        full40 = ws.cell(r, 38).value
        pickup, dropoff = _split_route(route)
        for price, wt in (
            (empty_p, "E40"), (empty_p, "E20"),
            (full20, "F20"), (full40, "F40"),
        ):
            if isinstance(price, (int, float)) and price > 0:
                rows.append(TariffRow(
                    pickup_raw=pickup,
                    dropoff_raw=dropoff,
                    work_type=wt,
                    unit_price=int(round(price)),
                    note=f"Trucking (HD) row {r}",
                ))
    wb.close()
    return PricingPreview(format="pan", sheet_name="Trucking (HD)", rows=rows)


def parse_hap_bytes(content: bytes) -> PricingPreview:
    """HAP tariff. Sheet `CUOC`, route col B from row 4, cols C/D/E/F."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if "CUOC" not in wb.sheetnames:
        wb.close()
        return PricingPreview(
            format="hap",
            sheet_name="",
            rows=[],
            warnings=["Không tìm thấy sheet 'CUOC' trong tệp HAP."],
        )
    ws = wb["CUOC"]
    rows: list[TariffRow] = []
    for r in range(4, ws.max_row + 1):
        route = ws.cell(r, 2).value
        if not route or not isinstance(route, str):
            continue
        route = route.strip()
        if _is_aggregate_row(route):
            continue
        full20 = ws.cell(r, 3).value
        full40 = ws.cell(r, 4).value
        empty20 = ws.cell(r, 5).value
        empty40 = ws.cell(r, 6).value
        pickup, dropoff = _split_route(route)
        for price, wt in (
            (full20, "F20"), (full40, "F40"),
            (empty20, "E20"), (empty40, "E40"),
        ):
            if isinstance(price, (int, float)) and price > 0:
                rows.append(TariffRow(
                    pickup_raw=pickup,
                    dropoff_raw=dropoff,
                    work_type=wt,
                    unit_price=int(round(price)),
                    note=f"CUOC row {r}",
                ))
    wb.close()
    return PricingPreview(format="hap", sheet_name="CUOC", rows=rows)


def parse_newway_bytes(content: bytes) -> PricingPreview:
    """NEWWAY: settlement-style data. Best-effort modal price per
    (route, work_type). Often produces 0 rows — flagged in warnings."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet_name = "Tháng 4" if "Tháng 4" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_seen: dict[tuple[str, str], dict[int, int]] = {}
    for r in range(8, ws.max_row + 1):
        route = ws.cell(r, 11).value
        if not route or not isinstance(route, str):
            continue
        route = route.strip()
        price = ws.cell(r, 12).value
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        empty20 = ws.cell(r, 5).value
        empty40 = ws.cell(r, 6).value
        full20 = ws.cell(r, 7).value
        full40 = ws.cell(r, 8).value
        wt = ("F40" if full40 else "F20" if full20
              else "E40" if empty40 else "E20" if empty20 else None)
        if wt is None:
            continue
        rows_seen.setdefault((route, wt), {})[int(round(price))] = (
            rows_seen.get((route, wt), {}).get(int(round(price)), 0) + 1
        )
    wb.close()
    rows: list[TariffRow] = []
    warnings: list[str] = []
    for (route, wt), counts in rows_seen.items():
        best_price = max(counts.items(), key=lambda kv: kv[1])[0]
        pickup, dropoff = _split_route(route)
        rows.append(TariffRow(
            pickup_raw=pickup,
            dropoff_raw=dropoff,
            work_type=wt,
            unit_price=best_price,
            note=f"settlement modal across {sum(counts.values())} trips",
        ))
    if not rows:
        warnings.append(
            "NEWWAY: không trích được dòng giá nào — định dạng có thể "
            "khác phiên bản tháng 4 hoặc thiếu cột giá."
        )
    return PricingPreview(
        format="newway",
        sheet_name=sheet_name,
        rows=rows,
        warnings=warnings,
    )


_PARSERS = {
    "pan": parse_pan_bytes,
    "hap": parse_hap_bytes,
    "newway": parse_newway_bytes,
}


def parse_tariff_bytes(content: bytes, fmt: str) -> PricingPreview:
    if fmt not in _PARSERS:
        raise ValueError(
            f"Định dạng không được hỗ trợ: {fmt!r}. "
            f"Các định dạng hợp lệ: {', '.join(SUPPORTED_FORMATS)}."
        )
    return _PARSERS[fmt](content)


# ── Commit ─────────────────────────────────────────────────────────


@dataclass
class CommitResult:
    pricings_created: int = 0
    pricings_existing: int = 0
    lines_created: int = 0
    lines_updated: int = 0
    lines_existing: int = 0
    skipped_no_locations: int = 0
    locations_created: int = 0


async def commit_tariff_rows(
    db: AsyncSession,
    *,
    client: Client,
    rows: Sequence[TariffRow],
    user_id: int | None,
    update_existing_lines: bool = False,
) -> CommitResult:
    """Upsert pricing + lines, idempotent on
    (client_id, work_type, pickup_location_id, dropoff_location_id) for the
    header and (pricing_id, quantity) for the line.

    If `update_existing_lines` is False (default), an existing line is left
    alone — common case for "import a fresh tariff but keep manually-tuned
    splits". When True, the line's `unit_price` is overwritten.
    """
    resolver = LocationResolverService(db)
    locations_before = await _location_count(db)
    result = CommitResult()

    for row in rows:
        if not row.pickup_raw or not row.dropoff_raw:
            result.skipped_no_locations += 1
            continue
        pickup_resp = await resolver.resolve_or_create(
            row.pickup_raw, source=ResolverSource.MANUAL, user_id=user_id,
        )
        dropoff_resp = await resolver.resolve_or_create(
            row.dropoff_raw, source=ResolverSource.MANUAL, user_id=user_id,
        )
        pickup_loc = pickup_resp.location
        dropoff_loc = dropoff_resp.location
        if pickup_loc is None or dropoff_loc is None:
            result.skipped_no_locations += 1
            continue

        existing_pricing = (await db.execute(
            select(Pricing).where(
                Pricing.client_id == client.id,
                Pricing.work_type == row.work_type,
                Pricing.pickup_location_id == pickup_loc.id,
                Pricing.dropoff_location_id == dropoff_loc.id,
            )
        )).scalar_one_or_none()
        if existing_pricing is None:
            pricing = Pricing(
                client_id=client.id,
                work_type=row.work_type,
                pickup_location_id=pickup_loc.id,
                dropoff_location_id=dropoff_loc.id,
                is_active=True,
            )
            db.add(pricing)
            await db.flush()
            result.pricings_created += 1
        else:
            pricing = existing_pricing
            result.pricings_existing += 1

        existing_line = (await db.execute(
            select(PricingLine).where(
                PricingLine.pricing_id == pricing.id,
                PricingLine.quantity == row.quantity,
            )
        )).scalar_one_or_none()
        if existing_line is None:
            db.add(PricingLine(
                pricing_id=pricing.id,
                quantity=row.quantity,
                unit_price=row.unit_price,
                driver_salary=row.driver_salary,
                allowance=row.allowance,
            ))
            result.lines_created += 1
        elif update_existing_lines and existing_line.unit_price != row.unit_price:
            existing_line.unit_price = row.unit_price
            existing_line.driver_salary = row.driver_salary
            existing_line.allowance = row.allowance
            result.lines_updated += 1
        else:
            result.lines_existing += 1

    locations_after = await _location_count(db)
    result.locations_created = max(0, locations_after - locations_before)
    await db.commit()
    return result


async def _location_count(db: AsyncSession) -> int:
    from sqlalchemy import func

    from app.models.domain import Location

    res = await db.execute(select(func.count()).select_from(Location))
    return int(res.scalar_one())


# ── Helpers for the API layer ──────────────────────────────────────


async def resolve_preview_locations(
    db: AsyncSession, rows: Iterable[TariffRow]
) -> dict[str, dict]:
    """Run the resolver in find-only mode for every unique pickup/dropoff
    string. Returns the same shape as the orders import preview so the
    frontend can render the (có sẵn) / (gợi ý) / (mới) badges identically.
    """
    resolver = LocationResolverService(db)
    seen: set[str] = set()
    for row in rows:
        for raw in (row.pickup_raw, row.dropoff_raw):
            if raw and raw.strip():
                seen.add(raw.strip())

    out: dict[str, dict] = {}
    for raw in seen:
        result = await resolver.find_match(raw)
        out[raw] = {
            "raw": raw,
            "match_kind": result.match_kind.value,
            "location_id": result.location.id if result.location else None,
            "location_name": result.location.name if result.location else None,
            "review_needed": result.review_needed,
            "suggestions": [
                {"location_id": s.location_id, "name": s.name, "score": s.score}
                for s in result.suggestions
            ],
        }
    return out
