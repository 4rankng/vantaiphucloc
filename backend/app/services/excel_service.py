"""Excel file processing for customer reconciliation uploads and trip order imports."""

import logging
import re
from datetime import date as date_type, datetime as dt
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import WorkOrder, WorkOrderContainer, TripOrder, TripOrderContainer, Client, PricingLine
from app.utils.iso6346 import normalize_container_number

_logger = logging.getLogger(__name__)


class ReconciliationResult:
    """Result of comparing customer Excel data with system records."""

    def __init__(
        self,
        container_number: str,
        normalized_number: str,
        work_order_id: int | None = None,
        trip_order_id: int | None = None,
        status: str = "pending",
        match_type: str = "none",
    ):
        self.container_number = container_number
        self.normalized_number = normalized_number
        self.work_order_id = work_order_id
        self.trip_order_id = trip_order_id
        self.status = status  # "confirmed" | "pending" | "rejected"
        self.is_duplicate = False
        self.match_type = match_type  # "exact" | "partial" | "none"

    def to_dict(self) -> dict[str, Any]:
        return {
            "container_number": self.container_number,
            "normalized_number": self.normalized_number,
            "work_order_id": self.work_order_id,
            "trip_order_id": self.trip_order_id,
            "status": self.status,
            "is_duplicate": self.is_duplicate,
            "match_type": self.match_type,
        }


async def parse_customer_excel(
    file_content: bytes,
    client_id: int | None = None,
) -> list[dict[str, Any]]:
    """
    Parse Excel file uploaded by customer.

    Expected columns (to be confirmed with customer):
    - Container Number (required)
    - Date (optional)
    - Pickup Location (optional)
    - Dropoff Location (optional)

    Returns list of dicts with parsed data.
    """
    try:
        import openpyxl
    except ImportError:
        _logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise ImportError("openpyxl is required for Excel processing")

    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active

    # Get header row
    headers = [cell.value for cell in sheet[1]]
    _logger.info(f"Excel headers: {headers}")

    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Skip empty rows
            continue

        row_dict = dict(zip(headers, row))
        results.append(row_dict)

    workbook.close()
    _logger.info(f"Parsed {len(results)} rows from Excel file")
    return results


async def compare_with_system_records(
    db: AsyncSession,
    client_id: int,
    excel_data: list[dict[str, Any]],
    date_from: str | None = None,
    date_to: str | None = None,
) -> list[ReconciliationResult]:
    """
    Compare customer Excel data with system work orders and trip orders.

    Returns list of ReconciliationResult objects with duplicate detection.
    """
    results: list[ReconciliationResult] = []

    # Extract and normalize container numbers from Excel
    excel_containers = []
    for row in excel_data:
        container_number = row.get("Container Number") or row.get("Số Container") or row.get("container_number")
        if container_number:
            normalized = normalize_container_number(str(container_number))
            excel_containers.append({
                "original": container_number,
                "normalized": normalized,
                "row_data": row,
            })

    # Query work orders for this client
    wo_query = select(WorkOrder).where(WorkOrder.client_id == client_id)
    if date_from:
        wo_query = wo_query.where(WorkOrder.created_at >= date_from)
    if date_to:
        wo_query = wo_query.where(WorkOrder.created_at <= date_to)

    wo_result = await db.execute(wo_query)
    work_orders = wo_result.scalars().all()

    # Query trip orders for this client
    to_query = select(TripOrder).where(TripOrder.client_id == client_id)
    if date_from:
        to_query = to_query.where(TripOrder.trip_date >= date_from)
    if date_to:
        to_query = to_query.where(TripOrder.trip_date <= date_to)

    to_result = await db.execute(to_query)
    trip_orders = to_result.scalars().all()

    # Load containers for all work orders
    wo_ids = [wo.id for wo in work_orders]
    if wo_ids:
        wo_cont_result = await db.execute(
            select(WorkOrderContainer).where(
                WorkOrderContainer.work_order_id.in_(wo_ids)
            )
        )
        wo_containers = wo_cont_result.scalars().all()

        # Build lookup: normalized container number -> list of work order IDs
        wo_container_map: dict[str, list[int]] = {}
        for wc in wo_containers:
            normalized = normalize_container_number(wc.container_number)
            if normalized not in wo_container_map:
                wo_container_map[normalized] = []
            wo_container_map[normalized].append(wc.work_order_id)

    # Load containers for all trip orders
    to_ids = [to.id for to in trip_orders]
    if to_ids:
        to_cont_result = await db.execute(
            select(TripOrderContainer).where(
                TripOrderContainer.trip_order_id.in_(to_ids)
            )
        )
        to_containers = to_cont_result.scalars().all()

        # Build lookup: normalized container number -> list of trip order IDs
        to_container_map: dict[str, list[int]] = {}
        for tc in to_containers:
            normalized = normalize_container_number(tc.container_number)
            if normalized not in to_container_map:
                to_container_map[normalized] = []
            to_container_map[normalized].append(tc.trip_order_id)

    # Compare Excel containers with system records
    # Build digits-only lookup for partial matching
    def _digits_only(s: str) -> str:
        return re.sub(r'[^0-9]', '', s)

    wo_digits_map: dict[str, list[int]] = {}
    for cn, ids in wo_container_map.items():
        digits = _digits_only(cn)
        if digits:
            wo_digits_map.setdefault(digits, []).extend(ids)

    to_digits_map: dict[str, list[int]] = {}
    for cn, ids in to_container_map.items():
        digits = _digits_only(cn)
        if digits:
            to_digits_map.setdefault(digits, []).extend(ids)

    for excel_cont in excel_containers:
        normalized = excel_cont["normalized"]

        # Check for exact matches
        wo_ids = wo_container_map.get(normalized, [])
        to_ids = to_container_map.get(normalized, [])

        match_type = "none"

        if wo_ids or to_ids:
            match_type = "exact"
        else:
            # Try digits-only partial match
            digits = _digits_only(normalized)
            if digits:
                wo_ids = list(set(wo_digits_map.get(digits, [])))
                to_ids = list(set(to_digits_map.get(digits, [])))
                if wo_ids or to_ids:
                    match_type = "partial"

        result = ReconciliationResult(
            container_number=excel_cont["original"],
            normalized_number=normalized,
            status="pending",
            match_type=match_type,
        )

        if wo_ids or to_ids:
            result.is_duplicate = True
            result.work_order_id = wo_ids[0] if wo_ids else None
            result.trip_order_id = to_ids[0] if to_ids else None
            if wo_ids and to_ids:
                result.status = "confirmed"

        results.append(result)

    _logger.info(
        f"Comparison complete: {len(results)} containers, "
        f"{sum(1 for r in results if r.is_duplicate)} duplicates found"
    )

    return results


async def generate_reconciliation_excel(
    db: AsyncSession,
    client_id: int,
    date_from: str | None = None,
    date_to: str | None = None,
) -> bytes:
    """
    Generate Excel file with reconciliation data for export.

    Returns Excel file content as bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        _logger.error("openpyxl not installed")
        raise ImportError("openpyxl is required for Excel generation")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đối soát"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    headers = [
        "Số Container",
        "Số cont chuẩn hóa",
        "Mã WO",
        "Mã TO",
        "Trạng thái",
        "Đã chốt",
        "Ngày chạy",
    ]
    ws.append(headers)

    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Query data
    wo_query = select(WorkOrder).where(WorkOrder.client_id == client_id)
    if date_from:
        wo_query = wo_query.where(WorkOrder.created_at >= date_from)
    if date_to:
        wo_query = wo_query.where(WorkOrder.created_at <= date_to)

    wo_result = await db.execute(wo_query)
    work_orders = wo_result.scalars().all()

    wo_ids = [wo.id for wo in work_orders]
    if wo_ids:
        wo_cont_result = await db.execute(
            select(WorkOrderContainer).where(
                WorkOrderContainer.work_order_id.in_(wo_ids)
            )
        )
        wo_containers = wo_cont_result.scalars().all()

        # Group containers by work order
        wo_cont_map: dict[int, list[WorkOrderContainer]] = {}
        for wc in wo_containers:
            if wc.work_order_id not in wo_cont_map:
                wo_cont_map[wc.work_order_id] = []
            wo_cont_map[wc.work_order_id].append(wc)

        # Check for matched trip orders
        to_result = await db.execute(
            select(TripOrder).where(TripOrder.client_id == client_id)
        )
        trip_orders = to_result.scalars().all()

        to_map: dict[int, TripOrder] = {to.id: to for to in trip_orders}

        # Query join table for matches
        from app.models.domain import TripOrderWorkOrder
        join_result = await db.execute(
            select(TripOrderWorkOrder).where(
                TripOrderWorkOrder.work_order_id.in_(wo_ids)
            )
        )
        joins = join_result.scalars().all()

        wo_to_map: dict[int, int] = {j.work_order_id: j.trip_order_id for j in joins}

        # Add data rows
        row_num = 2
        for wo in work_orders:
            containers = wo_cont_map.get(wo.id, [])
            for wc in containers:
                to_id = wo_to_map.get(wo.id)
                to = to_map.get(to_id) if to_id else None

                status = "Chưa khớp"
                if to:
                    status = "Đã khớp"

                ws.append([
                    wc.container_number,
                    normalize_container_number(wc.container_number),
                    f"WO#{wo.id}",
                    f"TO#{to_id}" if to_id else "",
                    status,
                    "✓" if to and to.is_confirmed else "",
                    wo.created_at.date() if wo.created_at else "",
                ])

                # Highlight duplicate containers
                if to:
                    cell = ws.cell(row=row_num, column=1)
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                row_num += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    content = excel_buffer.getvalue()
    wb.close()

    _logger.info(f"Generated Excel file with {row_num - 1} data rows")
    return content


# ---------------------------------------------------------------------------
# Trip Order Excel Import
# ---------------------------------------------------------------------------

# Vietnamese/English column name mappings
_TRIP_IMPORT_COLUMNS = {
    "ngay": "trip_date",
    "date": "trip_date",
    "ngay_chay": "trip_date",
    "trip_date": "trip_date",
    "ma_kh": "client_code",
    "client_code": "client_code",
    "khach_hang": "client_code",
    "ma_khach_hang": "client_code",
    "diem_lay": "pickup_location",
    "pickup": "pickup_location",
    "pickup_location": "pickup_location",
    "diem_tra": "dropoff_location",
    "dropoff": "dropoff_location",
    "dropoff_location": "dropoff_location",
    "cung_duong": "route",
    "route": "route",
    "so_cont": "container_number",
    "container": "container_number",
    "container_number": "container_number",
    "loai": "work_type",
    "work_type": "work_type",
    "loai_cont": "work_type",
    "don_gia": "unit_price",
    "unit_price": "unit_price",
    "luong_tx": "driver_salary",
    "driver_salary": "driver_salary",
    "phu_cap": "allowance",
    "allowance": "allowance",
}


def _normalize_trip_import_header(header: str) -> str:
    """Map Vietnamese/English header to standard field name."""
    if not header:
        return ""
    normalized = re.sub(r'[^a-zA-Z0-9_]', '_', str(header).strip().lower())
    return _TRIP_IMPORT_COLUMNS.get(normalized, normalized)


async def parse_trip_order_excel(file_content: bytes) -> list[dict[str, Any]]:
    """Parse Excel file for batch trip order import.

    Expected columns (Vietnamese or English):
    - trip_date (Ngày/Ngày chạy)
    - client_code (Mã KH/Mã khách hàng)
    - pickup_location (Điểm lấy) + dropoff_location (Điểm trả) or route (Cung đường)
    - container_number (Số cont/Container)
    - work_type (Loại/Loại cont) — E20/E40/F20/F40
    - unit_price (Đơn giá) — optional
    - driver_salary (Lương TX) — optional
    - allowance (Phụ cấp) — optional
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel processing")

    workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    sheet = workbook.active

    raw_headers = [cell.value for cell in sheet[1]]
    headers = [_normalize_trip_import_header(h) for h in raw_headers]

    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        row_dict = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                row_dict[headers[col_idx]] = value
        results.append(row_dict)

    workbook.close()
    _logger.info(f"Parsed {len(results)} trip order rows from Excel")
    return results


async def import_trip_orders(
    db: AsyncSession,
    rows: list[dict[str, Any]],
    user_id: int,
) -> dict[str, Any]:
    """Validate and create trip orders from parsed Excel rows.

    Groups rows by (trip_date, client_code) to form trip orders.
    Returns { created: int, errors: list[str] }.
    """
    from app.services.pricing_service import find_tiered_pricing

    created = 0
    errors: list[str] = []

    # Group rows by trip key
    groups: dict[tuple, list[dict]] = {}
    for i, row in enumerate(rows):
        trip_date = row.get("trip_date")
        client_code = row.get("client_code")

        if not trip_date or not client_code:
            errors.append(f"Dòng {i + 2}: thiếu ngày hoặc mã khách hàng")
            continue

        if isinstance(trip_date, str):
            try:
                trip_date = dt.strptime(trip_date, "%Y-%m-%d").date()
            except ValueError:
                try:
                    trip_date = dt.strptime(trip_date, "%d/%m/%Y").date()
                except ValueError:
                    errors.append(f"Dòng {i + 2}: định dạng ngày không hợp lệ '{trip_date}'")
                    continue
        elif hasattr(trip_date, "date"):
            pass  # already a date

        key = (str(trip_date), str(client_code))
        groups.setdefault(key, []).append(row)

    for key, group_rows in groups.items():
        trip_date, client_code_str = key

        # Look up client by code
        client_result = await db.execute(
            select(Client).where(Client.code == client_code_str)
        )
        client = client_result.scalar_one_or_none()
        if not client:
            errors.append(f"Nhóm {key}: không tìm thấy khách hàng mã '{client_code_str}'")
            continue

        first_row = group_rows[0]
        pickup = first_row.get("pickup_location")
        dropoff = first_row.get("dropoff_location")
        route = first_row.get("route") or (f"{pickup} - {dropoff}" if pickup and dropoff else "")

        # Build containers (normalize numbers — strip hyphens, uppercase)
        containers_data = []
        for row in group_rows:
            cn = normalize_container_number(str(row.get("container_number", "")).strip())
            wt = str(row.get("work_type", "")).strip().upper()
            if cn:
                containers_data.append({"container_number": cn, "work_type": wt or "E20"})

        if not containers_data:
            errors.append(f"Nhóm {key}: không có số container")
            continue

        work_type = containers_data[0]["work_type"]
        container_count = sum(1 for c in containers_data if c["work_type"] == work_type) or 1

        # Try auto-pricing
        unit_price = int(first_row.get("unit_price") or 0)
        driver_salary = int(first_row.get("driver_salary") or 0)
        allowance = int(first_row.get("allowance") or 0)
        pricing_id = None

        if not unit_price:
            tiered = await find_tiered_pricing(
                db, client_id=client.id, work_type=work_type,
                quantity=container_count, route=route,
                pickup_location=pickup, dropoff_location=dropoff,
            )
            if tiered:
                unit_price = tiered.unit_price
                driver_salary = tiered.driver_salary
                allowance = tiered.allowance
                pricing_id = tiered.pricing.id

        trip_date_val = date_type.fromisoformat(str(trip_date)) if isinstance(trip_date, str) else trip_date

        trip_order = TripOrder(
            trip_date=trip_date_val,
            client_id=client.id,
            client_name=client.name,
            work_type=work_type,
            route=route,
            pickup_location=pickup,
            dropoff_location=dropoff,
            container_number=containers_data[0]["container_number"],
            pricing_id=pricing_id,
            unit_price=unit_price,
            driver_salary=driver_salary,
            allowance=allowance,
            revenue=unit_price,
            status="PENDING" if unit_price > 0 else "DRAFT",
        )
        db.add(trip_order)
        await db.flush()

        for c in containers_data:
            db.add(TripOrderContainer(
                trip_order_id=trip_order.id,
                container_number=c["container_number"],
                work_type=c["work_type"],
            ))

        created += 1

    await db.commit()
    return {"created": created, "errors": errors}


def generate_trip_order_template() -> bytes:
    """Generate a blank Excel template for trip order import."""
    headers = [
        "Ngày", "Mã KH", "Điểm lấy", "Điểm trả",
        "Cung đường", "Số cont", "Loại cont", "Đơn giá",
        "Lương TX", "Phụ cấp",
    ]
    examples = [
        "01/05/2026", "KH001", "Cát lái", "KC Bình Dương",
        "Cát lái - KC Bình Dương", "TCLU1234567", "E20", "1500000",
        "500000", "100000",
    ]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Nhập chuyến"
    sheet.append(headers)
    sheet.append(examples)

    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)

    buf = io.BytesIO()
    workbook.save(buf)
    workbook.close()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel Export for Work Orders, Trip Orders, Salary
# ---------------------------------------------------------------------------

async def generate_work_orders_excel(
    db: AsyncSession,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = None,
) -> bytes:
    """Export work orders to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = select(WorkOrder).order_by(WorkOrder.id.desc())
    if date_from:
        query = query.where(WorkOrder.created_at >= date_from)
    if date_to:
        query = query.where(WorkOrder.created_at <= date_to)
    if status:
        query = query.where(WorkOrder.status == status)

    result = await db.execute(query)
    work_orders = result.scalars().all()

    wo_ids = [wo.id for wo in work_orders]
    containers_map: dict[int, list[WorkOrderContainer]] = {}
    if wo_ids:
        cont_result = await db.execute(
            select(WorkOrderContainer).where(WorkOrderContainer.work_order_id.in_(wo_ids))
        )
        for c in cont_result.scalars().all():
            containers_map.setdefault(c.work_order_id, []).append(c)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phiếu làm việc"

    headers = ["Mã WO", "Khách hàng", "Điểm lấy", "Điểm trả", "Tài xế", "Biển số", "Số cont", "Loại", "Lương TX", "Phụ cấp", "Thu nhập", "Trạng thái", "Ngày tạo"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_labels = {"PENDING": "Chờ đối soát", "MATCHED": "Đã đối soát", "COMPLETED": "Hoàn thành"}
    for wo in work_orders:
        containers = containers_map.get(wo.id, [])
        for c in containers:
            ws.append([
                f"WO#{wo.id}", wo.client_name,
                wo.pickup_location or "", wo.dropoff_location or "",
                wo.driver_name, wo.tractor_plate,
                c.container_number, c.work_type,
                wo.driver_salary, wo.allowance, wo.earning,
                status_labels.get(wo.status, wo.status),
                wo.created_at.date() if wo.created_at else "",
            ])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


async def generate_trip_orders_excel(
    db: AsyncSession,
    date_from: str | None = None,
    date_to: str | None = None,
    status: str | None = None,
) -> bytes:
    """Export trip orders to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = select(TripOrder).order_by(TripOrder.id.desc())
    if date_from:
        query = query.where(TripOrder.trip_date >= date_from)
    if date_to:
        query = query.where(TripOrder.trip_date <= date_to)
    if status:
        query = query.where(TripOrder.status == status)

    result = await db.execute(query)
    trip_orders = result.scalars().all()

    to_ids = [to.id for to in trip_orders]
    containers_map: dict[int, list[TripOrderContainer]] = {}
    if to_ids:
        cont_result = await db.execute(
            select(TripOrderContainer).where(TripOrderContainer.trip_order_id.in_(to_ids))
        )
        for c in cont_result.scalars().all():
            containers_map.setdefault(c.trip_order_id, []).append(c)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đơn hàng"

    headers = ["Mã TO", "Ngày chạy", "Khách hàng", "Điểm lấy", "Điểm trả", "Số cont", "Loại", "Đơn giá", "Lương TX", "Phụ cấp", "Doanh thu", "Trạng thái", "Đã chốt"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_labels = {"DRAFT": "Nháp", "PENDING": "Chờ đối soát", "COMPLETED": "Hoàn thành", "CANCELLED": "Đã huỷ"}
    for to in trip_orders:
        containers = containers_map.get(to.id, [])
        for c in containers:
            ws.append([
                f"TO#{to.id}", to.trip_date,
                to.client_name,
                to.pickup_location or "", to.dropoff_location or "",
                c.container_number, c.work_type,
                to.unit_price, to.driver_salary, to.allowance, to.revenue,
                status_labels.get(to.status, to.status),
                "✓" if to.is_confirmed else "",
            ])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


async def generate_salary_excel(
    db: AsyncSession,
    start_date: str,
    end_date: str,
) -> bytes:
    """Export salary breakdown per driver to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models.domain import SalaryPeriod

    result = await db.execute(
        select(SalaryPeriod).where(
            SalaryPeriod.start_date == start_date,
            SalaryPeriod.end_date == end_date,
        ).order_by(SalaryPeriod.driver_name)
    )
    periods = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng lương"

    headers = ["Tài xế", "Kỳ lương", "Số chuyến", "Giá/chuyến", "Tổng lương", "Phụ cấp", "Khấu trừ", "Thực nhận", "Trạng thái"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_labels = {"OPEN": "Chờ tính", "CALCULATED": "Đã tính", "PAID": "Đã trả"}
    for p in periods:
        ws.append([
            p.driver_name,
            f"{p.start_date} → {p.end_date}",
            p.work_order_count,
            p.price_per_order,
            p.total_salary,
            p.total_allowance,
            p.total_deduction,
            p.net_pay,
            status_labels.get(p.status, p.status),
        ])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()
