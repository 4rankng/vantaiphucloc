"""Integration tests for GET /booked-trips/export-doi-soat endpoint.

The đối soát (reconciliation) export now unions BookedTrip (containers the
customer listed in their ship file) with DeliveredTrip (containers drivers
actually moved) and tags each row with TRẠNG THÁI = "Đã ghép" / "Chưa ghép".
"""

import io
import random
import string
from uuid import uuid4

from openpyxl import load_workbook

_ISO_LETTER_MAP = {
    "A": 10,
    "B": 12,
    "C": 13,
    "D": 14,
    "E": 15,
    "F": 16,
    "G": 17,
    "H": 18,
    "I": 19,
    "J": 20,
    "K": 21,
    "L": 23,
    "M": 24,
    "N": 25,
    "O": 26,
    "P": 27,
    "Q": 28,
    "R": 29,
    "S": 30,
    "T": 31,
    "U": 32,
    "V": 34,
    "W": 35,
    "X": 36,
    "Y": 37,
    "Z": 38,
}
_ISO_POWERS = [2**i for i in range(10)]


def _uid() -> str:
    return uuid4().hex[:8]


def _container_number() -> str:
    prefix = "".join(random.choices(string.ascii_uppercase, k=4))
    serial = "".join(random.choices(string.digits, k=6))
    base = prefix + serial
    total = 0
    for i, ch in enumerate(base):
        value = _ISO_LETTER_MAP[ch] if ch.isalpha() else int(ch)
        total += value * _ISO_POWERS[i]
    check = total % 11
    if check == 10:
        check = 0
    return f"{base}{check}"


def _confirm_match(api_client, admin_headers, pairs):
    """Call the current match endpoint and return matched_count."""
    resp = api_client.post(
        "/auto-match/confirm",
        headers=admin_headers,
        json={"pairs": pairs},
    )
    assert resp.status_code == 200, f"Match failed: {resp.text}"
    return resp.json().get("matched_count", 0)


class TestDoiSoatExport:
    """Tests for the đối soát (reconciliation) Excel export endpoint."""

    EXPECTED_HEADERS = [
        "STT",
        "NGÀY ĐI",
        "CHỦ HÀNG",
        "SỐ CONTAINER",
        "F20'",
        "F40'",
        "E20'",
        "E40'",
        "SỐ XE CHẠY",
        "ĐIỂM ĐI",
        "ĐIỂM ĐẾN",
        "TÁC NGHIỆP",
        "CƯỚC",
        "LƯƠNG",
        "TRẠNG THÁI",
        "GHI CHÚ",
    ]

    def test_export_unions_matched_and_unmatched(
        self,
        api_client,
        admin_headers,
        create_partner,
        create_location,
        create_work_order,
        create_trip_order,
    ):
        """TC1: 3 BookedTrip + 2 matched DeliveredTrip → 3 rows (2 matched + 1 unmatched)."""
        uid = _uid()
        partner = create_partner(name=f"CTY DS {uid}", code=f"DS{uid.upper()}")

        pickup = create_location(name=f"Kho DS {uid}A")
        dropoff = create_location(name=f"Cảng DS {uid}B")

        cn1 = _container_number()
        cn2 = _container_number()
        cn3 = _container_number()

        bt1 = create_trip_order(
            client_id=partner["id"],
            pickup_location_id=pickup["id"],
            dropoff_location_id=dropoff["id"],
            trip_date="2026-05-10",
            cont_number=cn1,
            cont_type="E20",
            work_type="E20",
        )
        bt2 = create_trip_order(
            client_id=partner["id"],
            pickup_location_id=pickup["id"],
            dropoff_location_id=dropoff["id"],
            trip_date="2026-05-10",
            cont_number=cn2,
            cont_type="E40",
            work_type="E40",
        )
        create_trip_order(
            client_id=partner["id"],
            pickup_location_id=pickup["id"],
            dropoff_location_id=dropoff["id"],
            trip_date="2026-05-10",
            cont_number=cn3,
            cont_type="E20",
            work_type="E20",
        )

        wo1 = create_work_order(
            client_id=partner["id"],
            pickup_location_id=pickup["id"],
            dropoff_location_id=dropoff["id"],
            cont_number=cn1,
            cont_type="E20",
            work_type="E20",
            trip_date="2026-05-10",
            note="Tài xế ghi chú khi nộp chuyến",
        )
        wo2 = create_work_order(
            client_id=partner["id"],
            pickup_location_id=pickup["id"],
            dropoff_location_id=dropoff["id"],
            cont_number=cn2,
            cont_type="E40",
            work_type="E40",
            trip_date="2026-05-10",
        )

        matched_count = _confirm_match(
            api_client,
            admin_headers,
            [
                {"delivered_trip_id": wo1["id"], "booked_trip_id": bt1["id"]},
                {"delivered_trip_id": wo2["id"], "booked_trip_id": bt2["id"]},
            ],
        )
        assert matched_count == 2

        resp = api_client.get(
            "/booked-trips/export-doi-soat",
            params={
                "client_id": partner["id"],
                "date_from": "2026-05-01",
                "date_to": "2026-05-31",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200, f"Export failed: {resp.text}"
        assert "doisoat" in resp.headers.get("content-disposition", "").lower()

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active

        headers = [c.value for c in ws[10]]
        assert headers == self.EXPECTED_HEADERS

        data_rows = list(ws.iter_rows(min_row=12, values_only=True))
        assert len(data_rows) == 3, (
            f"Expected 3 rows, got {len(data_rows)}: {data_rows}"
        )

        status_by_cont = {row[3]: row[14] for row in data_rows}
        assert status_by_cont[cn1] == "Đã ghép"
        assert status_by_cont[cn2] == "Đã ghép"
        assert status_by_cont[cn3] == "Chưa ghép"
        note_by_cont = {row[3]: row[15] for row in data_rows}
        assert note_by_cont[cn1] == "Tài xế ghi chú khi nộp chuyến"

        assert [row[0] for row in data_rows] == [1, 2, 3]
        for row in data_rows:
            assert row[1] == "10/05/2026"

        wb.close()

    def test_export_includes_unmatched_delivered_trip(
        self,
        api_client,
        admin_headers,
        create_partner,
        create_location,
        create_work_order,
        create_trip_order,
    ):
        """TC4: 1 DeliveredTrip with no BookedTrip → 1 row 'Chưa ghép'."""
        uid = _uid()
        partner = create_partner(name=f"CTY Extra {uid}", code=f"EX{uid.upper()}")

        pickup = create_location(name=f"Kho Extra {uid}")
        dropoff = create_location(name=f"Cảng Extra {uid}")

        cn_unmatched = _container_number()
        create_work_order(
            client_id=partner["id"],
            pickup_location_id=pickup["id"],
            dropoff_location_id=dropoff["id"],
            cont_number=cn_unmatched,
            cont_type="E20",
            work_type="E20",
            trip_date="2026-05-15",
        )

        resp = api_client.get(
            "/booked-trips/export-doi-soat",
            params={
                "client_id": partner["id"],
                "date_from": "2026-05-01",
                "date_to": "2026-05-31",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=12, values_only=True))
        assert len(data_rows) == 1
        assert data_rows[0][3] == cn_unmatched
        assert data_rows[0][14] == "Chưa ghép"
        wb.close()

    def test_export_empty_date_range(
        self,
        api_client,
        admin_headers,
        create_partner,
        create_location,
        create_work_order,
        create_trip_order,
    ):
        """TC2: Date range with no trips returns Excel with header row only."""
        uid = _uid()
        partner = create_partner(name=f"CTY Empty {uid}", code=f"EM{uid.upper()}")
        pickup = create_location(name=f"Kho Empty {uid}")
        dropoff = create_location(name=f"Cảng Empty {uid}")

        create_trip_order(
            client_id=partner["id"],
            pickup_location_id=pickup["id"],
            dropoff_location_id=dropoff["id"],
            trip_date="2026-04-10",
            cont_number=_container_number(),
            cont_type="E20",
            work_type="E20",
        )

        resp = api_client.get(
            "/booked-trips/export-doi-soat",
            params={
                "client_id": partner["id"],
                "date_from": "2026-01-01",
                "date_to": "2026-01-31",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=12, values_only=True))
        assert len(data_rows) == 0
        wb.close()

    def test_export_nonexistent_client(self, api_client, admin_headers):
        """TC3: Non-existent client_id returns valid Excel with no data rows."""
        resp = api_client.get(
            "/booked-trips/export-doi-soat",
            params={
                "client_id": 99999,
                "date_from": "2026-05-01",
                "date_to": "2026-05-31",
            },
            headers=admin_headers,
        )
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=12, values_only=True))
        assert len(data_rows) == 0
        wb.close()
